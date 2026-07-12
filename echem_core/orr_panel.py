"""ORR Analysis panel — background-subtracted rotating disk electrode polarization curves.

Each SAMPLE groups N2 (background) and O2 (signal) CV file pairs by RPM.

Processing per pair:
  1. Extract the last cycle from each file.
  2. Apply IR correction:  E = Ewe/V − (I/mA / 1000) × R_sol
     (separate R_sol_N2 and R_sol_O2 — measured in independent sessions).
  3. Apply RHE conversion: E += E_ref  (shared per sample).
  4. Extract the anodic scan direction (from minimum-E vertex upward).
  5. Interpolate N2 onto O2's E grid and subtract: I_net = I_O2 − I_N2_interp.
  6. Optionally normalise to electrode area → J (mA/cm²).

Multiple samples are displayed in a scrollable grid of subplots (configurable columns),
mirroring the layout of the Multi E.Chem 2 tab.
"""

from collections import OrderedDict
import re
import os
import math
import colorsys

import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import pandas as pd
import matplotlib.colors as _mcolors
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

from .file_manager import _read_mpr, _PALETTE, _COLOR_NAMES, _COLOR_HEX
from .plotting import (apply_grid, draw_reflines, copy_figure_to_clipboard,
                        _cycle_colors, _scale_legend_spacing,
                        attach_plot_assistant)
from .checklist import CheckableListbox
from . import session_manager as _sm
from .legend_editor import open_legend_editor

# ── UI constants ────────────────────────────────────────────────────────
_SAMPLE_HDR_BG     = "#d1c4e9"   # light purple — distinct from ME1 blue / ME2 green
_SAMPLE_HDR_ACTIVE = "#ffd54f"   # gold  (matches other tabs)

# Extracts sample name from between the CV-type prefix and the "vs R…"
# reference in the filename. Handles both the old "vs RE3" form and the newer
# "vs RHEa" form, and strips a trailing "_RE<x>" working-/reference-electrode
# label when present.
# e.g. "P6_CVn2_Pt_disk vs RE3…"              → "Pt_disk"
#      "P8_CVo2_LTS-BDRDE_34(Pt) vs REa…"     → "LTS-BDRDE_34(Pt)"
#      "P6_CVn2_Pt_disk_M5_REa vs RHEa…"      → "Pt_disk_M5"
#      "P8_CVo2_LTS-BDRDE_40(Pt)_REa vs RHEa…"→ "LTS-BDRDE_40(Pt)"
_SAMPLE_NAME_PAT = re.compile(
    r'_CV[a-zA-Z0-9]+_(.+?)(?:_RE\w*)?\s+vs\s+R', re.IGNORECASE)

# Runtime-only keys stripped before JSON serialisation
_SAMPLE_RUNTIME = frozenset({
    "fig", "ax", "canvas", "toolbar", "plot_frame",
    "hdr_frame", "hdr_label", "legend",
    # ephemeral interaction state
    "panning", "pan_ax", "pan_start", "pan_moved",
    "leg_resize", "leg_resize_start_y", "leg_resize_start_sz",
    "ann", "ann_dot", "ann_last", "ann_idx", "_plot_data",
    "auto_xlim", "auto_ylim", "leg_size_live",
})
_PAIR_RUNTIME = frozenset({"df_n2", "df_o2"})

# Matches  _NN_CV_   or   _NNN_CV_   in a filename
_RPM_PAT = re.compile(r'_(\d{2,4})_CV_', re.IGNORECASE)

_GRID_STYLE_MAP = {
    "dashed": "--", "dotted": ":", "solid": "-", "dash-dot": "-."}

_ANN_DOT_LABEL = "_orr_dot"

_RPM_DEFAULTS = [400, 900, 1600, 2500]

# Electrolyte parameters for Levich / limiting-current theoretical lines
# (n, D_O2 cm²/s, ν cm²/s, C_O2 mol/cm³)
_ELECTROLYTES = {
    "0.1 M HClO₄":  (4, 1.93e-5, 1.005e-2, 1.26e-6),
    "0.1 M KOH":    (4, 1.90e-5, 1.012e-2, 1.21e-6),
    "0.5 M H₂SO₄":  (4, 1.40e-5, 1.000e-2, 1.10e-6),
    "1 M KOH":      (4, 1.90e-5, 1.012e-2, 1.21e-6),
    "1 M HClO₄":    (4, 1.93e-5, 1.000e-2, 1.21e-6),
}


def _safe_rpm_int(s) -> int:
    try: return int(s)
    except (ValueError, TypeError): return 0


# ── Module-level helpers ─────────────────────────────────────────────────

def _read_one_df(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".mpr":
        return _read_mpr(path)
    df = pd.read_csv(path, sep="\t", encoding="latin-1", on_bad_lines="skip")
    df.columns = [c.strip().replace("<", "").replace(">", "") for c in df.columns]
    return df.reset_index(drop=True)


def _detect_gas(stem: str) -> str:
    lo = stem.lower()
    # Match n2 / o2 as word-like fragments (not inside longer numbers)
    if re.search(r'(?<!\d)n2(?!\d)', lo):
        return "n2"
    if re.search(r'(?<!\d)o2(?!\d)', lo):
        return "o2"
    return ""


def _extract_rpm_id(stem: str) -> str:
    m = _RPM_PAT.search(stem)
    return m.group(1) if m else ""


def _detect_catalyst(stem: str) -> str:
    """Extract sample name from between the CV-type prefix and 'vs RE' in the filename."""
    m = _SAMPLE_NAME_PAT.search(stem)
    return m.group(1) if m else ""


def _detect_group_key(stem: str) -> str:
    """Treeview grouping key — same as the sample name extracted from the filename."""
    return _detect_catalyst(stem)


def _extract_anodic(E: np.ndarray, I: np.ndarray):
    """Return (E_sorted, I_sorted) for the anodic (E-increasing) half of the CV.

    Finds the cathodic vertex (minimum E), takes everything from that point
    onward, then sorts by E ascending so np.interp works correctly.
    """
    if len(E) < 4:
        return E, I
    min_idx = int(np.argmin(E))
    E_an = E[min_idx:]
    I_an = I[min_idx:]
    order = np.argsort(E_an)
    return E_an[order], I_an[order]


def _find_half_wave(E: np.ndarray, J: np.ndarray):
    """Return (E_half, J_half) for an ORR polarization curve (anodic scan, E ascending).

    J_lim = most negative J (diffusion-limited plateau at low E).
    E½    = E where J crosses J_lim / 2.
    Returns (None, None) when data is insufficient or has no cathodic current.
    """
    if len(E) < 4:
        return None, None
    j_lim = float(np.min(J))
    if j_lim >= 0:
        return None, None
    j_half = j_lim / 2.0
    diff = J - j_half
    sign_ch = np.where(np.diff(np.sign(diff)))[0]
    if len(sign_ch) == 0:
        return None, None
    idx = sign_ch[len(sign_ch) // 2]
    e0, e1 = E[idx], E[idx + 1]
    j0, j1 = J[idx], J[idx + 1]
    e_half = (e0 + (j_half - j0) / (j1 - j0) * (e1 - e0)
              if j1 != j0 else (e0 + e1) / 2.0)
    return float(e_half), float(j_half)


def _process_pair(pair: dict, r_sol_n2: float, r_sol_o2: float,
                  e_ref: float, area: float):
    """Return (E_plot, Y_plot) for one ORR pair, or None on failure."""
    df_n2 = pair.get("df_n2")
    df_o2 = pair.get("df_o2")
    if df_n2 is None or df_o2 is None:
        return None

    for df in (df_n2, df_o2):
        if "Ewe/V" not in df.columns or "I/mA" not in df.columns:
            return None

    def _last_cycle(df):
        if "cycle number" in df.columns and len(df["cycle number"].unique()) > 0:
            last = df["cycle number"].max()
            return df[df["cycle number"] == last].copy()
        return df.copy()

    lc_n2 = _last_cycle(df_n2)
    lc_o2 = _last_cycle(df_o2)
    if len(lc_n2) < 10 or len(lc_o2) < 10:
        return None

    E_n2 = lc_n2["Ewe/V"].values.astype(float)
    I_n2 = lc_n2["I/mA"].values.astype(float)
    E_o2 = lc_o2["Ewe/V"].values.astype(float)
    I_o2 = lc_o2["I/mA"].values.astype(float)

    # IR correction
    if r_sol_n2 != 0.0:
        E_n2 = E_n2 - (I_n2 / 1000.0) * r_sol_n2
    if r_sol_o2 != 0.0:
        E_o2 = E_o2 - (I_o2 / 1000.0) * r_sol_o2

    # RHE conversion
    if e_ref != 0.0:
        E_n2 = E_n2 + e_ref
        E_o2 = E_o2 + e_ref

    # Anodic scan
    E_n2_an, I_n2_an = _extract_anodic(E_n2, I_n2)
    E_o2_an, I_o2_an = _extract_anodic(E_o2, I_o2)
    if len(E_n2_an) < 5 or len(E_o2_an) < 5:
        return None

    # Restrict O2 to overlapping E range
    E_lo = max(E_n2_an[0],  E_o2_an[0])
    E_hi = min(E_n2_an[-1], E_o2_an[-1])
    if E_lo >= E_hi:
        return None

    mask    = (E_o2_an >= E_lo) & (E_o2_an <= E_hi)
    E_plot  = E_o2_an[mask]
    I_o2_cl = I_o2_an[mask]
    if len(E_plot) < 3:
        return None

    I_n2_interp = np.interp(E_plot, E_n2_an, I_n2_an)
    I_net       = I_o2_cl - I_n2_interp

    Y_plot = I_net / area if area > 0 else I_net
    return E_plot, Y_plot


# ════════════════════════════════════════════════════════════════════════
# ORRPanel
# ════════════════════════════════════════════════════════════════════════

class ORRPanel(ttk.Frame):
    """ORR analysis panel — group N2+O2 CV pairs into samples; plot background-subtracted
    ORR polarization curves (anodic scan, last cycle, per RPM)."""

    def __init__(self, master):
        ttk.Frame.__init__(self, master)
        self.loaded_files   = OrderedDict()  # short_name → {path, gas, rpm_id, df}
        self._loaded_keys   = []             # ordered list of short names (mirrors listbox)
        self.samples        = OrderedDict()  # sample_name → sample_entry
        self.active_sample  = None
        self._suppress_replot = False
        self._loading        = False
        self._drag                = None
        self._zoom_sample         = None
        self._copied_sample_fmt   = None  # clipboard for Copy/Paste format
        self._pair_sep_drag       = None  # catalyst-group drag-to-reorder state
        self._pair_sep_frames     = []    # [(sep_frame, cat_name)] for hit-testing
        self._pair_drop_line      = None  # blue drop indicator
        self._build_panel()
        self.after(500, self._auto_set_initial_size)

    # ════════════════════════════════════════════════════════════════
    # Panel construction
    # ════════════════════════════════════════════════════════════════
    def _build_panel(self):
        body = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        body.pack(fill=tk.BOTH, expand=True)

        # ── Scrollable left panel ─────────────────────────────────
        left_outer = ttk.Frame(body, width=360)
        body.add(left_outer, weight=0)

        _lc = tk.Canvas(left_outer, highlightthickness=0)
        _ls = ttk.Scrollbar(left_outer, orient=tk.VERTICAL, command=_lc.yview)
        _lc.configure(yscrollcommand=_ls.set)
        _ls.pack(side=tk.RIGHT, fill=tk.Y)
        _lc.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        left = tk.Frame(_lc)
        self._left_lc    = _lc    # outer scrollable canvas
        self._left_inner = left   # inner content frame
        _lwin = _lc.create_window((0, 0), window=left, anchor=tk.NW)
        left.bind("<Configure>", lambda e: _lc.configure(scrollregion=_lc.bbox("all")))
        _lc.bind("<Configure>", lambda e: _lc.itemconfig(_lwin, width=e.width))
        _lc.bind("<MouseWheel>", lambda e: _lc.yview_scroll(-1 * (e.delta // 120), "units"))

        # ══ LOADED FILES ════════════════════════════════════════════
        _lf_hdr = ttk.Frame(left)
        _lf_hdr.pack(fill=tk.X, padx=4, pady=(6, 0))
        ttk.Label(_lf_hdr, text="Loaded Files:", font=("", 9, "bold")).pack(side=tk.LEFT)

        ttk.Label(left, text="(N2/O2 and catalyst auto-detected from filename)",
                  foreground="gray", font=("", 8)).pack(anchor=tk.W, padx=4)
        _fb = ttk.Frame(left)
        _fb.pack(fill=tk.X, padx=4)
        ttk.Button(_fb, text="Load Files", command=self._load_files).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_fb, text="Remove",     command=self._remove_loaded_file).pack(side=tk.LEFT)

        _gb = ttk.Frame(left)
        _gb.pack(fill=tk.X, padx=4, pady=(1, 0))
        ttk.Button(_gb, text="Sel N2", width=7,
                   command=self._select_n2).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_gb, text="Sel O2", width=7,
                   command=self._select_o2).pack(side=tk.LEFT)

        _lf = ttk.Frame(left)
        _lf.pack(fill=tk.X, padx=4, pady=2)
        self.loaded_tv = ttk.Treeview(_lf, height=5, selectmode="extended", show="tree")
        self.loaded_tv.column("#0", minwidth=50, stretch=True)
        self.loaded_tv.tag_configure("n2",  background="#dbeafe", foreground="#1e40af")
        self.loaded_tv.tag_configure("o2",  background="#ffedd5", foreground="#9a3412")
        self.loaded_tv.tag_configure("unk", background="white")
        self.loaded_tv.tag_configure("hdr", background="#e8e8e8", foreground="#333333",
                                     font=("", 8, "bold"))
        self.loaded_tv.bind("<Double-1>", self._on_loaded_tv_dblclick)
        _lf_sb = ttk.Scrollbar(_lf, orient=tk.VERTICAL, command=self.loaded_tv.yview)
        self.loaded_tv.configure(yscrollcommand=_lf_sb.set)
        _lf_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.loaded_tv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Drag handle to resize the loaded-files treeview
        _lf_handle = tk.Frame(left, height=5, bg="#c8c8c8", cursor="sb_v_double_arrow")
        _lf_handle.pack(fill=tk.X, padx=4)
        _lf_handle.bind("<ButtonPress-1>", self._on_loaded_resize_start)
        _lf_handle.bind("<B1-Motion>",     self._on_loaded_resize_drag)

        # ══ SAMPLES ═════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=4)
        _sh_hdr = ttk.Frame(left)
        _sh_hdr.pack(fill=tk.X, padx=4)
        ttk.Label(_sh_hdr, text="ORR Samples:", font=("", 9, "bold")).pack(side=tk.LEFT)

        _sb = ttk.Frame(left)
        _sb.pack(fill=tk.X, padx=4, pady=(2, 0))
        ttk.Button(_sb, text="New Sample", command=self._new_sample).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(_sb, text="Rename",     command=self._rename_sample).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(_sb, text="Delete",     command=self._delete_sample).pack(side=tk.LEFT)

        _slf = ttk.Frame(left)
        _slf.pack(fill=tk.X, padx=4, pady=2)
        self.sample_lb = CheckableListbox(
            _slf, height=3,
            on_check=self._on_sample_visibility_change,
            on_reorder=self._on_sample_reorder)
        self.sample_lb.pack(fill=tk.X, expand=True)
        self.sample_lb.bind("<<ListboxSelect>>", self._on_sample_select)

        # Drag handle to resize the samples listbox
        _slf_handle = tk.Frame(left, height=5, bg="#c8c8c8", cursor="sb_v_double_arrow")
        _slf_handle.pack(fill=tk.X, padx=4)
        _slf_handle.bind("<ButtonPress-1>", self._on_sample_resize_start)
        _slf_handle.bind("<B1-Motion>",     self._on_sample_resize_drag)

        ttk.Button(left, text="↓ Add Selected Files to Sample",
                   command=self._add_files_to_sample).pack(fill=tk.X, padx=4, pady=(2, 0))

        _cp_row = ttk.Frame(left)
        _cp_row.pack(fill=tk.X, padx=4, pady=(2, 0))
        ttk.Button(_cp_row, text="Copy Format",
                   command=self._copy_sample_format).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_cp_row, text="Paste Format",
                   command=self._paste_sample_format).pack(side=tk.LEFT)
        ttk.Label(_cp_row, text="(grid/font/legend/reflines)",
                  foreground="gray", font=("", 7)).pack(side=tk.LEFT, padx=(6, 0))

        # ══ PAIR TABLE ══════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=4)
        ttk.Label(left, text="RPM Pairs  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        ttk.Label(left,
                  text="Edit RPM values; N2/O2 paired by filename RPM index",
                  foreground="gray", font=("", 8), wraplength=320).pack(
            anchor=tk.W, padx=4)

        _ptf = ttk.Frame(left)
        _ptf.pack(fill=tk.X, padx=4, pady=2)
        self._pair_tbl_canvas = tk.Canvas(_ptf, background="#f5f5f5",
                                          highlightthickness=1, highlightbackground="#cccccc",
                                          height=180)
        _pt_vs = ttk.Scrollbar(_ptf, orient=tk.VERTICAL, command=self._pair_tbl_canvas.yview)
        self._pair_tbl_canvas.configure(yscrollcommand=_pt_vs.set)
        _pt_vs.pack(side=tk.RIGHT, fill=tk.Y)
        self._pair_tbl_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._pair_tbl_inner = tk.Frame(self._pair_tbl_canvas, background="#f5f5f5")
        _pt_win = self._pair_tbl_canvas.create_window((0, 0), window=self._pair_tbl_inner, anchor=tk.NW)
        self._pair_tbl_inner.bind(
            "<Configure>",
            lambda e: self._pair_tbl_canvas.configure(scrollregion=self._pair_tbl_canvas.bbox("all")))
        self._pair_tbl_canvas.bind(
            "<Configure>", lambda e: self._pair_tbl_canvas.itemconfig(_pt_win, width=e.width))

        def _pt_wheel(e):
            self._pair_tbl_canvas.yview_scroll(-1 * (e.delta // 120), "units")
            return "break"
        self._pair_tbl_canvas.bind("<MouseWheel>", _pt_wheel)
        self._pair_tbl_inner.bind("<MouseWheel>", _pt_wheel)

        # Drag handle to resize the RPM pairs table
        _pt_handle = tk.Frame(left, height=5, bg="#c8c8c8", cursor="sb_v_double_arrow")
        _pt_handle.pack(fill=tk.X, padx=4)
        _pt_handle.bind("<ButtonPress-1>", self._on_pair_resize_start)
        _pt_handle.bind("<B1-Motion>",     self._on_pair_resize_drag)

        # ══ CORRECTION ══════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=4)
        ttk.Label(left, text="IR / RHE Correction  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)

        _cr0 = ttk.Frame(left)
        _cr0.pack(fill=tk.X, padx=4, pady=(2, 0))
        ttk.Label(_cr0, text="Catalyst:").pack(side=tk.LEFT)
        self._corr_catalyst_var = tk.StringVar(value="")
        self._corr_cat_cb = ttk.Combobox(
            _cr0, textvariable=self._corr_catalyst_var,
            values=[], state="readonly", width=14)
        self._corr_cat_cb.pack(side=tk.LEFT, padx=(4, 0))
        self._corr_cat_cb.bind("<<ComboboxSelected>>", self._on_corr_catalyst_select)

        _cst = ttk.Frame(left)
        _cst.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_cst, text="Color:").pack(side=tk.LEFT)
        self._cat_color_var = tk.StringVar(value="Blue")
        _cat_color_cb = ttk.Combobox(_cst, textvariable=self._cat_color_var,
                                      values=_COLOR_NAMES, state="readonly", width=12)
        _cat_color_cb.pack(side=tk.LEFT, padx=(2, 4))
        ttk.Label(_cst, text="Grad:").pack(side=tk.LEFT)
        self._gradient_step_var = tk.StringVar(value="0.10")
        _grad_e = ttk.Entry(_cst, textvariable=self._gradient_step_var, width=5)
        _grad_e.pack(side=tk.LEFT, padx=(2, 6))
        ttk.Label(_cst, text="Style:").pack(side=tk.LEFT)
        self._cat_ls_var = tk.StringVar(value="solid")
        _cat_ls_cb = ttk.Combobox(_cst, textvariable=self._cat_ls_var,
                                   values=["solid", "dashed", "dotted", "dash-dot"],
                                   state="readonly", width=8)
        _cat_ls_cb.pack(side=tk.LEFT, padx=(2, 6))
        ttk.Label(_cst, text="W:").pack(side=tk.LEFT)
        self._cat_lw_var = tk.StringVar(value="1.5")
        _cat_lw_e = ttk.Entry(_cst, textvariable=self._cat_lw_var, width=4)
        _cat_lw_e.pack(side=tk.LEFT, padx=(2, 6))
        ttk.Label(_cst, text="Marker:").pack(side=tk.LEFT)
        self._cat_mk_var = tk.StringVar(value="none")
        _cat_mk_cb = ttk.Combobox(_cst, textvariable=self._cat_mk_var,
                                   values=["none", "o", "s", "^", "D", "v", "x", "+"],
                                   state="readonly", width=6)
        _cat_mk_cb.pack(side=tk.LEFT, padx=(2, 0))

        _cat_lw_e.bind("<Return>",   lambda e: self._on_cat_style_change())
        _cat_lw_e.bind("<FocusOut>", lambda e: self._on_cat_style_change())
        _cat_color_cb.bind("<<ComboboxSelected>>", lambda e: self._on_cat_style_change())
        _cat_ls_cb.bind("<<ComboboxSelected>>", lambda e: self._on_cat_style_change())
        _cat_mk_cb.bind("<<ComboboxSelected>>", lambda e: self._on_cat_style_change())
        _grad_e.bind("<Return>",   lambda e: self._on_gradient_change())
        _grad_e.bind("<FocusOut>", lambda e: self._on_gradient_change())

        _cr1 = ttk.Frame(left)
        _cr1.pack(fill=tk.X, padx=4, pady=2)
        ttk.Label(_cr1, text="R_sol N2 (Ω):").pack(side=tk.LEFT)
        self.r_sol_n2_var = tk.StringVar(value="0")
        _n2_e = ttk.Entry(_cr1, textvariable=self.r_sol_n2_var, width=8)
        _n2_e.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(_cr1, text="R_sol O2 (Ω):").pack(side=tk.LEFT)
        self.r_sol_o2_var = tk.StringVar(value="0")
        _o2_e = ttk.Entry(_cr1, textvariable=self.r_sol_o2_var, width=8)
        _o2_e.pack(side=tk.LEFT, padx=(4, 0))
        for _e in (_n2_e, _o2_e):
            _e.bind("<Return>",   lambda ev: self._on_correction_change())
            _e.bind("<FocusOut>", lambda ev: self._on_correction_change())

        _cr2 = ttk.Frame(left)
        _cr2.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_cr2, text="E_ref (V vs RHE):").pack(side=tk.LEFT)
        self.e_ref_var = tk.StringVar(value="0")
        _eref_e = ttk.Entry(_cr2, textvariable=self.e_ref_var, width=8)
        _eref_e.pack(side=tk.LEFT, padx=(4, 12))
        _eref_e.bind("<Return>",   lambda ev: self._on_correction_change())
        _eref_e.bind("<FocusOut>", lambda ev: self._on_correction_change())
        ttk.Label(_cr2, text="Ref:").pack(side=tk.LEFT)
        self.ref_electrode_var = tk.StringVar(value="RHE")
        _ref_cb = ttk.Combobox(
            _cr2, textvariable=self.ref_electrode_var,
            values=["RHE", "Ag/AgCl", "SCE", "NHE", "SHE", "Hg/HgO", "Fc/Fc+"],
            state="readonly", width=10)
        _ref_cb.pack(side=tk.LEFT, padx=(2, 0))
        _ref_cb.bind("<<ComboboxSelected>>", lambda e: self._auto_replot())

        _cr3 = ttk.Frame(left)
        _cr3.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_cr3, text="Area (cm²):").pack(side=tk.LEFT)
        self.area_var = tk.StringVar(value="")
        _area_e = ttk.Entry(_cr3, textvariable=self.area_var, width=8)
        _area_e.pack(side=tk.LEFT, padx=(4, 6))
        ttk.Label(_cr3, text="(leave blank for I/mA; enter for J/mA/cm²)",
                  foreground="gray", font=("", 8)).pack(side=tk.LEFT)
        _area_e.bind("<Return>",   lambda ev: self._on_correction_change())
        _area_e.bind("<FocusOut>", lambda ev: self._on_correction_change())

        _cr4 = ttk.Frame(left)
        _cr4.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_cr4, text="ECSA_Hupd (cm²):").pack(side=tk.LEFT)
        self.ecsa_var = tk.StringVar(value="")
        _ecsa_hupd_e = ttk.Entry(_cr4, textvariable=self.ecsa_var, width=8)
        _ecsa_hupd_e.pack(side=tk.LEFT, padx=(4, 6))
        ttk.Label(_cr4, text="(from Hupd panel; used for SA in report)",
                  foreground="gray", font=("", 8)).pack(side=tk.LEFT)
        _ecsa_hupd_e.bind("<Return>",   lambda ev: self._on_correction_change())
        _ecsa_hupd_e.bind("<FocusOut>", lambda ev: self._on_correction_change())

        # Traces: immediately write any change to the active sample's dict so that
        # FocusOut events firing after a sample switch never overwrite the wrong sample.
        for _ck, _cv in [("r_sol_n2", self.r_sol_n2_var),
                         ("r_sol_o2", self.r_sol_o2_var),
                         ("e_ref",    self.e_ref_var)]:
            _cv.trace_add("write",
                          lambda *_a, k=_ck, v=_cv: self._on_corr_var_trace(k, v))
        self.area_var.trace_add("write",
                                lambda *_a: self._on_corr_var_trace("area", self.area_var))
        self.ecsa_var.trace_add("write",
                                lambda *_a: self._on_corr_var_trace("ecsa", self.ecsa_var))

        ttk.Label(left, text="(auto-applied on Enter / focus change)",
                  foreground="gray", font=("", 8)).pack(anchor=tk.W, padx=4)

        # ══ AXIS / RANGE ════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Label(left, text="Plot Range  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        _xr = ttk.Frame(left)
        _xr.pack(fill=tk.X, padx=4, pady=(1, 0))
        ttk.Label(_xr, text="X min:").pack(side=tk.LEFT)
        self.x_min_var = tk.StringVar()
        _xmin = ttk.Entry(_xr, textvariable=self.x_min_var, width=7)
        _xmin.pack(side=tk.LEFT, padx=(2, 4))
        ttk.Label(_xr, text="X max:").pack(side=tk.LEFT)
        self.x_max_var = tk.StringVar()
        _xmax = ttk.Entry(_xr, textvariable=self.x_max_var, width=7)
        _xmax.pack(side=tk.LEFT, padx=(2, 4))
        ttk.Label(_xr, text="Int:").pack(side=tk.LEFT)
        self.x_grid_int_var = tk.StringVar(value="0")
        _xgi = ttk.Entry(_xr, textvariable=self.x_grid_int_var, width=5)
        _xgi.pack(side=tk.LEFT, padx=(2, 0))

        _yr = ttk.Frame(left)
        _yr.pack(fill=tk.X, padx=4, pady=(1, 0))
        ttk.Label(_yr, text="Y min:").pack(side=tk.LEFT)
        self.y_min_var = tk.StringVar()
        _ymin = ttk.Entry(_yr, textvariable=self.y_min_var, width=7)
        _ymin.pack(side=tk.LEFT, padx=(2, 4))
        ttk.Label(_yr, text="Y max:").pack(side=tk.LEFT)
        self.y_max_var = tk.StringVar()
        _ymax = ttk.Entry(_yr, textvariable=self.y_max_var, width=7)
        _ymax.pack(side=tk.LEFT, padx=(2, 4))
        ttk.Label(_yr, text="Int:").pack(side=tk.LEFT)
        self.y_grid_int_var = tk.StringVar(value="0")
        _ygi = ttk.Entry(_yr, textvariable=self.y_grid_int_var, width=5)
        _ygi.pack(side=tk.LEFT, padx=(2, 0))

        ttk.Label(left, text="(blank = auto)", foreground="gray",
                  font=("", 8)).pack(anchor=tk.W, padx=4)
        for _re in (_xmin, _xmax, _ymin, _ymax, _xgi, _ygi):
            _re.bind("<Return>",   lambda e: self._auto_replot())
            _re.bind("<FocusOut>", lambda e: self._auto_replot())

        _flip_row = ttk.Frame(left)
        _flip_row.pack(fill=tk.X, padx=4, pady=(2, 2))
        self.x_flip_var = tk.BooleanVar(value=False)
        self.y_flip_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(_flip_row, text="Flip X", variable=self.x_flip_var,
                        command=self._auto_replot).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(_flip_row, text="Flip Y", variable=self.y_flip_var,
                        command=self._auto_replot).pack(side=tk.LEFT)

        _ehalf_row = ttk.Frame(left)
        _ehalf_row.pack(fill=tk.X, padx=4, pady=(2, 0))
        self.show_half_wave_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(_ehalf_row, text="Show E½ markers",
                        variable=self.show_half_wave_var,
                        command=self._auto_replot).pack(side=tk.LEFT)

        # ══ TITLE ═══════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        _title_row = ttk.Frame(left)
        _title_row.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_title_row, text="Title:").pack(side=tk.LEFT)
        self.plot_title_var = tk.StringVar(value="")
        _title_e = ttk.Entry(_title_row, textvariable=self.plot_title_var)
        _title_e.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))
        _title_e.bind("<Return>",   lambda e: self._auto_replot())
        _title_e.bind("<FocusOut>", lambda e: self._auto_replot())

        # ══ LEGEND ══════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Label(left, text="Legend  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        _lr1 = ttk.Frame(left)
        _lr1.pack(fill=tk.X, padx=4, pady=2)
        self.legend_show_var  = tk.BooleanVar(value=True)
        self.legend_frame_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(_lr1, text="Show Legend", variable=self.legend_show_var,
                        command=self._auto_replot).pack(side=tk.LEFT)
        ttk.Checkbutton(_lr1, text="Frame", variable=self.legend_frame_var,
                        command=self._auto_replot).pack(side=tk.LEFT, padx=(8, 0))
        _lr2 = ttk.Frame(left)
        _lr2.pack(fill=tk.X, padx=4, pady=2)
        ttk.Label(_lr2, text="Size:").pack(side=tk.LEFT)
        self.legend_size_var = tk.StringVar(value="8")
        _lsz_e = ttk.Entry(_lr2, textvariable=self.legend_size_var, width=4)
        _lsz_e.pack(side=tk.LEFT, padx=(2, 8))
        _lsz_e.bind("<Return>",   lambda e: self._auto_replot())
        _lsz_e.bind("<FocusOut>", lambda e: self._auto_replot())
        ttk.Label(_lr2, text="Loc:").pack(side=tk.LEFT)
        self.legend_loc_var = tk.StringVar(value="best")
        _leg_loc_cb = ttk.Combobox(
            _lr2, textvariable=self.legend_loc_var,
            values=["best", "upper right", "upper left", "lower left", "lower right",
                    "right", "center left", "center right", "lower center",
                    "upper center", "center"],
            state="readonly", width=11)
        _leg_loc_cb.pack(side=tk.LEFT, padx=2)

        def _on_leg_loc_select(e=None):
            if self.active_sample and self.active_sample in self.samples:
                self.samples[self.active_sample].pop("legend_manual_pos", None)
                _leg = self.samples[self.active_sample].get("legend")
                if _leg is not None:
                    _leg._loc = 0
            self._auto_replot()
        _leg_loc_cb.bind("<<ComboboxSelected>>", _on_leg_loc_select)
        ttk.Label(left, text="(left-drag to move)",
                  foreground="gray", font=("", 8)).pack(anchor=tk.W, padx=4)

        # ══ GRID ════════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Label(left, text="Grid  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        _gxy = ttk.Frame(left)
        _gxy.pack(fill=tk.X, padx=4, pady=2)
        self.x_grid_var = tk.BooleanVar(value=False)
        self.y_grid_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(_gxy, text="X", variable=self.x_grid_var,
                        command=self._auto_replot).pack(side=tk.LEFT)
        ttk.Checkbutton(_gxy, text="Y", variable=self.y_grid_var,
                        command=self._auto_replot).pack(side=tk.LEFT, padx=(10, 0))
        _gst = ttk.Frame(left)
        _gst.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_gst, text="Style:").pack(side=tk.LEFT)
        self.grid_style_var = tk.StringVar(value="dashed")
        _gscb = ttk.Combobox(_gst, textvariable=self.grid_style_var,
                              values=["dashed", "dotted", "solid", "dash-dot"],
                              state="readonly", width=9)
        _gscb.pack(side=tk.LEFT, padx=(2, 6))
        _gscb.bind("<<ComboboxSelected>>", lambda e: self._auto_replot())
        ttk.Label(_gst, text="Color:").pack(side=tk.LEFT)
        self.grid_color_var = tk.StringVar(value="gray")
        _gccb = ttk.Combobox(_gst, textvariable=self.grid_color_var,
                              values=["gray", "black", "red", "blue", "green",
                                      "orange", "purple", "crimson", "royalblue",
                                      "darkorange", "teal"],
                              state="readonly", width=9)
        _gccb.pack(side=tk.LEFT, padx=(2, 6))
        _gccb.bind("<<ComboboxSelected>>", lambda e: self._auto_replot())
        ttk.Label(_gst, text="Width:").pack(side=tk.LEFT)
        self.grid_linewidth_var = tk.StringVar(value="0.8")
        _glw = ttk.Entry(_gst, textvariable=self.grid_linewidth_var, width=4)
        _glw.pack(side=tk.LEFT, padx=(2, 0))
        _glw.bind("<Return>",   lambda e: self._auto_replot())
        _glw.bind("<FocusOut>", lambda e: self._auto_replot())

        # ══ FONT ════════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Label(left, text="Font  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        self.font_title_size_var = tk.StringVar(value="10")
        self.font_title_bold_var = tk.BooleanVar(value=False)
        self.font_label_size_var = tk.StringVar(value="10")
        self.font_label_bold_var = tk.BooleanVar(value=False)
        self.font_tick_size_var  = tk.StringVar(value="8")
        self.font_tick_bold_var  = tk.BooleanVar(value=False)
        for lbl, sz_v, bd_v in (
            ("Title:    Size", self.font_title_size_var, self.font_title_bold_var),
            ("Axis Lbl: Size", self.font_label_size_var, self.font_label_bold_var),
            ("Tick Nos: Size", self.font_tick_size_var,  self.font_tick_bold_var),
        ):
            _fr = ttk.Frame(left)
            _fr.pack(fill=tk.X, padx=4, pady=(2, 0))
            ttk.Label(_fr, text=lbl).pack(side=tk.LEFT)
            _fe = ttk.Entry(_fr, textvariable=sz_v, width=4)
            _fe.pack(side=tk.LEFT, padx=(2, 4))
            _fe.bind("<Return>",   lambda e: self._auto_replot())
            _fe.bind("<FocusOut>", lambda e: self._auto_replot())
            ttk.Checkbutton(_fr, text="Bold", variable=bd_v,
                            command=self._auto_replot).pack(side=tk.LEFT)
        _spc = ttk.Frame(left)
        _spc.pack(fill=tk.X, padx=4, pady=(2, 0))
        ttk.Label(_spc, text="Spacing: Title").pack(side=tk.LEFT)
        self.title_pad_var = tk.StringVar(value="6")
        _tpe = ttk.Entry(_spc, textvariable=self.title_pad_var, width=4)
        _tpe.pack(side=tk.LEFT, padx=(2, 6))
        _tpe.bind("<Return>",   lambda e: self._auto_replot())
        _tpe.bind("<FocusOut>", lambda e: self._auto_replot())
        ttk.Label(_spc, text="Label").pack(side=tk.LEFT)
        self.label_pad_var = tk.StringVar(value="4")
        _lpe = ttk.Entry(_spc, textvariable=self.label_pad_var, width=4)
        _lpe.pack(side=tk.LEFT, padx=(2, 0))
        _lpe.bind("<Return>",   lambda e: self._auto_replot())
        _lpe.bind("<FocusOut>", lambda e: self._auto_replot())

        # ══ REFERENCE LINES ══════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Label(left, text="Reference Lines  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        _ra = ttk.Frame(left)
        _ra.pack(fill=tk.X, padx=4, pady=2)
        ttk.Label(_ra, text="X:").pack(side=tk.LEFT)
        self._ref_x_var = tk.StringVar()
        _rx_e = ttk.Entry(_ra, textvariable=self._ref_x_var, width=7)
        _rx_e.pack(side=tk.LEFT, padx=(2, 0))
        ttk.Button(_ra, text="+X", width=3, command=self._add_xrefline).pack(
            side=tk.LEFT, padx=(2, 8))
        ttk.Label(_ra, text="Y:").pack(side=tk.LEFT)
        self._ref_y_var = tk.StringVar()
        _ry_e = ttk.Entry(_ra, textvariable=self._ref_y_var, width=7)
        _ry_e.pack(side=tk.LEFT, padx=(2, 0))
        ttk.Button(_ra, text="+Y", width=3, command=self._add_yrefline).pack(
            side=tk.LEFT, padx=2)
        _rx_e.bind("<Return>", lambda e: self._add_xrefline())
        _ry_e.bind("<Return>", lambda e: self._add_yrefline())

        _rl_row = ttk.Frame(left)
        _rl_row.pack(fill=tk.X, padx=4, pady=(0, 2))
        self._reflines_lb = tk.Listbox(_rl_row, height=3,
                                       selectmode=tk.SINGLE, exportselection=False)
        self._reflines_lb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._reflines_lb.bind("<<ListboxSelect>>", lambda e: self._on_refline_select())
        ttk.Button(_rl_row, text="Remove",
                   command=self._remove_refline).pack(side=tk.RIGHT, padx=(4, 0))
        _ro = ttk.Frame(left)
        _ro.pack(fill=tk.X, padx=4, pady=(0, 2))
        ttk.Label(_ro, text="Style:").pack(side=tk.LEFT)
        self._refline_style_var = tk.StringVar(value="dashed")
        _rls_cb = ttk.Combobox(_ro, textvariable=self._refline_style_var,
                                values=["dashed", "dotted", "solid", "dash-dot"],
                                state="readonly", width=9)
        _rls_cb.pack(side=tk.LEFT, padx=(2, 6))
        ttk.Label(_ro, text="Color:").pack(side=tk.LEFT)
        self._refline_color_var = tk.StringVar(value="dimgray")
        _rlc_cb = ttk.Combobox(_ro, textvariable=self._refline_color_var,
                                values=["dimgray", "black", "red", "blue", "green",
                                        "orange", "purple", "crimson", "royalblue",
                                        "darkorange", "teal", "saddlebrown"],
                                state="readonly", width=9)
        _rlc_cb.pack(side=tk.LEFT, padx=(2, 6))
        ttk.Label(_ro, text="Width:").pack(side=tk.LEFT)
        self._refline_lw_var = tk.StringVar(value="1.0")
        _rllw = ttk.Entry(_ro, textvariable=self._refline_lw_var, width=4)
        _rllw.pack(side=tk.LEFT, padx=(2, 0))
        _rls_cb.bind("<<ComboboxSelected>>", lambda e: self._on_refline_style_change())
        _rlc_cb.bind("<<ComboboxSelected>>", lambda e: self._on_refline_style_change())
        _rllw.bind("<Return>",   lambda e: self._on_refline_style_change())
        _rllw.bind("<FocusOut>", lambda e: self._on_refline_style_change())

        # ══ PLOT BUTTON + LOG ═══════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Button(left, text="Plot Active Sample",
                   command=self._auto_replot).pack(anchor=tk.W, padx=4, pady=(0, 4))

        # ══ ANALYSIS ════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Label(left, text="Analysis  (active sample)",
                  font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        _an_row = ttk.Frame(left)
        _an_row.pack(fill=tk.X, padx=4, pady=(2, 1))
        ttk.Button(_an_row, text="Tafel Analysis",
                   command=self._open_tafel_window).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_an_row, text="KL Analysis",
                   command=self._open_kl_window).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_an_row, text="SA Analysis",
                   command=self._open_sa_window).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_an_row, text="Levich Plot",
                   command=self._open_levich_window).pack(side=tk.LEFT)
        _an_row2 = ttk.Frame(left)
        _an_row2.pack(fill=tk.X, padx=4, pady=(1, 2))
        ttk.Button(_an_row2, text="Sample Comparison",
                   command=self._open_comparison_window).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(_an_row2, text="Lim. Current Compare",
                   command=self._open_lc_comparison_window).pack(side=tk.LEFT)
        _an_row3 = ttk.Frame(left)
        _an_row3.pack(fill=tk.X, padx=4, pady=(1, 2))
        ttk.Button(_an_row3, text="Extract Report",
                   command=self._open_report_window).pack(side=tk.LEFT)

        # ══ EXPORT ══════════════════════════════════════════════════
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=4, pady=6)
        ttk.Button(left, text="Export Active Sample to Excel",
                   command=self._export_sample_excel).pack(anchor=tk.W, padx=4, pady=(0, 4))

        ttk.Label(left, text="Log", font=("", 9, "bold")).pack(anchor=tk.W, padx=4)
        _logf = ttk.Frame(left)
        _logf.pack(fill=tk.X, padx=4, pady=2)
        self.log_text = tk.Text(_logf, height=4, state=tk.DISABLED,
                                wrap=tk.WORD, relief=tk.SUNKEN, borderwidth=1)
        _log_sc = ttk.Scrollbar(_logf, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=_log_sc.set)
        _log_sc.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # ── Right panel ────────────────────────────────────────────
        right_outer = ttk.Frame(body)
        body.add(right_outer, weight=1)
        right_outer.rowconfigure(0, weight=0)
        right_outer.rowconfigure(1, weight=0)
        right_outer.rowconfigure(2, weight=1)
        right_outer.columnconfigure(0, weight=1)

        self.plot_w_var     = tk.StringVar(value="10.5")
        self.plot_h_var     = tk.StringVar(value="5.5")
        self._grid_cols_var = tk.StringVar(value="2")
        _size_bar = ttk.Frame(right_outer)
        _size_bar.grid(row=0, column=0, sticky="ew", padx=4, pady=2)
        ttk.Label(_size_bar, text="Plot size (in):").pack(side=tk.LEFT, padx=(4, 2))
        ttk.Label(_size_bar, text="W").pack(side=tk.LEFT)
        _pw_e = ttk.Entry(_size_bar, textvariable=self.plot_w_var, width=5)
        _pw_e.pack(side=tk.LEFT, padx=(1, 6))
        ttk.Label(_size_bar, text="H").pack(side=tk.LEFT)
        _ph_e = ttk.Entry(_size_bar, textvariable=self.plot_h_var, width=5)
        _ph_e.pack(side=tk.LEFT, padx=(1, 0))
        ttk.Label(_size_bar, text="Cols:").pack(side=tk.LEFT, padx=(10, 2))
        _gc_e = ttk.Entry(_size_bar, textvariable=self._grid_cols_var, width=3)
        _gc_e.pack(side=tk.LEFT, padx=(1, 0))
        for _e in (_pw_e, _ph_e):
            _e.bind("<Return>",   lambda ev: self._apply_plot_size())
            _e.bind("<FocusOut>", lambda ev: self._apply_plot_size())
        _gc_e.bind("<Return>",   lambda ev: self._on_grid_cols_change())
        _gc_e.bind("<FocusOut>", lambda ev: self._on_grid_cols_change())

        self._zoom_bar = ttk.Frame(right_outer)
        ttk.Button(self._zoom_bar, text="← Back to Grid",
                   command=self._unzoom_sample_view).pack(side=tk.LEFT, padx=6, pady=3)
        self._zoom_bar.grid(row=1, column=0, sticky="ew")
        self._zoom_bar.grid_remove()

        _ri = ttk.Frame(right_outer)
        _ri.grid(row=2, column=0, sticky="nsew")
        _ri.rowconfigure(0, weight=1)
        _ri.columnconfigure(0, weight=1)

        self._right_canvas = tk.Canvas(_ri, highlightthickness=0)
        _rvs = ttk.Scrollbar(_ri, orient=tk.VERTICAL,   command=self._right_canvas.yview)
        _rhs = ttk.Scrollbar(_ri, orient=tk.HORIZONTAL, command=self._right_canvas.xview)
        self._right_canvas.configure(yscrollcommand=_rvs.set, xscrollcommand=_rhs.set)
        _rvs.grid(row=0, column=1, sticky="ns")
        _rhs.grid(row=1, column=0, sticky="ew")
        self._right_canvas.grid(row=0, column=0, sticky="nsew")

        self._plots_frame = ttk.Frame(self._right_canvas)
        self._plots_win   = self._right_canvas.create_window(
            (0, 0), window=self._plots_frame, anchor=tk.NW)
        self._plots_frame.bind(
            "<Configure>",
            lambda e: self._right_canvas.configure(
                scrollregion=self._right_canvas.bbox("all")))
        self._right_canvas.bind("<Configure>", self._on_right_canvas_configure)
        self._right_canvas.bind(
            "<MouseWheel>",
            lambda e: self._right_canvas.yview_scroll(-1 * (e.delta // 120), "units"))
        self._right_canvas.bind(
            "<Shift-MouseWheel>",
            lambda e: self._right_canvas.xview_scroll(-1 * (e.delta // 120), "units"))

        self._drop_line = tk.Frame(self._plots_frame, bg="#1a73e8", height=3)

        self._placeholder = ttk.Label(
            self._plots_frame,
            text="Create samples and load N2+O2 files to display ORR polarization curves.",
            foreground="gray", font=("", 10))
        self._placeholder.grid(row=0, column=0, columnspan=2, pady=60)

    # ════════════════════════════════════════════════════════════════
    # File loading (no auto-merge)
    # ════════════════════════════════════════════════════════════════
    def _load_files(self):
        paths = filedialog.askopenfilenames(
            title="Load ORR Files (N2 and O2)",
            filetypes=[("EC-Lab files", "*.mpr *.txt"), ("All files", "*.*")])
        if not paths:
            return
        added, errors = [], []
        for path in paths:
            stem  = os.path.splitext(os.path.basename(path))[0]
            short = os.path.basename(path)
            if short in self.loaded_files:
                continue
            try:
                df = _read_one_df(path)
            except Exception as exc:
                errors.append(f"{short}: {exc}")
                continue
            gas      = _detect_gas(stem)
            rpm_id   = _extract_rpm_id(stem)
            catalyst = _detect_catalyst(stem)
            self.loaded_files[short] = {
                "path": path, "gas": gas, "rpm_id": rpm_id,
                "catalyst": catalyst, "df": df}
            self._loaded_keys.append(short)
            group_key = _detect_group_key(stem)
            cat_iid   = f"_cat_:{group_key}"
            cat_label = group_key if group_key else "(no catalyst)"
            if not self.loaded_tv.exists(cat_iid):
                self.loaded_tv.insert("", tk.END, iid=cat_iid,
                                      text=f"  ▸ {cat_label}",
                                      tags=("hdr",), open=True)
            gas_tag = gas.upper() if gas else "??"
            self.loaded_tv.insert(cat_iid, tk.END, iid=short,
                                  text=f"    ({gas_tag})  {short}",
                                  tags=(gas or "unk",))
            added.append(short)

        if errors:
            messagebox.showerror("Load Error", "\n".join(errors[:5]))
        if added:
            self._log(f"Loaded {len(added)} file(s).")

    def _remove_loaded_file(self):
        sel = [iid for iid in self.loaded_tv.selection()
               if not iid.startswith("_cat_:")]
        if not sel:
            return
        for short in sel:
            if short not in self.loaded_files:
                continue
            parent = self.loaded_tv.parent(short)
            self.loaded_tv.delete(short)
            self.loaded_files.pop(short, None)
            if short in self._loaded_keys:
                self._loaded_keys.remove(short)
            if parent and not self.loaded_tv.get_children(parent):
                self.loaded_tv.delete(parent)

    def _on_loaded_tv_dblclick(self, event):
        """Double-click on a catalyst header → rename it."""
        iid = self.loaded_tv.identify_row(event.y)
        if iid and iid.startswith("_cat_:"):
            self._set_catalyst_name(iid)

    def _set_catalyst_name(self, cat_iid=None):
        """Rename the catalyst group and propagate to pairs."""
        if cat_iid is None:
            sel = self.loaded_tv.selection()
            if not sel:
                return
            iid = sel[0]
            cat_iid = iid if iid.startswith("_cat_:") else self.loaded_tv.parent(iid)
        if not cat_iid or not cat_iid.startswith("_cat_:"):
            return

        old_cat = cat_iid[len("_cat_:"):]
        if not cat_iid or not cat_iid.startswith("_cat_:"):
            messagebox.showwarning("ORR", "Could not determine catalyst group.")
            return

        old_cat = cat_iid[len("_cat_:"):]

        from tkinter.simpledialog import askstring
        new_cat = askstring(
            "Rename Catalyst",
            "Catalyst name:",
            initialvalue=old_cat,
            parent=self,
        )
        if new_cat is None:
            return
        new_cat = new_cat.strip()
        if new_cat == old_cat:
            return

        file_shorts = set(self.loaded_tv.get_children(cat_iid))

        # Update loaded_files
        for short in file_shorts:
            if short in self.loaded_files:
                self.loaded_files[short]["catalyst"] = new_cat

        # Update treeview header text
        self.loaded_tv.item(cat_iid,
                            text=f"  ▸ {new_cat if new_cat else '(no catalyst)'}")

        # Propagate to all sample pairs that reference these files
        affected_samples = set()
        for sname, sentry in self.samples.items():
            cc = sentry.setdefault("catalyst_corrections", {})
            cs = sentry.setdefault("catalyst_styles", {})
            renamed: dict = {}  # old_id → new_id

            for pair in sentry["pairs"]:
                if (pair.get("n2_short") not in file_shorts
                        and pair.get("o2_short") not in file_shorts):
                    continue
                old_base = pair.get("catalyst_base", "")
                old_id   = pair.get("catalyst_id",   "")
                suffix   = old_id[len(old_base):] if old_id.startswith(old_base) else ""
                new_id   = new_cat + suffix
                renamed[old_id] = new_id
                pair["catalyst_base"] = new_cat
                pair["catalyst_id"]   = new_id
                affected_samples.add(sname)

            for old_id, new_id in renamed.items():
                if old_id != new_id:
                    if old_id in cc:
                        cc[new_id] = cc.pop(old_id)
                    if old_id in cs:
                        cs[new_id] = cs.pop(old_id)

        for sname in affected_samples:
            self._plot_sample(sname)
        if self.active_sample in affected_samples:
            self._rebuild_pair_table(self.active_sample)
            self._update_catalyst_selector(self.active_sample)

    def _select_n2(self):
        shorts = [k for k in self._loaded_keys
                  if self.loaded_files.get(k, {}).get("gas") == "n2"]
        self.loaded_tv.selection_set(shorts)

    def _select_o2(self):
        shorts = [k for k in self._loaded_keys
                  if self.loaded_files.get(k, {}).get("gas") == "o2"]
        self.loaded_tv.selection_set(shorts)

    # ════════════════════════════════════════════════════════════════
    # Sample management
    # ════════════════════════════════════════════════════════════════
    def _new_sample(self):
        name = simpledialog.askstring("New ORR Sample", "Sample name:", parent=self)
        if not name:
            return
        name = name.strip()
        if not name or name in self.samples:
            return
        self.samples[name] = {"pairs": [], "reflines": []}
        self._rebuild_sample_listbox()
        self._create_sample_figure(name)
        self._save_active_sample_state()
        self._switch_active_sample(name)

    def _rename_sample(self):
        sel = self.sample_lb.curselection()
        if not sel:
            return
        old = list(self.samples.keys())[sel[0]]
        new = simpledialog.askstring("Rename Sample", "New name:",
                                     initialvalue=old, parent=self)
        if not new:
            return
        new = new.strip()
        if not new or new == old or new in self.samples:
            return
        new_samples = OrderedDict()
        for k, v in self.samples.items():
            new_samples[new if k == old else k] = v
        self.samples = new_samples
        if self.active_sample == old:
            self.active_sample = new
        self._rebuild_sample_listbox()
        gentry = self.samples.get(new, {})
        if "ax" in gentry:
            ax = gentry["ax"]
            ax.set_title(gentry.get("custom_title", "") or new, fontsize=9)
            gentry["canvas"].draw_idle()

    def _delete_sample(self):
        sel = self.sample_lb.curselection()
        if not sel:
            return
        name = list(self.samples.keys())[sel[0]]
        self._destroy_sample_figure(name)
        del self.samples[name]
        if self.active_sample == name:
            self.active_sample = None
            self._rebuild_pair_table(None)
            self._refresh_reflines_lb()
        if not self.samples:
            self._rebuild_sample_listbox()
            self._placeholder.grid(row=0, column=0, columnspan=2, pady=60)
        else:
            self._rebuild_sample_listbox()
            self._relayout_figures()
            new_idx = min(sel[0], self.sample_lb.size() - 1)
            if new_idx >= 0:
                self._switch_active_sample(list(self.samples.keys())[new_idx])

    def _rebuild_sample_listbox(self):
        self.sample_lb.clear()
        for sn, sentry in self.samples.items():
            vis = not sentry.get("hidden", False)
            self.sample_lb.insert(tk.END, sn, checked=vis)
        if self.active_sample and self.active_sample in self.samples:
            idx = list(self.samples.keys()).index(self.active_sample)
            self._loading = True
            try:
                self.sample_lb.selection_clear(0, tk.END)
                self.sample_lb.selection_set(idx)
            finally:
                self._loading = False

    def _on_sample_select(self, event):
        if self._loading:
            return
        sel = self.sample_lb.curselection()
        if not sel:
            return
        keys = list(self.samples.keys())
        if sel[0] >= len(keys):
            return
        name = keys[sel[0]]
        if name != self.active_sample:
            self._save_active_sample_state()
            self._switch_active_sample(name)

    def _on_sample_visibility_change(self, sample_name, visible):
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return
        sentry["hidden"] = not visible
        self._relayout_figures()
        self._right_canvas.after(50, lambda: self._right_canvas.configure(
            scrollregion=self._right_canvas.bbox("all")))

    def _on_sample_reorder(self, new_names):
        self.samples = OrderedDict(
            (n, self.samples[n]) for n in new_names if n in self.samples)
        self._relayout_figures()

    # ════════════════════════════════════════════════════════════════
    # File → Sample assignment + pairing
    # ════════════════════════════════════════════════════════════════
    def _add_files_to_sample(self):
        if not self.active_sample:
            messagebox.showwarning("ORR", "Create or select a sample first.")
            return
        sel_iids = self.loaded_tv.selection()
        if not sel_iids:
            return
        sentry = self.samples[self.active_sample]

        expanded = []
        for iid in sel_iids:
            if iid.startswith("_cat_:"):
                expanded.extend(self.loaded_tv.get_children(iid))
            else:
                expanded.append(iid)
        selected_shorts = [iid for iid in expanded if iid in self.loaded_files]

        unknown = []
        file_entries = []
        for short in selected_shorts:
            fe = self.loaded_files.get(short)
            if fe is None:
                continue
            if fe["gas"] not in ("n2", "o2"):
                unknown.append(short)
            else:
                file_entries.append((short, fe))

        if unknown:
            ans = messagebox.askyesno(
                "Undetected Gas",
                f"{len(unknown)} file(s) have no 'n2'/'o2' in their name.\n"
                "Skip them?  (Yes = skip, No = cancel)")
            if not ans:
                return

        # Track all paths already committed to any pair slot (exact-dup guard).
        all_existing_paths = set()
        for p in sentry["pairs"]:
            if p.get("n2_path"):
                all_existing_paths.add(p["n2_path"])
            if p.get("o2_path"):
                all_existing_paths.add(p["o2_path"])

        added = 0

        # ── Batch-isolation: group this batch by (catalyst_base, rpm_id) ──
        # First-seen gas wins each slot; duplicates fall through to lone-file path.
        batch_groups: dict = {}
        for short, fe in file_entries:
            key = (fe.get("catalyst", ""), fe["rpm_id"])
            gas = fe["gas"]
            batch_groups.setdefault(key, {})
            if gas not in batch_groups[key]:
                batch_groups[key][gas] = (short, fe)

        # Step 1: create complete pairs from within this batch.
        # Batch-internal pairs are never merged into existing incomplete pairs —
        # this prevents cross-contamination when the same catalyst/rpm appears in
        # multiple experiments loaded in separate batches.
        batch_paired: set = set()
        for (cat_base, rpm_id), gas_map in batch_groups.items():
            if "n2" not in gas_map or "o2" not in gas_map:
                continue
            n2_short, n2_fe = gas_map["n2"]
            o2_short, o2_fe = gas_map["o2"]
            if n2_fe["path"] in all_existing_paths or o2_fe["path"] in all_existing_paths:
                batch_paired.add(n2_short)
                batch_paired.add(o2_short)
                continue
            used_cat_rpm = {(p.get("catalyst_id", ""), p.get("rpm_id", ""))
                            for p in sentry["pairs"]}
            cat_label = cat_base
            suffix = 2
            while (cat_label, rpm_id) in used_cat_rpm:
                cat_label = f"{cat_base}_{suffix}"
                suffix += 1
            new_pair = {
                "catalyst_id":   cat_label,
                "catalyst_base": cat_base,
                "rpm_id":        rpm_id,
                "rpm_val":       rpm_id,
                "n2_short":      n2_short,
                "o2_short":      o2_short,
                "n2_path":       n2_fe["path"],
                "o2_path":       o2_fe["path"],
                "df_n2":         n2_fe["df"],
                "df_o2":         o2_fe["df"],
            }
            sentry["pairs"].append(new_pair)
            all_existing_paths.add(n2_fe["path"])
            all_existing_paths.add(o2_fe["path"])
            batch_paired.add(n2_short)
            batch_paired.add(o2_short)
            cc = sentry.setdefault("catalyst_corrections", {})
            cc.setdefault(cat_label, {"r_sol_n2": 0.0, "r_sol_o2": 0.0,
                                      "e_ref": 0.0, "area": "", "ecsa": ""})
            cs = sentry.setdefault("catalyst_styles", {})
            cs.setdefault(cat_label, {"color": "", "linestyle": "solid",
                                      "linewidth": "1.5", "marker": "none"})
            added += 2

        # Step 2: process lone files not handled by batch-internal pairing.
        for short, fe in file_entries:
            if short in batch_paired:
                continue
            gas      = fe["gas"]
            rpm_id   = fe["rpm_id"]
            catalyst = fe.get("catalyst", "")
            path     = fe["path"]

            if path in all_existing_paths:
                continue

            # Try to merge into an existing incomplete pair that matches
            # (catalyst_base, rpm_id) and has this gas slot empty.
            merged = False
            for pair in sentry["pairs"]:
                pair_base = pair.get("catalyst_base", pair.get("catalyst_id", ""))
                if (pair_base == catalyst
                        and pair.get("rpm_id") == rpm_id
                        and not pair.get(f"{gas}_path")):
                    pair[f"{gas}_path"]  = path
                    pair[f"{gas}_short"] = short
                    pair[f"df_{gas}"]    = fe["df"]
                    merged = True
                    break

            if merged:
                all_existing_paths.add(path)
                added += 1
                continue

            # No mergeable slot — create a new (possibly incomplete) pair.
            # Auto-suffix catalyst label if (label, rpm_id) is already taken.
            used_cat_rpm = {(p.get("catalyst_id", ""), p.get("rpm_id", ""))
                            for p in sentry["pairs"]}
            cat_label = catalyst
            suffix = 2
            while (cat_label, rpm_id) in used_cat_rpm:
                cat_label = f"{catalyst}_{suffix}"
                suffix += 1
            new_pair = {
                "catalyst_id":   cat_label,
                "catalyst_base": catalyst,
                "rpm_id":        rpm_id,
                "rpm_val":       rpm_id,
                "n2_short":      short if gas == "n2" else "",
                "o2_short":      short if gas == "o2" else "",
                "n2_path":       path  if gas == "n2" else "",
                "o2_path":       path  if gas == "o2" else "",
                "df_n2":         fe["df"] if gas == "n2" else None,
                "df_o2":         fe["df"] if gas == "o2" else None,
            }
            sentry["pairs"].append(new_pair)
            all_existing_paths.add(path)
            cc = sentry.setdefault("catalyst_corrections", {})
            cc.setdefault(cat_label, {"r_sol_n2": 0.0, "r_sol_o2": 0.0,
                                      "e_ref": 0.0, "area": "", "ecsa": ""})
            cs = sentry.setdefault("catalyst_styles", {})
            cs.setdefault(cat_label, {"color": "", "linestyle": "solid",
                                      "linewidth": "1.5", "marker": "none"})
            added += 1

        sentry["pairs"].sort(key=lambda p: (p.get("catalyst_id", ""),
                                            _safe_rpm_int(p.get("rpm_id", ""))))
        # Positional default rpm_val: 1st pair→400, 2nd→900, 3rd→1600, 4th→2500
        # Only applies when rpm_val == rpm_id (never manually edited by user)
        _cat_pos: dict = {}
        for _p in sentry["pairs"]:
            _c = _p.get("catalyst_id", "")
            _pos = _cat_pos.get(_c, 0)
            _cat_pos[_c] = _pos + 1
            if _p.get("rpm_val", "") == _p.get("rpm_id", "") and _pos < len(_RPM_DEFAULTS):
                _p["rpm_val"] = str(_RPM_DEFAULTS[_pos])
        if added:
            self._rebuild_pair_table(self.active_sample)
            self._auto_replot()
            self._log(f"Added {added} pair(s) to sample '{self.active_sample}'.")
            self._update_catalyst_selector(self.active_sample)
        else:
            self._log("No new pairs added (already exist or no matching IDs).")

    # ════════════════════════════════════════════════════════════════
    # Pair table
    # ════════════════════════════════════════════════════════════════
    def _rebuild_pair_table(self, sample_name):
        for w in self._pair_tbl_inner.winfo_children():
            w.destroy()

        sentry = self.samples.get(sample_name) if sample_name else None
        pairs  = sentry["pairs"] if sentry else []

        if not pairs:
            tk.Label(self._pair_tbl_inner,
                     text="No pairs yet. Load files and click\n"
                          "'↓ Add Selected Files to Sample'.",
                     bg="#f5f5f5", fg="gray", font=("", 8),
                     justify=tk.LEFT).pack(padx=6, pady=6, anchor=tk.W)
            return

        # Header — checkbox col + catalyst + rpm fixed; N2/O2 expand; remove btn right
        hdr = tk.Frame(self._pair_tbl_inner, bg="#f5f5f5")
        hdr.pack(fill=tk.X, padx=2, pady=(2, 0))
        tk.Label(hdr, text="Plot", width=4, bg="#f5f5f5",
                 font=("", 8, "bold"), anchor=tk.W).pack(side=tk.LEFT)
        tk.Label(hdr, text="Catalyst", width=7, bg="#f5f5f5",
                 font=("", 8, "bold"), anchor=tk.W).pack(side=tk.LEFT)
        tk.Label(hdr, text="RPM", width=5, bg="#f5f5f5",
                 font=("", 8, "bold"), anchor=tk.W).pack(side=tk.LEFT)
        tk.Label(hdr, text="", width=2, bg="#f5f5f5").pack(side=tk.RIGHT)
        tk.Label(hdr, text="O2 file", bg="#f5f5f5",
                 font=("", 8, "bold"), anchor=tk.W).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(2, 0))
        tk.Label(hdr, text="N2 file", bg="#f5f5f5",
                 font=("", 8, "bold"), anchor=tk.W).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0))

        self._pair_sep_frames = []
        prev_cat = None
        row_idx  = 0
        for pair in pairs:
            cat = pair.get("catalyst_id", "")
            if cat != prev_cat:
                _sep_bg = "#e8d5f5"
                _sep = tk.Frame(self._pair_tbl_inner, bg=_sep_bg)
                _sep.pack(fill=tk.X, padx=2, pady=(6 if prev_cat is not None else 2, 0))
                _sh = tk.Label(_sep, text="⠿", bg=_sep_bg, cursor="fleur", font=("", 10))
                _sh.pack(side=tk.LEFT, padx=(4, 0))
                # Group-level checkbox — toggles all pairs in this catalyst group
                _cat_pairs = [p for p in pairs if p.get("catalyst_id") == cat]
                _cat_en_val = all(p.get("enabled", True) for p in _cat_pairs)
                _cat_en_var = tk.BooleanVar(value=_cat_en_val)
                def _toggle_cat(bv=_cat_en_var, c=cat, sn=sample_name):
                    v = bv.get()
                    for _p in self.samples.get(sn, {}).get("pairs", []):
                        if _p.get("catalyst_id") == c:
                            _p["enabled"] = v
                    self._rebuild_pair_table(sn)
                    self._auto_replot()
                tk.Checkbutton(_sep, variable=_cat_en_var, command=_toggle_cat,
                               bg=_sep_bg, activebackground=_sep_bg,
                               relief=tk.FLAT).pack(side=tk.LEFT, padx=(0, 0))
                tk.Label(_sep, text=f"── {cat} ──", bg=_sep_bg,
                         font=("", 8, "bold"), anchor=tk.W, padx=4).pack(
                    side=tk.LEFT, fill=tk.X, expand=True)
                _c = cat
                _sh.bind("<ButtonPress-1>",
                         lambda e, c=_c, sn=sample_name: self._on_pair_sep_press(e, c, sn))
                _sh.bind("<B1-Motion>",
                         lambda e: self._on_pair_sep_drag(e))
                _sh.bind("<ButtonRelease-1>",
                         lambda e, sn=sample_name: self._on_pair_sep_release(e, sn))
                self._pair_sep_frames.append((_sep, cat))
                prev_cat = cat
                row_idx  = 0

            i = row_idx
            row_idx += 1
            row_bg = "#ffffff" if i % 2 == 0 else "#eeeeee"
            row = tk.Frame(self._pair_tbl_inner, bg=row_bg)
            row.pack(fill=tk.X, padx=2, pady=1)

            cat_var     = tk.StringVar(value=pair.get("catalyst_id", ""))
            rpm_var     = tk.StringVar(value=pair.get("rpm_val", pair.get("rpm_id", "")))
            enabled_var = tk.BooleanVar(value=pair.get("enabled", True))

            def _toggle_enabled(p=pair, ev=enabled_var):
                p["enabled"] = ev.get()
                self._auto_replot()

            def _save_pair(cv=cat_var, rv=rpm_var, p=pair, sn=sample_name):
                new_cat = cv.get().strip()
                old_cat = p.get("catalyst_id", "")
                if new_cat and new_cat != old_cat:
                    sentry_ref = self.samples.get(sn, {})
                    for _p in sentry_ref.get("pairs", []):
                        if _p.get("catalyst_id") == old_cat:
                            _p["catalyst_id"] = new_cat
                    for _store_key in ("catalyst_corrections", "catalyst_styles"):
                        _store = sentry_ref.get(_store_key, {})
                        if old_cat in _store:
                            _store[new_cat] = _store.pop(old_cat)
                    if getattr(self, "_active_catalyst", None) == old_cat:
                        self._active_catalyst = new_cat
                    self._rebuild_pair_table(sn)
                    self._update_catalyst_selector(sn)
                else:
                    p["catalyst_id"] = new_cat or old_cat
                p["rpm_val"] = rv.get()
                self._auto_replot()

            # Checkbox — left edge
            tk.Checkbutton(row, variable=enabled_var, command=_toggle_enabled,
                           bg=row_bg, activebackground=row_bg,
                           relief=tk.FLAT).pack(side=tk.LEFT, padx=(2, 0))

            cat_e = tk.Entry(row, textvariable=cat_var, width=7, bg=row_bg,
                             relief=tk.GROOVE)
            cat_e.pack(side=tk.LEFT, padx=(1, 1))
            cat_e.bind("<Return>",   lambda e, fn=_save_pair: fn())
            cat_e.bind("<FocusOut>", lambda e, fn=_save_pair: fn())

            rpm_e = tk.Entry(row, textvariable=rpm_var, width=5, bg=row_bg,
                             relief=tk.GROOVE)
            rpm_e.pack(side=tk.LEFT, padx=(0, 2))
            rpm_e.bind("<Return>",   lambda e, fn=_save_pair: fn())
            rpm_e.bind("<FocusOut>", lambda e, fn=_save_pair: fn())

            n2_s = pair.get("n2_short", "") or "—"
            o2_s = pair.get("o2_short", "") or "—"
            n2_bg = "#d4edda" if pair.get("n2_short") else "#f8d7da"
            o2_bg = "#d1ecf1" if pair.get("o2_short") else "#f8d7da"

            def _remove(p=pair, sn=sample_name):
                if sn and sn in self.samples:
                    try:
                        self.samples[sn]["pairs"].remove(p)
                    except ValueError:
                        pass
                    self._rebuild_pair_table(sn)
                    self._auto_replot()

            # Remove button right-anchored so filename labels get all remaining space
            tk.Button(row, text="✕", width=2, command=_remove,
                      bg=row_bg, relief=tk.FLAT, font=("", 8)).pack(side=tk.RIGHT, padx=(1, 2))
            tk.Label(row, text=o2_s, bg=o2_bg, font=("", 7), anchor=tk.W,
                     relief=tk.GROOVE).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(0, 1))
            tk.Label(row, text=n2_s, bg=n2_bg, font=("", 7), anchor=tk.W,
                     relief=tk.GROOVE).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 1))

    # ════════════════════════════════════════════════════════════════
    # Pair catalyst-group drag-to-reorder
    # ════════════════════════════════════════════════════════════════
    def _on_pair_sep_press(self, event, cat, sample_name):
        self._pair_sep_drag = {"from_cat": cat, "sample_name": sample_name,
                               "target_cat": None}

    def _on_pair_sep_drag(self, event):
        d = self._pair_sep_drag
        if d is None:
            return
        y_root = event.y_root
        target_cat = None
        for frame, c in self._pair_sep_frames:
            try:
                fy = frame.winfo_rooty()
                fh = max(frame.winfo_height(), 1)
                if y_root < fy + fh / 2:
                    target_cat = c
                    break
                target_cat = c
            except Exception:
                pass
        d["target_cat"] = target_cat
        # Drop indicator
        if self._pair_drop_line is None:
            self._pair_drop_line = tk.Frame(self._pair_tbl_inner,
                                            bg="#1a73e8", height=2)
        for frame, c in self._pair_sep_frames:
            if c == target_cat:
                try:
                    self._pair_drop_line.place(
                        x=0, y=frame.winfo_y(), relwidth=1.0, height=2)
                    self._pair_drop_line.lift()
                except Exception:
                    pass
                return
        self._pair_drop_line.place_forget()

    def _on_pair_sep_release(self, event, sample_name):
        if self._pair_drop_line is not None:
            try:
                self._pair_drop_line.place_forget()
            except Exception:
                pass
        d = self._pair_sep_drag
        self._pair_sep_drag = None
        if d is None:
            return
        from_cat   = d["from_cat"]
        target_cat = d.get("target_cat")
        if target_cat is None or target_cat == from_cat:
            return
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return
        pairs = sentry["pairs"]
        # Determine current catalyst order
        cats_ordered = []
        for p in pairs:
            c = p.get("catalyst_id", "")
            if c not in cats_ordered:
                cats_ordered.append(c)
        if from_cat not in cats_ordered or target_cat not in cats_ordered:
            return
        cats_ordered.remove(from_cat)
        to_idx = cats_ordered.index(target_cat)
        cats_ordered.insert(to_idx, from_cat)
        # Rebuild pairs in new catalyst order
        new_pairs = []
        for c in cats_ordered:
            new_pairs.extend(p for p in pairs if p.get("catalyst_id", "") == c)
        sentry["pairs"] = new_pairs
        self._rebuild_pair_table(sample_name)
        self._plot_sample(sample_name)

    # ════════════════════════════════════════════════════════════════
    # Copy / Paste sample display format
    # ════════════════════════════════════════════════════════════════
    _FMT_KEYS = (
        "x_grid", "y_grid", "x_grid_int", "y_grid_int",
        "grid_style", "grid_color", "grid_linewidth",
        "x_flip", "y_flip",
        "legend_show", "legend_frame", "leg_size", "legend_loc",
        "font_title_size", "font_title_bold",
        "font_label_size", "font_label_bold",
        "font_tick_size",  "font_tick_bold",
        "title_pad", "label_pad",
        "ref_electrode",
    )

    def _copy_sample_format(self):
        if not self.active_sample or self.active_sample not in self.samples:
            return
        self._save_active_sample_state()
        g = self.samples[self.active_sample]
        self._copied_sample_fmt = {k: g.get(k) for k in self._FMT_KEYS}
        self._copied_sample_fmt["reflines"] = list(g.get("reflines", []))

    def _paste_sample_format(self):
        if not self._copied_sample_fmt:
            return
        if not self.active_sample or self.active_sample not in self.samples:
            return
        g = self.samples[self.active_sample]
        p = self._copied_sample_fmt
        for k in self._FMT_KEYS:
            if k in p and p[k] is not None:
                g[k] = p[k]
        if "reflines" in p:
            g["reflines"] = list(p["reflines"])
        self._switch_active_sample(self.active_sample)
        self._plot_sample(self.active_sample)

    # ════════════════════════════════════════════════════════════════
    # Active sample state save / restore
    # ════════════════════════════════════════════════════════════════
    def _save_active_sample_state(self):
        if not self.active_sample or self.active_sample not in self.samples:
            return
        g = self.samples[self.active_sample]

        def _fv(var, default=0.0):
            try:
                return float(var.get())
            except (ValueError, tk.TclError):
                return default

        g["ref_electrode"]   = self.ref_electrode_var.get()
        g["x_min"]           = self.x_min_var.get()
        g["x_max"]           = self.x_max_var.get()
        g["y_min"]           = self.y_min_var.get()
        g["y_max"]           = self.y_max_var.get()
        g["x_grid_int"]      = self.x_grid_int_var.get()
        g["y_grid_int"]      = self.y_grid_int_var.get()
        g["x_flip"]          = self.x_flip_var.get()
        g["y_flip"]          = self.y_flip_var.get()
        g["legend_show"]     = self.legend_show_var.get()
        g["legend_frame"]    = self.legend_frame_var.get()
        try:
            g["leg_size"]    = float(self.legend_size_var.get())
        except (ValueError, tk.TclError):
            pass
        g["legend_loc"]      = self.legend_loc_var.get()
        g["x_grid"]          = self.x_grid_var.get()
        g["y_grid"]          = self.y_grid_var.get()
        g["grid_style"]      = self.grid_style_var.get()
        g["grid_color"]      = self.grid_color_var.get()
        g["grid_linewidth"]  = self.grid_linewidth_var.get()
        g["font_title_size"] = self.font_title_size_var.get()
        g["font_title_bold"] = self.font_title_bold_var.get()
        g["font_label_size"] = self.font_label_size_var.get()
        g["font_label_bold"] = self.font_label_bold_var.get()
        g["font_tick_size"]  = self.font_tick_size_var.get()
        g["font_tick_bold"]  = self.font_tick_bold_var.get()
        g["title_pad"]       = self.title_pad_var.get()
        g["label_pad"]       = self.label_pad_var.get()
        g["custom_title"]    = self.plot_title_var.get()
        g["show_half_wave"]  = self.show_half_wave_var.get()
        try:
            g["gradient_step"] = float(self._gradient_step_var.get())
        except ValueError:
            pass

    def _switch_active_sample(self, sample_name):
        self._switching_sample = True
        self.active_sample = sample_name
        g = self.samples.get(sample_name, {})

        g.setdefault("ref_electrode",   "RHE")
        g.setdefault("x_min",           "")
        g.setdefault("x_max",           "")
        g.setdefault("y_min",           "")
        g.setdefault("y_max",           "")
        g.setdefault("x_grid_int",      "0")
        g.setdefault("y_grid_int",      "0")
        g.setdefault("x_flip",          False)
        g.setdefault("y_flip",          False)
        g.setdefault("legend_show",     True)
        g.setdefault("legend_frame",    True)
        g.setdefault("leg_size",        8.0)
        g.setdefault("legend_loc",      "best")
        g.setdefault("x_grid",          False)
        g.setdefault("y_grid",          False)
        g.setdefault("grid_style",      "dashed")
        g.setdefault("grid_color",      "gray")
        g.setdefault("grid_linewidth",  "0.8")
        g.setdefault("font_title_size", "10")
        g.setdefault("font_title_bold", False)
        g.setdefault("font_label_size", "10")
        g.setdefault("font_label_bold", False)
        g.setdefault("font_tick_size",  "8")
        g.setdefault("font_tick_bold",  False)
        g.setdefault("title_pad",       "6")
        g.setdefault("label_pad",       "4")
        g.setdefault("custom_title",    "")
        g.setdefault("show_half_wave",  True)
        g.setdefault("gradient_step",   0.10)
        g.setdefault("reflines",        [])
        g.setdefault("custom_xlabel",   None)
        g.setdefault("custom_ylabel",   None)

        old = self._suppress_replot
        self._suppress_replot = True
        try:
            def _sv(var, key, default=""):
                v = g.get(key, default)
                var.set("" if v == 0.0 and default == "" else str(v))

            self._update_catalyst_selector(sample_name)
            self.ref_electrode_var.set(g["ref_electrode"])
            self.x_min_var.set(g["x_min"])
            self.x_max_var.set(g["x_max"])
            self.y_min_var.set(g["y_min"])
            self.y_max_var.set(g["y_max"])
            self.x_grid_int_var.set(g["x_grid_int"])
            self.y_grid_int_var.set(g["y_grid_int"])
            self.x_flip_var.set(g["x_flip"])
            self.y_flip_var.set(g["y_flip"])
            self.legend_show_var.set(g["legend_show"])
            self.legend_frame_var.set(g["legend_frame"])
            _ls = g["leg_size"]
            self.legend_size_var.set(str(int(_ls) if float(_ls) == int(_ls) else _ls))
            self.legend_loc_var.set(g["legend_loc"])
            self.x_grid_var.set(g["x_grid"])
            self.y_grid_var.set(g["y_grid"])
            self.grid_style_var.set(g["grid_style"])
            self.grid_color_var.set(g["grid_color"])
            self.grid_linewidth_var.set(g["grid_linewidth"])
            self.font_title_size_var.set(g["font_title_size"])
            self.font_title_bold_var.set(g["font_title_bold"])
            self.font_label_size_var.set(g["font_label_size"])
            self.font_label_bold_var.set(g["font_label_bold"])
            self.font_tick_size_var.set(g["font_tick_size"])
            self.font_tick_bold_var.set(g["font_tick_bold"])
            self.title_pad_var.set(g["title_pad"])
            self.label_pad_var.set(g["label_pad"])
            self.plot_title_var.set(g.get("custom_title", ""))
            self.show_half_wave_var.set(g.get("show_half_wave", True))
            self._gradient_step_var.set(str(g.get("gradient_step", 0.10)))
            self._rebuild_pair_table(sample_name)
            self._refresh_reflines_lb()
        finally:
            self._suppress_replot = old
            self._switching_sample = False

        self._highlight_active_headers()
        self._auto_replot()

    # ════════════════════════════════════════════════════════════════
    # Per-catalyst correction UI helpers
    # ════════════════════════════════════════════════════════════════
    def _update_catalyst_selector(self, sample_name):
        """Rebuild the catalyst combobox and load the first (or previously selected) catalyst."""
        sentry = self.samples.get(sample_name, {})
        cats = list(sentry.get("catalyst_corrections", {}).keys())
        # Ensure any catalyst that has pairs but no entry yet is included
        for p in sentry.get("pairs", []):
            c = p.get("catalyst_id", "")
            if c and c not in cats:
                cats.append(c)
        self._corr_cat_cb["values"] = cats
        if cats:
            # Try to keep the previously selected catalyst if it still exists
            prev = getattr(self, "_active_catalyst", None)
            sel = prev if prev in cats else cats[0]
            self._corr_catalyst_var.set(sel)
            self._load_catalyst_corrections(sample_name, sel)
        else:
            self._corr_catalyst_var.set("")
            self._active_catalyst = None
            self._switching_sample = True
            try:
                self.r_sol_n2_var.set("0")
                self.r_sol_o2_var.set("0")
                self.e_ref_var.set("0")
                self.area_var.set("")
                self.ecsa_var.set("")
                self._cat_color_var.set("Blue")
                self._cat_ls_var.set("solid")
                self._cat_lw_var.set("1.5")
                self._cat_mk_var.set("none")
            finally:
                self._switching_sample = False

    def _load_catalyst_corrections(self, sample_name, catalyst_id):
        """Load a catalyst's stored correction values into the shared UI vars."""
        sentry = self.samples.get(sample_name, {})
        cc = sentry.get("catalyst_corrections", {}).get(catalyst_id, {})
        self._switching_sample = True
        try:
            self.r_sol_n2_var.set(str(cc.get("r_sol_n2", 0.0)))
            self.r_sol_o2_var.set(str(cc.get("r_sol_o2", 0.0)))
            self.e_ref_var.set(str(cc.get("e_ref", 0.0)))
            self.area_var.set(cc.get("area", ""))
            self.ecsa_var.set(cc.get("ecsa", ""))
        finally:
            self._switching_sample = False
        self._active_catalyst = catalyst_id
        sentry = self.samples.get(sample_name, {})
        cs = sentry.get("catalyst_styles", {}).get(catalyst_id, {})
        self._switching_sample = True
        try:
            # Stored color is hex (or empty); reverse-lookup to display name
            _hex = (cs.get("color", "") or "").strip()
            _name = next((n for n, h in _COLOR_HEX.items() if h.lower() == _hex.lower()),
                         "Blue")
            self._cat_color_var.set(_name)
            self._cat_ls_var.set(cs.get("linestyle", "solid"))
            self._cat_lw_var.set(cs.get("linewidth", "1.5"))
            self._cat_mk_var.set(cs.get("marker", "none"))
        finally:
            self._switching_sample = False

    def _on_corr_catalyst_select(self, event=None):
        """User picked a different catalyst in the combobox — load its corrections."""
        cat = self._corr_catalyst_var.get()
        if cat and self.active_sample and self.active_sample in self.samples:
            self._load_catalyst_corrections(self.active_sample, cat)

    def _on_cat_style_change(self):
        if getattr(self, "_switching_sample", False):
            return
        if not self.active_sample or self.active_sample not in self.samples:
            return
        cat = getattr(self, "_active_catalyst", None)
        if not cat:
            return
        sentry = self.samples[self.active_sample]
        cs = sentry.setdefault("catalyst_styles", {})
        cat_cs = cs.setdefault(cat, {"color": "", "linestyle": "solid",
                                      "linewidth": "1.5", "marker": "none"})
        # Convert color name → hex for storage (so existing color-resolution
        # logic in _get_curves_for_sample / catalyst_styles continues to work).
        _name = self._cat_color_var.get().strip()
        cat_cs["color"]     = _COLOR_HEX.get(_name, _name)
        cat_cs["linestyle"] = self._cat_ls_var.get()
        cat_cs["linewidth"] = self._cat_lw_var.get()
        cat_cs["marker"]    = self._cat_mk_var.get()
        self._auto_replot()

    def _on_gradient_change(self):
        if getattr(self, "_switching_sample", False):
            return
        if not self.active_sample or self.active_sample not in self.samples:
            return
        try:
            step = float(self._gradient_step_var.get())
        except ValueError:
            return
        self.samples[self.active_sample]["gradient_step"] = step
        self._auto_replot()

    # ════════════════════════════════════════════════════════════════
    # Correction trigger + per-sample immediate write
    # ════════════════════════════════════════════════════════════════
    def _on_corr_var_trace(self, key: str, var):
        if getattr(self, "_switching_sample", False):
            return
        if not self.active_sample or self.active_sample not in self.samples:
            return
        cat = getattr(self, "_active_catalyst", None)
        if not cat:
            return
        sentry = self.samples[self.active_sample]
        cc = sentry.setdefault("catalyst_corrections", {})
        cat_cc = cc.setdefault(cat, {"r_sol_n2": 0.0, "r_sol_o2": 0.0,
                                      "e_ref": 0.0, "area": "", "ecsa": ""})
        try:
            raw = var.get()
        except tk.TclError:
            return
        if key in ("area", "ecsa"):
            cat_cc[key] = raw
        else:
            try:
                cat_cc[key] = float(raw or 0)
            except ValueError:
                pass

    def _on_correction_change(self):
        if getattr(self, "_switching_sample", False):
            return
        self._save_active_sample_state()
        self._auto_replot()

    # ════════════════════════════════════════════════════════════════
    # Plotting
    # ════════════════════════════════════════════════════════════════
    def _auto_replot(self):
        if self._suppress_replot or not self.active_sample:
            return
        self._plot_sample(self.active_sample)

    def _plot_sample(self, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None or "ax" not in sentry:
            return

        is_active = (sample_name == self.active_sample)

        def _gv(key, default=""):
            var = getattr(self, key + "_var", None)
            return (var.get() if (is_active and var is not None)
                    else sentry.get(key, default))

        ref      = _gv("ref_electrode", "RHE")
        leg_show = sentry.get("legend_show", True)  if not is_active else self.legend_show_var.get()
        leg_frm  = sentry.get("legend_frame", True) if not is_active else self.legend_frame_var.get()
        leg_loc  = sentry.get("legend_loc", "best") if not is_active else self.legend_loc_var.get()
        try:
            leg_size = float(sentry.get("leg_size", 8.0) if not is_active
                             else self.legend_size_var.get())
        except (ValueError, TypeError):
            leg_size = 8.0

        ax     = sentry["ax"]
        canvas = sentry["canvas"]

        # Preserve zoom
        _prev_view = ((ax.get_xlim(), ax.get_ylim())
                      if sentry.get("auto_xlim") is not None else None)

        # Save legend manual position
        _old_leg = sentry.get("legend")
        if _old_leg is not None:
            _loc = getattr(_old_leg, "_loc", None)
            if isinstance(_loc, (tuple, list)):
                sentry["legend_manual_pos"] = tuple(_loc)
        sentry["legend"] = None
        self._clear_ann(sample_name, redraw=False)
        ax.clear()

        pairs = sentry.get("pairs", [])
        plot_data = []    # (E_arr, Y_arr, label, color) — cached for annotation

        # Group pairs by catalyst_id to assign per-catalyst base color + RPM gradient
        catalyst_order = []
        pairs_by_cat   = {}
        for pair in pairs:
            if not pair.get("enabled", True):
                continue
            cat = pair.get("catalyst_id", "")
            if cat not in catalyst_order:
                catalyst_order.append(cat)
                pairs_by_cat[cat] = []
            pairs_by_cat[cat].append(pair)

        try:
            _gstep = float(sentry.get("gradient_step", 0.10))
        except (ValueError, TypeError):
            _gstep = 0.10

        for ci, cat in enumerate(catalyst_order):
            _cat_st  = sentry.get("catalyst_styles", {}).get(cat, {})
            _user_col = _cat_st.get("color", "").strip()
            # User-set color becomes the gradient base; fall back to palette
            base_col  = _user_col if _user_col else _PALETTE[ci % len(_PALETTE)]
            cat_pairs = sorted(pairs_by_cat[cat],
                               key=lambda p: _safe_rpm_int(
                                   p.get("rpm_val") or p.get("rpm_id", "")))
            n_cat = len(cat_pairs)
            cat_colors = (_cycle_colors(base_col, n_cat, step=_gstep, reverse=False)
                          if n_cat > 1 else [base_col])
            for j, pair in enumerate(cat_pairs):
                if not pair.get("n2_short") or not pair.get("o2_short"):
                    continue
                _cat_cc = sentry.get("catalyst_corrections", {}).get(cat, {})
                _r_n2 = float(_cat_cc.get("r_sol_n2", 0) or 0)
                _r_o2 = float(_cat_cc.get("r_sol_o2", 0) or 0)
                _eref  = float(_cat_cc.get("e_ref", 0) or 0)
                _area  = float(_cat_cc.get("area", "") or 0)
                result = _process_pair(pair, _r_n2, _r_o2, _eref, _area)
                if result is None:
                    continue
                E_plot, Y_plot = result
                rpm_val = pair.get("rpm_val") or pair.get("rpm_id") or f"#{j+1}"
                prefix  = f"[{cat}] " if cat else ""
                label   = f"{prefix}{rpm_val} rpm"
                _col = cat_colors[j]
                try:
                    _lw = float(_cat_st.get("linewidth", "1.5") or 1.5)
                except (ValueError, TypeError):
                    _lw = 1.5
                _ls = {"solid": "-", "dashed": "--", "dotted": ":",
                       "dash-dot": "-."}.get(_cat_st.get("linestyle", "solid"), "-")
                _mk = _cat_st.get("marker", "none")
                _mk = None if _mk == "none" else _mk
                ax.plot(E_plot, Y_plot, color=_col, linewidth=_lw, linestyle=_ls,
                        marker=_mk, markersize=4, label=label, zorder=2)
                plot_data.append((E_plot, Y_plot, label, _col))

        sentry["_plot_data"] = plot_data

        # E½ markers
        show_hw = (self.show_half_wave_var.get() if is_active
                   else sentry.get("show_half_wave", True))
        if show_hw:
            for E_pd, Y_pd, lbl_pd, col_pd in plot_data:
                e_half, j_half = _find_half_wave(E_pd, Y_pd)
                if e_half is not None:
                    ax.axvline(e_half, color=col_pd, linestyle="--",
                               linewidth=0.8, alpha=0.55, label="_ehalf")
                    ax.plot([e_half], [j_half], marker="o", color=col_pd,
                            ms=4, zorder=5, linestyle="none", label="_ehalf_dot")
                    ax.annotate(
                        f"E½={e_half:.3f} V",
                        xy=(e_half, j_half),
                        xytext=(4, -14), textcoords="offset points",
                        fontsize=6, color=col_pd,
                        bbox=dict(boxstyle="round,pad=0.2", fc="white",
                                  alpha=0.75, ec=col_pd, lw=0.5),
                    ).set_in_layout(False)

        # Axis labels
        x_label = f"E (V vs {ref})"
        _any_area = any(
            float(sentry.get("catalyst_corrections", {}).get(c, {}).get("area", "") or 0) > 0
            for c in catalyst_order)
        y_label  = "J (mA cm⁻²)" if _any_area else "I (mA)"
        try:
            lbl_sz  = int(sentry.get("font_label_size", "10") if not is_active
                          else self.font_label_size_var.get())
            lbl_wt  = "bold" if (sentry.get("font_label_bold", False) if not is_active
                                 else self.font_label_bold_var.get()) else "normal"
            tick_sz = int(sentry.get("font_tick_size", "8") if not is_active
                          else self.font_tick_size_var.get())
            tick_wt = "bold" if (sentry.get("font_tick_bold", False) if not is_active
                                 else self.font_tick_bold_var.get()) else "normal"
            tit_sz  = int(sentry.get("font_title_size", "10") if not is_active
                          else self.font_title_size_var.get())
            tit_wt  = "bold" if (sentry.get("font_title_bold", False) if not is_active
                                 else self.font_title_bold_var.get()) else "normal"
            t_pad   = int(sentry.get("title_pad", "6") if not is_active
                          else self.title_pad_var.get())
            l_pad   = int(sentry.get("label_pad", "4") if not is_active
                          else self.label_pad_var.get())
        except (ValueError, TypeError):
            lbl_sz = 10; lbl_wt = "normal"; tick_sz = 8; tick_wt = "normal"
            tit_sz = 10; tit_wt = "normal"; t_pad = 6; l_pad = 4

        ax.set_xlabel(sentry.get("custom_xlabel") or x_label,
                      fontsize=lbl_sz, fontweight=lbl_wt, labelpad=l_pad)
        ax.set_ylabel(sentry.get("custom_ylabel") or y_label,
                      fontsize=lbl_sz, fontweight=lbl_wt, labelpad=l_pad)
        ax.tick_params(labelsize=tick_sz)
        for lbl in ax.get_xticklabels() + ax.get_yticklabels():
            lbl.set_fontweight(tick_wt)

        title = (sentry.get("custom_title", "") if not is_active
                 else self.plot_title_var.get())
        ax.set_title(title, fontsize=tit_sz, fontweight=tit_wt, pad=t_pad)

        # Legend
        _leg = None
        if leg_show and plot_data:
            _leg = ax.legend(fontsize=leg_size, frameon=leg_frm, loc=leg_loc)
            _mp  = sentry.get("legend_manual_pos")
            if isinstance(_mp, (tuple, list)):
                _set = getattr(_leg, "_set_loc", None)
                if callable(_set):
                    _set(tuple(_mp))
                else:
                    _leg._loc = tuple(_mp)
            sentry["legend_label_order"] = [t.get_text() for t in _leg.get_texts()]
            _leg.set_draggable(True)
        sentry["legend"] = _leg

        # Layout
        _lv = ax.get_legend()
        if _lv:
            _lv.set_visible(False)
        try:
            sentry["fig"].tight_layout(pad=0.5)
            sentry["fig"].set_layout_engine("none")
        except Exception:
            pass
        if _lv:
            _lv.set_visible(True)

        canvas.draw()

        sentry["auto_xlim"] = ax.get_xlim()
        sentry["auto_ylim"] = ax.get_ylim()

        if _prev_view:
            ax.set_xlim(_prev_view[0])
            ax.set_ylim(_prev_view[1])

        # Manual range
        self._apply_sample_range(sample_name, is_active)

        # Grid + tick intervals
        try:
            x_g  = sentry.get("x_grid", False) if not is_active else self.x_grid_var.get()
            y_g  = sentry.get("y_grid", False) if not is_active else self.y_grid_var.get()
            g_st = sentry.get("grid_style", "dashed") if not is_active else self.grid_style_var.get()
            g_co = sentry.get("grid_color",  "gray")  if not is_active else self.grid_color_var.get()
            g_lw = float(sentry.get("grid_linewidth", "0.8") if not is_active
                         else self.grid_linewidth_var.get())
            xi   = sentry.get("x_grid_int", "0") if not is_active else self.x_grid_int_var.get()
            yi   = sentry.get("y_grid_int", "0") if not is_active else self.y_grid_int_var.get()
            apply_grid(ax, x_g, y_g, xi, yi, g_st, linewidth=g_lw, color=g_co)
        except Exception:
            pass

        # Reference lines
        reflines = (sentry.get("reflines", []) if not is_active
                    else self._get_current_reflines())
        draw_reflines(ax, reflines)

        # Re-apply catalyst highlight (legend handles already drawn at alpha=1)
        self._apply_orr_highlight(sample_name)

        canvas.draw_idle()

    def _apply_sample_range(self, sample_name, is_active=None):
        sentry = self.samples.get(sample_name)
        if sentry is None or "ax" not in sentry:
            return
        if is_active is None:
            is_active = (sample_name == self.active_sample)
        ax = sentry["ax"]

        def _val(key, var_attr):
            if is_active:
                var = getattr(self, var_attr, None)
                return var.get() if var else ""
            return sentry.get(key, "")

        try:
            xmin = float(_val("x_min", "x_min_var")); ax.set_xlim(left=xmin)
        except (ValueError, TypeError): pass
        try:
            xmax = float(_val("x_max", "x_max_var")); ax.set_xlim(right=xmax)
        except (ValueError, TypeError): pass
        try:
            ymin = float(_val("y_min", "y_min_var")); ax.set_ylim(bottom=ymin)
        except (ValueError, TypeError): pass
        try:
            ymax = float(_val("y_max", "y_max_var")); ax.set_ylim(top=ymax)
        except (ValueError, TypeError): pass

        xl = ax.get_xlim()
        yl = ax.get_ylim()
        flip_x = sentry.get("x_flip", False) if not is_active else self.x_flip_var.get()
        flip_y = sentry.get("y_flip", False) if not is_active else self.y_flip_var.get()
        if flip_x and xl[0] < xl[1]:
            ax.set_xlim(xl[1], xl[0])
        elif not flip_x and xl[0] > xl[1]:
            ax.set_xlim(xl[1], xl[0])
        if flip_y and yl[0] < yl[1]:
            ax.set_ylim(yl[1], yl[0])
        elif not flip_y and yl[0] > yl[1]:
            ax.set_ylim(yl[1], yl[0])

    def _reset_sample_view(self, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry and sentry.get("auto_xlim") is not None:
            sentry["ax"].set_xlim(sentry["auto_xlim"])
            sentry["ax"].set_ylim(sentry["auto_ylim"])
            sentry["canvas"].draw_idle()

    # ════════════════════════════════════════════════════════════════
    # Reference lines
    # ════════════════════════════════════════════════════════════════
    def _get_current_reflines(self):
        if not self.active_sample:
            return []
        return self.samples.get(self.active_sample, {}).get("reflines", [])

    def _add_xrefline(self):
        try:
            val = float(self._ref_x_var.get())
        except ValueError:
            return
        self._add_refline("x", val)

    def _add_yrefline(self):
        try:
            val = float(self._ref_y_var.get())
        except ValueError:
            return
        self._add_refline("y", val)

    def _add_refline(self, axis, val):
        if not self.active_sample or self.active_sample not in self.samples:
            return
        self._save_active_sample_state()
        style = self._refline_style_var.get()
        color = self._refline_color_var.get()
        try:
            lw = float(self._refline_lw_var.get())
        except ValueError:
            lw = 1.0
        self.samples[self.active_sample]["reflines"].append((axis, val, style, color, lw))
        self._refresh_reflines_lb()
        self._auto_replot()

    def _remove_refline(self):
        if not self.active_sample:
            return
        sel = self._reflines_lb.curselection()
        if not sel:
            return
        reflines = self.samples[self.active_sample].get("reflines", [])
        idx = sel[0]
        if idx < len(reflines):
            reflines.pop(idx)
        self._refresh_reflines_lb()
        self._auto_replot()

    def _on_refline_select(self):
        if not self.active_sample:
            return
        sel = self._reflines_lb.curselection()
        if not sel:
            return
        reflines = self.samples[self.active_sample].get("reflines", [])
        idx = sel[0]
        if idx < len(reflines):
            rl = reflines[idx]
            style = rl[2] if len(rl) > 2 else "dashed"
            color = rl[3] if len(rl) > 3 else "dimgray"
            lw    = rl[4] if len(rl) > 4 else 1.0
            self._refline_style_var.set(style)
            self._refline_color_var.set(color)
            self._refline_lw_var.set(str(lw))

    def _on_refline_style_change(self):
        if not self.active_sample:
            return
        sel = self._reflines_lb.curselection()
        if not sel:
            return
        reflines = self.samples[self.active_sample].get("reflines", [])
        idx = sel[0]
        if idx < len(reflines):
            rl = list(reflines[idx])
            rl[2] = self._refline_style_var.get()
            rl[3] = self._refline_color_var.get()
            try:
                rl_lw = float(self._refline_lw_var.get())
            except ValueError:
                rl_lw = 1.0
            if len(rl) > 4:
                rl[4] = rl_lw
            else:
                rl.append(rl_lw)
            reflines[idx] = tuple(rl)
        self._auto_replot()

    def _refresh_reflines_lb(self):
        self._reflines_lb.delete(0, tk.END)
        if not self.active_sample:
            return
        for rl in self.samples.get(self.active_sample, {}).get("reflines", []):
            axis = rl[0]; val = rl[1]
            self._reflines_lb.insert(tk.END, f"{axis.upper()}={val:.4g}")

    # ════════════════════════════════════════════════════════════════
    # Figure creation / destruction / layout
    # ════════════════════════════════════════════════════════════════
    def _create_sample_figure(self, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None or "fig" in sentry:
            return
        panel_ref = self
        self._placeholder.grid_remove()

        frame  = tk.Frame(self._plots_frame, relief="groove", bd=2)
        header = tk.Frame(frame, bg=_SAMPLE_HDR_BG, cursor="fleur")
        header.pack(fill=tk.X, side=tk.TOP)
        lbl = tk.Label(header, text=f"⠿  {sample_name}",
                       bg=_SAMPLE_HDR_BG, font=("", 9, "bold"), anchor=tk.W)
        lbl.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6, pady=3)

        inner = ttk.Frame(frame, padding=(4, 2, 4, 2))
        inner.pack()

        try:
            _fw = float(self.plot_w_var.get())
            _fh = float(self.plot_h_var.get())
        except (ValueError, AttributeError):
            _fw, _fh = 9.5, 5.5

        fig = Figure(figsize=(_fw, _fh), dpi=100)
        ax  = fig.add_subplot(111)
        canvas = FigureCanvasTkAgg(fig, master=inner)
        canvas.get_tk_widget().pack()

        tb_frame = ttk.Frame(inner)
        tb_frame.pack(fill=tk.X)

        class _Toolbar(NavigationToolbar2Tk):
            def home(tb_self, *a):
                panel_ref._reset_sample_view(sample_name)

        _tb = _Toolbar(canvas, tb_frame, pack_toolbar=False)
        _tb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        _tb.update()
        tk.Button(tb_frame, text="Copy",
                  command=lambda f=fig: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
            side=tk.LEFT, padx=(4, 2), pady=1)

        def _fwd(e):
            self._right_canvas.yview_scroll(-1 * (e.delta // 120), "units")
        frame.bind("<MouseWheel>",    _fwd)
        header.bind("<MouseWheel>",   _fwd)
        tb_frame.bind("<MouseWheel>", _fwd)

        for _w in (header, lbl):
            _w.bind("<ButtonPress-1>",   lambda e, s=sample_name: self._on_frame_press(e, s))
            _w.bind("<B1-Motion>",       lambda e, s=sample_name: self._on_frame_drag(e, s))
            _w.bind("<ButtonRelease-1>", lambda e, s=sample_name: self._on_frame_release(e, s))
            _w.bind("<Double-Button-1>", lambda e, s=sample_name: self._toggle_zoom(s))

        def _activate(e=None):
            self._activate_sample(sample_name)
        frame.bind("<Button-1>",    _activate, add="+")
        tb_frame.bind("<Button-1>", _activate, add="+")

        canvas.mpl_connect("scroll_event",         lambda ev: self._on_scroll(ev, sample_name))
        canvas.mpl_connect("button_press_event",   lambda ev: self._on_press(ev, sample_name))
        canvas.mpl_connect("button_release_event", lambda ev: self._on_release(ev, sample_name))
        canvas.mpl_connect("motion_notify_event",  lambda ev: self._on_motion(ev, sample_name))
        canvas.mpl_connect("button_press_event",
                           lambda ev, sn=sample_name: self._on_legend_dblclick(ev, sn))

        sentry.update({
            "fig": fig, "ax": ax, "canvas": canvas, "toolbar": _tb,
            "plot_frame": frame, "hdr_frame": header, "hdr_label": lbl,
            "legend": None, "leg_size": 8.0,
            "auto_xlim": None, "auto_ylim": None,
            "panning": False, "pan_ax": None, "pan_start": None, "pan_moved": False,
            "leg_resize": False, "leg_resize_start_y": None, "leg_resize_start_sz": None,
            "ann": None, "ann_dot": None, "ann_last": None, "ann_idx": 0,
            "_plot_data": [],
        })
        ax.set_title("", fontsize=9)
        ax.set_xlabel("E (V vs RHE)")
        ax.set_ylabel("I (mA)")
        canvas.draw()
        self._relayout_figures()

    def _destroy_sample_figure(self, sample_name):
        frame = self.samples.get(sample_name, {}).get("plot_frame")
        if frame is not None:
            frame.destroy()

    def _activate_sample(self, sample_name):
        items = list(self.samples.keys())
        if sample_name in items:
            idx = items.index(sample_name)
            self.sample_lb.selection_clear(0, tk.END)
            self.sample_lb.selection_set(idx)
            self.sample_lb.see(idx)
            self._scroll_left_to_widget(self.sample_lb)
        if sample_name != self.active_sample:
            self._save_active_sample_state()
            self._switch_active_sample(sample_name)

    def _scroll_left_to_widget(self, widget):
        """Scroll the left panel so that widget is visible."""
        try:
            self.update_idletasks()
            y = 0
            w = widget
            while w is not None and w is not self._left_inner:
                y += w.winfo_y()
                w = w.master
            inner_h = self._left_inner.winfo_reqheight()
            if inner_h > 0:
                self._left_lc.yview_moveto(max(0.0, (y - 10) / inner_h))
        except Exception:
            pass

    def _highlight_active_headers(self):
        for sn, sentry in self.samples.items():
            hdr = sentry.get("hdr_frame")
            lbl = sentry.get("hdr_label")
            if hdr is None:
                continue
            color = (_SAMPLE_HDR_ACTIVE if sn == self.active_sample
                     else _SAMPLE_HDR_BG)
            hdr.configure(bg=color)
            if lbl:
                lbl.configure(bg=color)

    # ════════════════════════════════════════════════════════════════
    # Right panel layout
    # ════════════════════════════════════════════════════════════════
    def _on_right_canvas_configure(self, event):
        if self._zoom_sample:
            self._right_canvas.itemconfig(
                self._plots_win, width=event.width, height=event.height)

    def _on_grid_cols_change(self):
        try:
            c = max(1, int(self._grid_cols_var.get()))
            self._grid_cols_var.set(str(c))
        except (ValueError, AttributeError):
            self._grid_cols_var.set("2")
        self._relayout_figures()

    def _relayout_figures(self):
        try:
            MAX_COLS = max(1, int(self._grid_cols_var.get()))
        except (ValueError, AttributeError):
            MAX_COLS = 2
        valid = [(sn, self.samples[sn]) for sn in self.samples
                 if "plot_frame" in self.samples[sn]
                 and not self.samples[sn].get("hidden", False)]

        if self._zoom_sample and any(sn == self._zoom_sample for sn, _ in valid):
            for sn, sentry in valid:
                if sn == self._zoom_sample:
                    sentry["plot_frame"].grid(row=0, column=0,
                                             columnspan=MAX_COLS,
                                             sticky="nsew", padx=4, pady=4)
                else:
                    sentry["plot_frame"].grid_remove()
            self._plots_frame.rowconfigure(0, weight=1)
            return

        for sn in self.samples:
            pf = self.samples[sn].get("plot_frame")
            if pf is not None:
                pf.grid_remove()
        for i, (sname, sentry) in enumerate(valid):
            row = i // MAX_COLS
            col = i % MAX_COLS
            sentry["plot_frame"].grid(row=row, column=col, columnspan=1,
                                     sticky="nsew", padx=4, pady=4)
        n_rows = (len(valid) + MAX_COLS - 1) // MAX_COLS if valid else 0
        for r in range(n_rows):
            self._plots_frame.rowconfigure(r, weight=0)

    def _apply_plot_size(self, event=None):
        try:
            w = float(self.plot_w_var.get())
            h = float(self.plot_h_var.get())
        except ValueError:
            return
        w = max(2.0, min(50.0, w))
        h = max(1.5, min(50.0, h))
        dpi = 100
        self._right_canvas.itemconfig(self._plots_win, width=0, height=0)
        for sentry in self.samples.values():
            fig = sentry.get("fig")
            cv  = sentry.get("canvas")
            if fig and cv:
                fig.set_size_inches(w, h)
                cv.get_tk_widget().config(width=int(w * dpi), height=int(h * dpi))
                _lgs = [a.get_legend() for a in fig.get_axes() if a.get_legend()]
                for _l in _lgs:
                    _l.set_visible(False)
                fig.tight_layout(pad=0.5)
                fig.set_layout_engine("none")
                for _l in _lgs:
                    _l.set_visible(True)
                cv.draw_idle()

        def _upd():
            self._plots_frame.update_idletasks()
            self._right_canvas.configure(scrollregion=self._right_canvas.bbox("all"))
        self._right_canvas.after(100, _upd)

    def _auto_set_initial_size(self):
        w = self._right_canvas.winfo_width()
        h = self._right_canvas.winfo_height()
        if w <= 1 or h <= 1:
            self.after(100, self._auto_set_initial_size)
            return
        try:
            _ncols = max(1, int(self._grid_cols_var.get()))
        except (ValueError, AttributeError):
            _ncols = 2
        plot_w = max(3.0, (w / _ncols - 30) / 100)
        plot_h = max(2.0, round(plot_w * 0.6, 1))
        self.plot_w_var.set(f"{plot_w:.1f}")
        self.plot_h_var.set(f"{plot_h:.1f}")
        self._apply_plot_size()

    # ════════════════════════════════════════════════════════════════
    # Zoom (single-sample full-panel view)
    # ════════════════════════════════════════════════════════════════
    def _toggle_zoom(self, sample_name):
        if self._zoom_sample is None:
            self._zoom_sample_view(sample_name)
        else:
            self._unzoom_sample_view()

    def _zoom_sample_view(self, sample_name):
        self._zoom_sample = sample_name
        self._zoom_bar.grid()
        self.update_idletasks()
        w = self._right_canvas.winfo_width()
        h = self._right_canvas.winfo_height()
        if w > 1 and h > 1:
            self._right_canvas.itemconfig(self._plots_win, width=w, height=h)
            sentry = self.samples.get(sample_name, {})
            fig = sentry.get("fig")
            cv  = sentry.get("canvas")
            if fig and cv:
                fig.set_size_inches(w / 100, h / 100)
                cv.get_tk_widget().config(width=w, height=h)
                cv.draw_idle()
        self._relayout_figures()
        self._right_canvas.yview_moveto(0)

    def _unzoom_sample_view(self):
        self._zoom_sample = None
        self._zoom_bar.grid_remove()
        self._right_canvas.itemconfig(self._plots_win, width=0, height=0)
        self._apply_plot_size()
        self._relayout_figures()
        self._plots_frame.update_idletasks()
        self._right_canvas.configure(scrollregion=self._right_canvas.bbox("all"))

    # ════════════════════════════════════════════════════════════════
    # Drag-to-reorder subplots
    # ════════════════════════════════════════════════════════════════
    def _on_frame_press(self, event, sample_name):
        self._drag = {
            "sample": sample_name,
            "start_x": event.x_root, "start_y": event.y_root,
            "active": False, "target": None, "target_top": True,
        }

    def _on_frame_drag(self, event, sample_name):
        drag = self._drag
        if drag is None or drag["sample"] != sample_name:
            return
        if not drag["active"]:
            if abs(event.x_root - drag["start_x"]) + abs(event.y_root - drag["start_y"]) < 6:
                return
            drag["active"] = True
        target = None; target_top = True
        for sn, sentry in self.samples.items():
            if sn == sample_name or sentry.get("hidden"):
                continue
            pf = sentry.get("plot_frame")
            if pf is None:
                continue
            x0, y0 = pf.winfo_rootx(), pf.winfo_rooty()
            w_,  h_ = pf.winfo_width(), pf.winfo_height()
            if x0 <= event.x_root <= x0 + w_ and y0 <= event.y_root <= y0 + h_:
                target = sn
                target_top = (event.y_root - y0) < h_ / 2
                break
        drag["target"] = target; drag["target_top"] = target_top
        if target is not None:
            pf = self.samples[target]["plot_frame"]
            rx, ry = pf.winfo_x(), pf.winfo_y()
            rw, rh = pf.winfo_width(), pf.winfo_height()
            line_y = ry if target_top else ry + rh - 3
            self._drop_line.place(x=rx, y=line_y, width=rw, height=3)
            self._drop_line.lift()
        else:
            self._drop_line.place_forget()

    def _on_frame_release(self, event, sample_name):
        drag = self._drag; self._drag = None
        self._drop_line.place_forget()
        if drag is None or not drag["active"]:
            return
        target = drag.get("target")
        if target is None or target == sample_name:
            return
        keys = list(self.samples.keys())
        if sample_name not in keys or target not in keys:
            return
        keys.remove(sample_name)
        to_idx = keys.index(target)
        keys.insert(to_idx if drag.get("target_top", True) else to_idx + 1, sample_name)
        self.samples = OrderedDict((k, self.samples[k]) for k in keys)
        self._rebuild_sample_listbox()
        self._relayout_figures()

    # ════════════════════════════════════════════════════════════════
    # Mouse interactions (scroll / pan / annotate)
    # ════════════════════════════════════════════════════════════════
    # Listbox drag-to-resize handlers
    # ════════════════════════════════════════════════════════════════
    def _on_loaded_resize_start(self, event):
        self._loaded_drag = {
            "y0": event.y_root,
            "h0": int(self.loaded_tv.cget("height")),
        }

    def _on_loaded_resize_drag(self, event):
        d = getattr(self, "_loaded_drag", None)
        if d is None:
            return
        dy = event.y_root - d["y0"]
        new_rows = max(2, int(d["h0"] + dy / 20))
        self.loaded_tv.config(height=new_rows)

    def _on_sample_resize_start(self, event):
        h = self.sample_lb._canvas.winfo_height()
        if h <= 1:
            h = self.sample_lb._canvas.cget("height")
        self._sample_drag = {"y0": event.y_root, "h0": int(h)}

    def _on_sample_resize_drag(self, event):
        d = getattr(self, "_sample_drag", None)
        if d is None:
            return
        dy = event.y_root - d["y0"]
        new_h = max(40, d["h0"] + dy)
        self.sample_lb._canvas.config(height=int(new_h))

    def _on_pair_resize_start(self, event):
        h = self._pair_tbl_canvas.winfo_height()
        if h <= 1:
            h = int(self._pair_tbl_canvas.cget("height"))
        self._pair_drag = {"y0": event.y_root, "h0": h}

    def _on_pair_resize_drag(self, event):
        d = getattr(self, "_pair_drag", None)
        if d is None:
            return
        dy = event.y_root - d["y0"]
        new_h = max(40, d["h0"] + dy)
        self._pair_tbl_canvas.config(height=int(new_h))

    # ════════════════════════════════════════════════════════════════
    def _on_scroll(self, event, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None or "ax" not in sentry:
            return
        ax = sentry["ax"]
        canvas = sentry["canvas"]
        if event.xdata is None or event.ydata is None:
            return
        factor = (1 / 1.15) if event.button == "up" else 1.15
        cx, cy = event.xdata, event.ydata
        xl = ax.get_xlim(); yl = ax.get_ylim()
        ax.set_xlim(cx + (xl[0] - cx) * factor, cx + (xl[1] - cx) * factor)
        ax.set_ylim(cy + (yl[0] - cy) * factor, cy + (yl[1] - cy) * factor)
        canvas.draw_idle()

    def _on_press(self, event, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return
        self._activate_sample(sample_name)
        sentry["pan_moved"] = False

        # Double-click on axis labels → inline edit
        if event.button == 1 and getattr(event, "dblclick", False):
            ax = sentry.get("ax")
            if ax is not None:
                try:
                    renderer = event.canvas.get_renderer()
                    xl = ax.xaxis.label
                    if xl.get_window_extent(renderer).contains(event.x, event.y):
                        self._edit_axis_label(sample_name, "x")
                        return
                    yl = ax.yaxis.label
                    if yl.get_window_extent(renderer).contains(event.x, event.y):
                        self._edit_axis_label(sample_name, "y")
                        return
                except Exception:
                    pass

        if event.button == 3:
            # Right-drag on legend → resize; right-click elsewhere → dismiss annotation
            leg = sentry.get("legend")
            if leg is not None:
                try:
                    hit, _ = leg.contains(event)
                    if hit:
                        sentry["leg_resize"] = True
                        sentry["leg_resize_start_y"]  = event.y
                        sentry["leg_resize_start_sz"] = sentry.get(
                            "leg_size_live", sentry.get("leg_size", 8.0))
                        return
                except Exception:
                    pass
            sentry["highlighted_cat"] = None
            self._apply_orr_highlight(sample_name)
            self._clear_ann(sample_name, redraw=True)
            return
        if event.button != 1 or event.xdata is None:
            return
        # Don't start panning if the click landed on the legend — let DraggableLegend handle it.
        leg = sentry.get("legend")
        if leg is not None:
            try:
                hit, _ = leg.contains(event)
                if hit:
                    return
            except Exception:
                pass
        sentry["panning"]   = True
        sentry["pan_start"] = (event.xdata, event.ydata,
                               *sentry["ax"].get_xlim(), *sentry["ax"].get_ylim())

    def _edit_axis_label(self, sample_name, which):
        """Double-click X or Y label → askstring dialog; blank reverts to auto."""
        from tkinter.simpledialog import askstring
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return
        ax = sentry.get("ax")
        if ax is None:
            return
        if which == "x":
            current = ax.get_xlabel()
            new = askstring("Edit X Label", "X axis label\n(blank = auto):",
                            initialvalue=current, parent=self)
            if new is not None:
                sentry["custom_xlabel"] = new.strip() or None
                ax.set_xlabel(new.strip() if new.strip() else current)
                sentry["canvas"].draw_idle()
        else:
            current = ax.get_ylabel()
            new = askstring("Edit Y Label", "Y axis label\n(blank = auto):",
                            initialvalue=current, parent=self)
            if new is not None:
                sentry["custom_ylabel"] = new.strip() or None
                ax.set_ylabel(new.strip() if new.strip() else current)
                sentry["canvas"].draw_idle()

    def _on_motion(self, event, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return

        # Legend resize (right-drag on legend)
        if sentry.get("leg_resize"):
            leg = sentry.get("legend")
            if leg is not None and event.y is not None:
                dy = event.y - sentry["leg_resize_start_y"]
                new_sz = max(4.0, min(30.0,
                             sentry["leg_resize_start_sz"] + dy / 5.0))
                sentry["leg_size_live"] = new_sz
                prev_sz = sentry.get("_leg_prev_sz", sentry.get("leg_size", 8.0))
                sentry["_leg_prev_sz"] = new_sz
                try:
                    if prev_sz > 0:
                        _scale_legend_spacing(leg, new_sz / prev_sz)
                except Exception:
                    pass
                for txt in leg.get_texts():
                    txt.set_fontsize(new_sz)
                ttl = leg.get_title()
                if ttl:
                    ttl.set_fontsize(new_sz)
                sentry["canvas"].draw_idle()
            return

        if not sentry.get("panning"):
            return
        if event.xdata is None or sentry.get("pan_start") is None:
            return
        sentry["pan_moved"] = True
        x0, y0, x1, x2, y1, y2 = sentry["pan_start"]
        dx = event.xdata - x0; dy = event.ydata - y0
        ax = sentry["ax"]
        ax.set_xlim(x1 - dx, x2 - dx)
        ax.set_ylim(y1 - dy, y2 - dy)
        sentry["canvas"].draw_idle()

    def _on_release(self, event, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return

        # Legend resize release → commit final size
        if sentry.get("leg_resize"):
            sentry["leg_resize"] = False
            live = sentry.get("leg_size_live")
            if live is not None:
                sentry["leg_size"] = live
                sentry["leg_size_live"] = None
            return

        sentry["panning"] = False
        if event.button == 1 and not sentry.get("pan_moved"):
            self._try_annotate(event, sample_name)
        sentry["pan_moved"] = False
        # Deferred legend-position save: fires after matplotlib's DraggableLegend
        # finalize_offset() has completed, so _loc / _loc_real reflects new position.
        if event.button == 1:
            self.after(10, lambda s=sample_name: self._save_leg_pos(s))

    def _save_leg_pos(self, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return
        leg = sentry.get("legend")
        if leg is None:
            return
        for attr in ("_loc_real", "_loc"):
            _loc = getattr(leg, attr, None)
            if isinstance(_loc, (tuple, list)) and len(_loc) == 2:
                sentry["legend_manual_pos"] = tuple(_loc)
                return

    def _try_annotate(self, event, sample_name):
        sentry = self.samples.get(sample_name)
        if sentry is None or event.xdata is None:
            return
        plot_data = sentry.get("_plot_data", [])
        if not plot_data:
            return
        ax = sentry["ax"]; canvas = sentry["canvas"]
        # Find nearest point across all curves
        best_dist = float("inf"); best_x = best_y = None; best_label = ""
        xl = ax.get_xlim(); yl = ax.get_ylim()
        xspan = abs(xl[1] - xl[0]) or 1; yspan = abs(yl[1] - yl[0]) or 1
        for E_arr, Y_arr, label, color in plot_data:
            if len(E_arr) == 0:
                continue
            dx = (E_arr - event.xdata) / xspan
            dy = (Y_arr - event.ydata) / yspan
            dist = dx**2 + dy**2
            idx  = int(np.argmin(dist))
            if dist[idx] < best_dist:
                best_dist = dist[idx]; best_x = E_arr[idx]; best_y = Y_arr[idx]
                best_label = label
        if best_x is None or best_dist > 0.04:
            return
        # Switch catalyst correction display + highlight all lines for that catalyst
        _cat_m = re.match(r'^\[([^\]]+)\]', best_label)
        _clicked_cat = _cat_m.group(1) if _cat_m else ""
        sentry["highlighted_cat"] = _clicked_cat
        self._apply_orr_highlight(sample_name)
        if _cat_m:
            _avail = list(self._corr_cat_cb["values"])
            if _clicked_cat in _avail and _clicked_cat != getattr(self, "_active_catalyst", None):
                self._corr_catalyst_var.set(_clicked_cat)
                self._load_catalyst_corrections(sample_name, _clicked_cat)
        # Increment annotation index for overlap cycling
        ann_key = (round(best_x, 6), round(best_y, 6))
        if sentry.get("ann_last") == ann_key:
            sentry["ann_idx"] = (sentry.get("ann_idx", 0) + 1) % 4
        else:
            sentry["ann_idx"] = 0; sentry["ann_last"] = ann_key
        self._clear_ann(sample_name, redraw=False)
        offsets = [(8, 8), (-8, 8), (-8, -8), (8, -8)]
        xytext  = offsets[sentry["ann_idx"] % 4]
        ann = ax.annotate(
            f"{best_label}\n({best_x:.4f}, {best_y:.4f})",
            xy=(best_x, best_y), xytext=xytext,
            textcoords="offset points", fontsize=8,
            bbox=dict(boxstyle="round,pad=0.3", fc="lightyellow", alpha=0.85),
            arrowprops=dict(arrowstyle="->", lw=0.8),
        )
        ann.set_in_layout(False)
        dot, = ax.plot([best_x], [best_y], "ko", ms=5,
                       label=_ANN_DOT_LABEL, zorder=10)
        sentry["ann"] = ann; sentry["ann_dot"] = dot
        canvas.draw_idle()

    def _apply_orr_highlight(self, sample_name):
        """Dim all catalyst lines except the highlighted one; bring highlighted to front."""
        sentry = self.samples.get(sample_name)
        if sentry is None or "ax" not in sentry:
            return
        ax = sentry["ax"]
        hl_cat = sentry.get("highlighted_cat")  # None = no highlight; str = catalyst name
        for line in ax.get_lines():
            lbl = line.get_label()
            if lbl.startswith("_"):
                continue  # skip ehalf, annotation dot, glow lines
            if hl_cat is None:
                line.set_alpha(1.0)
                line.set_zorder(2)
            else:
                m = re.match(r'^\[([^\]]+)\]', lbl)
                line_cat = m.group(1) if m else ""
                if line_cat == hl_cat:
                    line.set_alpha(1.0)
                    line.set_zorder(5)
                else:
                    line.set_alpha(0.15)
                    line.set_zorder(2)

    def _clear_ann(self, sample_name, *, redraw=True):
        sentry = self.samples.get(sample_name)
        if sentry is None:
            return
        for key in ("ann", "ann_dot"):
            obj = sentry.get(key)
            if obj is not None:
                try:
                    obj.remove()
                except Exception:
                    pass
            sentry[key] = None
        if redraw:
            cv = sentry.get("canvas")
            if cv:
                cv.draw_idle()

    def _on_legend_dblclick(self, event, sample_name):
        if event.dblclick and event.button == 1:
            sentry = self.samples.get(sample_name)
            if sentry is None:
                return
            leg = sentry.get("legend")
            if leg is None:
                return
            try:
                hit, _ = leg.contains(event)
            except Exception:
                hit = False
            if not hit:
                return
            leg.set_draggable(False)
            sentry["legend"], perm = open_legend_editor(
                self, leg, sentry["canvas"], sentry.get("leg_size", 8.0))
            if sentry.get("legend") is not None:
                sentry["legend"].set_draggable(True)
            # Store permutation so next replot respects the new order
            orig_labels = sentry.get("legend_label_order", [])
            if perm and orig_labels:
                sentry["legend_label_order"] = [orig_labels[j] for j in perm
                                                if j < len(orig_labels)]
            sentry["canvas"].draw()

    # ════════════════════════════════════════════════════════════════
    # Analysis windows
    # ════════════════════════════════════════════════════════════════
    @staticmethod
    def _gradient_shades(base_hex, n):
        """Return n shades of base_hex ordered dark→light (vary HSL lightness)."""
        try:
            r, g, b = _mcolors.to_rgb(base_hex)
        except Exception:
            return [base_hex] * max(n, 1)
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        if n <= 1:
            return [_mcolors.to_hex((r, g, b))]
        out = []
        for j in range(n):
            t = j / max(n - 1, 1)
            l_new = max(0.22, min(0.78, l - 0.22 + 0.46 * t))
            s_new = max(0.0, min(1.0, s + (0.10 if t < 0.4 else 0.0)))
            r2, g2, b2 = colorsys.hls_to_rgb(h, l_new, s_new)
            out.append(_mcolors.to_hex((r2, g2, b2)))
        return out

    @classmethod
    def _gradient_recolor(cls, curves, *, label_idx=3, color_idx=4,
                          sort_idx=2, sn_idx=5):
        """Re-color a list of curve tuples so that curves in the same
        (sample, catalyst) group share a base color and vary by lightness
        along *sort_idx* (usually RPM).

        Curves with a unique (sample, catalyst) group keep their original colour.
        Tuples must have length > max(label_idx, color_idx, sort_idx, sn_idx).
        Returns a new list of tuples with updated colour at *color_idx*.
        """
        groups = {}
        for i, c in enumerate(curves):
            if not isinstance(c, tuple) or len(c) <= max(label_idx, color_idx,
                                                          sort_idx, sn_idx):
                continue
            sn  = c[sn_idx]
            lbl = c[label_idx]
            m   = re.match(r'^\[([^\]]+)\]', lbl) if isinstance(lbl, str) else None
            cat = m.group(1) if m else ""
            groups.setdefault((sn, cat), []).append(i)

        out = [list(c) for c in curves]
        for key, idxs in groups.items():
            n = len(idxs)
            if n <= 1:
                continue
            try:
                idxs.sort(key=lambda i: float(curves[i][sort_idx] or 0))
            except (TypeError, ValueError):
                pass
            base   = curves[idxs[0]][color_idx]
            shades = cls._gradient_shades(base, n)
            for j, i in enumerate(idxs):
                out[i][color_idx] = shades[j]
        return [tuple(c) for c in out]

    def _build_curve_records(self, sname):
        """Return the fully-processed, background-subtracted curves for a sample.

        Single source of truth shared by the plot (`_get_curves_for_sample`) and
        the Excel export. Only enabled pairs are included; per-catalyst
        corrections (r_sol, e_ref, area, ECSA_Hupd) are applied via
        `_process_pair`; pairs are ordered per catalyst by ascending RPM.

        Each record is a dict:
            E, J          np.ndarray  (J already ÷ area when area > 0)
            rpm           float       (numeric, for sorting/plotting)
            rpm_v         str         (raw RPM label as shown)
            label, color              (plot label / gradient colour)
            catalyst                  (catalyst_id)
            r_n2, r_o2, e_ref, area   float  (corrections actually applied)
            ecsa                      (ECSA_Hupd, raw string/'' if unset)
        """
        sentry = self.samples.get(sname)
        if sentry is None:
            return []
        cat_corrections = sentry.get("catalyst_corrections", {})

        # Build per-catalyst ordered pair lists (enabled only)
        _cat_order: list = []
        _pairs_by_cat: dict = {}
        for pair in sentry.get("pairs", []):
            if not pair.get("n2_short") or not pair.get("o2_short"):
                continue
            if not pair.get("enabled", True):
                continue
            cat = pair.get("catalyst_id", "")
            if cat not in _cat_order:
                _cat_order.append(cat)
                _pairs_by_cat[cat] = []
            _pairs_by_cat[cat].append(pair)

        try:
            _gstep = float(sentry.get("gradient_step", 0.10))
        except (ValueError, TypeError):
            _gstep = 0.10

        records = []
        for ci, cat in enumerate(_cat_order):
            _cat_st   = sentry.get("catalyst_styles", {}).get(cat, {})
            _user_col = _cat_st.get("color", "").strip()
            base_col  = _user_col if _user_col else _PALETTE[ci % len(_PALETTE)]
            cat_pairs = sorted(_pairs_by_cat[cat],
                               key=lambda p: _safe_rpm_int(
                                   p.get("rpm_val") or p.get("rpm_id", "")))
            n_cat = len(cat_pairs)
            cat_colors = (_cycle_colors(base_col, n_cat, step=_gstep, reverse=False)
                          if n_cat > 1 else [base_col])
            _cc = cat_corrections.get(cat, {})
            try:
                r_n2  = float(_cc.get("r_sol_n2", 0) or 0)
                r_o2  = float(_cc.get("r_sol_o2", 0) or 0)
                e_ref = float(_cc.get("e_ref", 0) or 0)
                area  = float(_cc.get("area", "") or 0)
            except (ValueError, TypeError):
                r_n2 = r_o2 = e_ref = area = 0.0
            ecsa = _cc.get("ecsa", "")
            for j, pair in enumerate(cat_pairs):
                result = _process_pair(pair, r_n2, r_o2, e_ref, area)
                if result is None:
                    continue
                E_arr, J_arr = result
                try:
                    rpm = float(pair.get("rpm_val") or pair.get("rpm_id") or 0)
                except (ValueError, TypeError):
                    rpm = 0.0
                prefix = f"[{cat}] " if cat else ""
                rpm_v  = pair.get("rpm_val") or pair.get("rpm_id") or f"#{j+1}"
                label  = f"{prefix}{rpm_v} rpm"
                records.append({
                    "E": E_arr, "J": J_arr, "rpm": rpm, "rpm_v": rpm_v,
                    "label": label, "color": cat_colors[j], "catalyst": cat,
                    "r_n2": r_n2, "r_o2": r_o2, "e_ref": e_ref, "area": area,
                    "ecsa": ecsa,
                })
        return records

    def _get_curves_for_sample(self, sname):
        """Return list of (E_arr, J_arr, rpm_float, label, color) for a named sample."""
        return [(r["E"], r["J"], r["rpm"], r["label"], r["color"])
                for r in self._build_curve_records(sname)]

    def _get_active_curves(self):
        """Return list of (E_arr, J_arr, rpm_float, label, color) for the active sample."""
        if not self.active_sample:
            return []
        return self._get_curves_for_sample(self.active_sample)

    def _open_tafel_window(self):
        # Gather from ALL loaded samples
        # (curves recoloured to share base colour per sample/catalyst group)
        all_curves = []  # (E_arr, J_arr, rpm, label, color, sname)
        for sn in self.samples:
            for E, J, rpm, lbl, col in self._get_curves_for_sample(sn):
                all_curves.append((E, J, rpm, lbl, col, sn))
        if not all_curves:
            messagebox.showwarning("Tafel", "No processed curves in any sample.")
            return
        all_curves = self._gradient_recolor(all_curves)
        ref = self.ref_electrode_var.get()

        win = tk.Toplevel(self)
        win.title("Tafel Analysis")
        win.geometry("820x720")
        try: win.state('zoomed')
        except Exception: pass

        # Debounce
        _recompute_id = [None]
        def _schedule(*_):
            if _recompute_id[0]:
                try: win.after_cancel(_recompute_id[0])
                except Exception: pass
            _recompute_id[0] = win.after(350, _compute)

        # ── Theory ──────────────────────────────────────────────────────
        _th = ttk.LabelFrame(win, text="Tafel Theory")
        _th.pack(fill=tk.X, padx=8, pady=(6, 0))
        _th_g = ttk.Frame(_th); _th_g.pack(fill=tk.X, padx=6, pady=3)
        for _r, (_hd, _bd) in enumerate([
            ("Tafel equation:",       "E = a + b · log₁₀|J|   (b = Tafel slope, mV/dec)"),
            ("Diffusion correction:", "Jᵏ = J · J_lim / (J_lim − J)   (Koutecky, optional)")
        ]):
            ttk.Label(_th_g, text=_hd, font=("TkDefaultFont", 9, "bold"),
                      anchor=tk.W).grid(row=_r, column=0, sticky=tk.W, pady=2)
            ttk.Label(_th_g, text=f"  {_bd}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=1, sticky=tk.W, padx=(4,0), pady=2)

        # ── Notation — 2-column grid ──────────────────────────────────
        _sym = ttk.LabelFrame(win, text="Notation")
        _sym.pack(fill=tk.X, padx=8, pady=(0, 2))
        _sym_g = ttk.Frame(_sym); _sym_g.pack(fill=tk.X, padx=4, pady=3)
        for _i, (_s, _d) in enumerate([
            ("J",     "current density (mA cm⁻²)"),
            ("Jᵏ",    "kinetic J = J · J_lim / (J_lim − J)"),
            ("J_lim", "diffusion plateau (most negative J)"),
            ("b",     "Tafel slope (mV/decade)"),
            ("E",     "electrode potential vs reference"),
        ]):
            _r, _c = _i // 2, (_i % 2) * 2
            ttk.Label(_sym_g, text=_s, font=("TkDefaultFont", 9, "bold"),
                      width=7, anchor=tk.E).grid(row=_r, column=_c, sticky=tk.E, padx=(8,2), pady=2)
            ttk.Label(_sym_g, text=f"= {_d}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=_c+1, sticky=tk.W, padx=(0,16), pady=2)

        # ── Curve selector — Group label (not interactive) → Sample [cat] checkbox → RPM ──
        _ksel_fr = ttk.LabelFrame(
            win, text="Select curves  (☑ [Cat] = one sample = one catalyst at multiple RPMs)")
        _ksel_fr.pack(fill=tk.X, padx=8, pady=(4, 0))
        _sel_cv = tk.Canvas(_ksel_fr, height=80, bd=0, highlightthickness=0)
        _sel_sb = ttk.Scrollbar(_ksel_fr, orient=tk.VERTICAL, command=_sel_cv.yview)
        _sel_inner = tk.Frame(_sel_cv)
        _sel_inner.bind("<Configure>",
                        lambda e: _sel_cv.configure(scrollregion=_sel_cv.bbox("all")))
        _sel_cv.create_window((0, 0), window=_sel_inner, anchor="nw")
        _sel_cv.configure(yscrollcommand=_sel_sb.set)
        _sel_sb.pack(side=tk.RIGHT, fill=tk.Y)
        _sel_cv.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        _sel_cv.bind("<MouseWheel>",
                     lambda e: _sel_cv.yview_scroll(-1*(e.delta//120), "units"))

        _by_samp = {}; _samp_order = []
        for idx, (E_a, J_a, rpm, lbl, col, sn) in enumerate(all_curves):
            if sn not in _by_samp:
                _by_samp[sn] = []; _samp_order.append(sn)
            m = re.match(r'^\[([^\]]+)\]', lbl)
            cat = m.group(1) if m else ""
            _by_samp[sn].append((idx, E_a, J_a, rpm, lbl, col, cat))

        _tsel_vars = {}  # idx → BooleanVar

        for sn in _samp_order:
            # Group label — informational, not a sample in user's model
            tk.Label(_sel_inner, text=f"  ▸ Group: {sn}",
                     font=("TkDefaultFont", 8, "italic"), fg="#555555",
                     anchor=tk.W).pack(fill=tk.X, anchor=tk.W, pady=(4, 0))
            _by_cat = {}; _cat_ord = []
            for idx, E_a, J_a, rpm, lbl, col, cat in _by_samp[sn]:
                if cat not in _by_cat:
                    _by_cat[cat] = []; _cat_ord.append(cat)
                _by_cat[cat].append((idx, rpm, lbl, col))
            for cat in _cat_ord:
                cat_idxs = [it[0] for it in _by_cat[cat]]
                cat_bv = tk.BooleanVar(value=True)
                cat_row = tk.Frame(_sel_inner)
                cat_row.pack(fill=tk.X, anchor=tk.W, padx=(18, 0), pady=1)
                # Sample-level checkbox = one catalyst = one sample (user's definition)
                tk.Checkbutton(
                    cat_row,
                    text=f"[{cat}]" if cat else "(no cat)",
                    variable=cat_bv,
                    command=lambda idxs=cat_idxs, bv=cat_bv:
                        [_tsel_vars[i].set(bv.get()) for i in idxs],
                    font=("TkDefaultFont", 8, "bold")
                ).pack(side=tk.LEFT, padx=(0, 8))
                for idx, rpm, lbl, col in _by_cat[cat]:
                    bv = tk.BooleanVar(value=True)
                    _tsel_vars[idx] = bv
                    display = re.sub(r'^\[[^\]]+\]\s*', '',lbl)
                    tk.Checkbutton(cat_row, text=display, variable=bv).pack(
                        side=tk.LEFT, padx=3)
                    bv.trace_add("write", _schedule)

        # ── Controls ──────────────────────────────────────────────────
        ctrl = ttk.Frame(win); ctrl.pack(fill=tk.X, padx=8, pady=(4, 0))
        ttk.Label(ctrl, text="Kinetic E range (V):").pack(side=tk.LEFT)
        e_lo_var = tk.StringVar(value="0.85"); e_hi_var = tk.StringVar(value="0.95")
        ttk.Entry(ctrl, textvariable=e_lo_var, width=6).pack(side=tk.LEFT, padx=(2, 0))
        ttk.Label(ctrl, text="to").pack(side=tk.LEFT, padx=3)
        ttk.Entry(ctrl, textvariable=e_hi_var, width=6).pack(side=tk.LEFT, padx=(0, 10))
        use_jk_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(ctrl, text="Diffusion-correct J → Jᵏ",
                        variable=use_jk_var).pack(side=tk.LEFT, padx=(0, 10))
        e_lo_var.trace_add("write", _schedule)
        e_hi_var.trace_add("write", _schedule)
        use_jk_var.trace_add("write", _schedule)

        # ── Figure — pack bottom items first ──────────────────────────
        fig = Figure(figsize=(7.0, 4.0), dpi=100, constrained_layout=True)
        ax  = fig.add_subplot(111)
        cv  = FigureCanvasTkAgg(fig, master=win)
        _tb_row = ttk.Frame(win)
        tb  = NavigationToolbar2Tk(cv, _tb_row, pack_toolbar=False)
        tb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tb_row, text="Copy",
                  command=lambda f=fig: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        _passist = attach_plot_assistant(win, fig, ax, cv)
        _passist.pack(side=tk.TOP, fill=tk.X, padx=8, pady=(2, 0))
        res = tk.Text(win, height=5, state=tk.DISABLED, font=("Courier", 9), wrap=tk.WORD)
        _tres_handle = tk.Frame(win, height=6, bg="#888888", cursor="sb_v_double_arrow")
        _tb_row.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        res.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        _tres_handle.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(1, 1))
        cv.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8)
        _tres_y0 = [0]; _tres_h0 = [5]
        def _tres_start(e): _tres_y0[0] = e.y_root; _tres_h0[0] = int(res.cget("height"))
        def _tres_drag(e):
            res.configure(height=max(2, _tres_h0[0] - int((e.y_root - _tres_y0[0]) / 16)))
        _tres_handle.bind("<ButtonPress-1>", _tres_start)
        _tres_handle.bind("<B1-Motion>", _tres_drag)

        def _compute():
            try:
                e_lo = float(e_lo_var.get()); e_hi = float(e_hi_var.get())
            except ValueError:
                return
            if e_lo >= e_hi:
                return
            ax.clear(); lines = []
            for idx, (E_arr, J_arr, rpm, label, color, sn) in enumerate(all_curves):
                if not _tsel_vars.get(idx, tk.BooleanVar(value=True)).get():
                    continue
                j_lim = float(np.min(J_arr))
                mask = (E_arr >= e_lo) & (E_arr <= e_hi)
                if mask.sum() < 3:
                    lines.append(f"{label}: < 3 pts in [{e_lo},{e_hi}] V — skipped")
                    continue
                E_k = E_arr[mask]; J_k = J_arr[mask]
                if use_jk_var.get() and j_lim < 0:
                    with np.errstate(divide="ignore", invalid="ignore"):
                        jk = J_k * j_lim / (j_lim - J_k)
                    good = np.isfinite(jk) & (jk < 0)
                    if good.sum() < 3:
                        lines.append(f"{label}: Jᵏ < 3 pts — skipped")
                        continue
                    E_k, J_k = E_k[good], jk[good]
                with np.errstate(divide="ignore", invalid="ignore"):
                    log_j = np.log10(np.abs(J_k))
                good = np.isfinite(log_j)
                if good.sum() < 3:
                    continue
                E_f, log_j_f = E_k[good], log_j[good]
                coeffs = np.polyfit(log_j_f, E_f, 1)
                b_mV = coeffs[0] * 1000.0
                ax.plot(log_j_f, E_f, color=color, linewidth=1.5, label=label)
                xfit = np.linspace(log_j_f.min(), log_j_f.max(), 60)
                ax.plot(xfit, np.polyval(coeffs, xfit), color=color,
                        linestyle="--", linewidth=0.8, label="_fit")
                lines.append(f"{label:32s}  b = {b_mV:+.1f} mV/dec")
            j_lbl = "Jᵏ" if use_jk_var.get() else "J"
            ax.set_xlabel(f"log₁₀|{j_lbl}|  (mA cm⁻² or mA)")
            ax.set_ylabel(f"E  (V vs {ref})")
            ax.set_title("Tafel Analysis")
            ax.legend(fontsize=7, frameon=True)
            cv.draw()
            res.configure(state=tk.NORMAL)
            res.delete("1.0", tk.END)
            res.insert(tk.END, "\n".join(lines))
            res.configure(state=tk.DISABLED)

        _compute()

    def _open_kl_window(self):
        # Gather curves from ALL loaded samples (not just active)
        all_valid = []  # (E_arr, J_arr, rpm, label, color, sname)
        for sn in self.samples:
            for E, J, rpm, lbl, col in self._get_curves_for_sample(sn):
                if rpm > 0:
                    all_valid.append((E, J, rpm, lbl, col, sn))

        if len(all_valid) < 2:
            messagebox.showwarning(
                "KL Analysis",
                "Need at least 2 RPM curves with numeric RPM values.")
            return
        # NOTE: KL plots fitted lines per (group, e_val), not raw curves —
        # so we apply gradient at plot time, not via _gradient_recolor here.

        ref   = self.ref_electrode_var.get()
        sname = self.active_sample or "ORR"

        win = tk.Toplevel(self)
        win.title(f"Koutecky-Levich Analysis — {sname}")
        win.geometry("860x760")
        try: win.state('zoomed')
        except Exception: pass

        # ── Debounced auto-recompute ─────────────────────────────────
        _recompute_id = [None]
        def _schedule(*_):
            if _recompute_id[0]:
                try: win.after_cancel(_recompute_id[0])
                except Exception: pass
            _recompute_id[0] = win.after(350, _compute)

        # ── Curve selector — Group label → Sample [cat] checkbox → RPM ──
        _ksel_fr = ttk.LabelFrame(
            win, text="Select curves  (☑ [Cat] = one sample = one catalyst at multiple RPMs)")
        _ksel_fr.pack(fill=tk.X, padx=8, pady=(6, 0))

        _sel_cv = tk.Canvas(_ksel_fr, height=90, bd=0, highlightthickness=0)
        _sel_sb = ttk.Scrollbar(_ksel_fr, orient=tk.VERTICAL, command=_sel_cv.yview)
        _sel_inner = tk.Frame(_sel_cv)
        _sel_inner.bind("<Configure>",
                        lambda e: _sel_cv.configure(scrollregion=_sel_cv.bbox("all")))
        _sel_cv.create_window((0, 0), window=_sel_inner, anchor="nw")
        _sel_cv.configure(yscrollcommand=_sel_sb.set)
        _sel_sb.pack(side=tk.RIGHT, fill=tk.Y)
        _sel_cv.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        _sel_cv.bind("<MouseWheel>",
                     lambda e: _sel_cv.yview_scroll(-1*(e.delta//120), "units"))

        _by_samp = {}; _samp_order = []
        for idx, (E_a, J_a, rpm, lbl, col, sn) in enumerate(all_valid):
            if sn not in _by_samp:
                _by_samp[sn] = []; _samp_order.append(sn)
            m = re.match(r'^\[([^\]]+)\]', lbl)
            cat = m.group(1) if m else ""
            _by_samp[sn].append((idx, E_a, J_a, rpm, lbl, col, cat))

        _ksel_vars = {}  # idx → BooleanVar

        for sn in _samp_order:
            tk.Label(_sel_inner, text=f"  ▸ Group: {sn}",
                     font=("TkDefaultFont", 8, "italic"), fg="#555555",
                     anchor=tk.W).pack(fill=tk.X, anchor=tk.W, pady=(4, 0))
            _by_cat = {}; _cat_ord = []
            for idx, E_a, J_a, rpm, lbl, col, cat in _by_samp[sn]:
                if cat not in _by_cat:
                    _by_cat[cat] = []; _cat_ord.append(cat)
                _by_cat[cat].append((idx, rpm, lbl, col))
            for cat in _cat_ord:
                cat_idxs = [it[0] for it in _by_cat[cat]]
                cat_bv = tk.BooleanVar(value=True)
                cat_row = tk.Frame(_sel_inner)
                cat_row.pack(fill=tk.X, anchor=tk.W, padx=(18, 0), pady=1)
                tk.Checkbutton(
                    cat_row,
                    text=f"[{cat}]" if cat else "(no cat)",
                    variable=cat_bv,
                    command=lambda idxs=cat_idxs, bv=cat_bv:
                        [_ksel_vars[i].set(bv.get()) for i in idxs],
                    font=("TkDefaultFont", 8, "bold")
                ).pack(side=tk.LEFT, padx=(0, 8))
                for idx, rpm, lbl, col in _by_cat[cat]:
                    bv = tk.BooleanVar(value=True)
                    _ksel_vars[idx] = bv
                    display = re.sub(r'^\[[^\]]+\]\s*', '',lbl)
                    tk.Checkbutton(cat_row, text=display, variable=bv).pack(
                        side=tk.LEFT, padx=3)
                    bv.trace_add("write", _schedule)

        # ── Theory ───────────────────────────────────────────────────
        _th = ttk.LabelFrame(win, text="Koutecky-Levich Theory")
        _th.pack(fill=tk.X, padx=8, pady=(4, 0))
        _th_g = ttk.Frame(_th); _th_g.pack(fill=tk.X, padx=6, pady=3)
        for _r, (_hd, _bd) in enumerate([
            ("KL equation:",   "1/J = 1/Jᵏ + 1/(B·ω^½)"),
            ("B factor:",      "B = 0.62 · n · F · D^(2/3) · ν^(-1/6) · C"),
            ("Electron count:","n = 1 / (|slope| · B_factor)   ω = 2π·RPM/60"),
        ]):
            ttk.Label(_th_g, text=_hd, font=("TkDefaultFont", 9, "bold"),
                      anchor=tk.W).grid(row=_r, column=0, sticky=tk.W, pady=2)
            ttk.Label(_th_g, text=f"  {_bd}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=1, sticky=tk.W, padx=(4,0), pady=2)

        # ── Notation — 2-column grid ──────────────────────────────────
        _sym = ttk.LabelFrame(win, text="Notation")
        _sym.pack(fill=tk.X, padx=8, pady=(0, 2))
        _sym_g = ttk.Frame(_sym); _sym_g.pack(fill=tk.X, padx=4, pady=3)
        for _i, (_s, _d) in enumerate([
            ("J",  "current density (mA cm⁻²)"),
            ("Jᵏ", "kinetic J  (1/Jᵏ = KL y-intercept)"),
            ("n",  "electrons per O₂  (4 = full 4e⁻,  2 = peroxide)"),
            ("ω",  "angular velocity = 2π·RPM/60  (rad s⁻¹)"),
            ("D",  "O₂ diffusion coefficient (cm² s⁻¹)"),
            ("ν",  "kinematic viscosity (cm² s⁻¹)"),
            ("C",  "O₂ solubility (mol cm⁻³)"),
            ("F",  "Faraday constant = 96 485 C mol⁻¹"),
        ]):
            _r, _c = _i // 2, (_i % 2) * 2
            ttk.Label(_sym_g, text=_s, font=("TkDefaultFont", 9, "bold"),
                      width=5, anchor=tk.E).grid(row=_r, column=_c, sticky=tk.E, padx=(8,2), pady=2)
            ttk.Label(_sym_g, text=f"= {_d}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=_c+1, sticky=tk.W, padx=(0,16), pady=2)

        # ── Electrolyte parameters ────────────────────────────────────
        _ELYTE_PRESETS = {
            "0.1 M KOH  (25 °C)":    ("1.90e-5", "1.00e-2", "1.20e-6"),
            "0.5 M KOH  (25 °C)":    ("1.75e-5", "1.05e-2", "1.15e-6"),
            "1.0 M KOH  (25 °C)":    ("1.65e-5", "1.15e-2", "1.10e-6"),
            "0.1 M HClO₄ (25 °C)":   ("1.93e-5", "1.007e-2","1.26e-6"),
            "0.5 M HClO₄ (25 °C)":   ("1.70e-5", "1.06e-2", "1.15e-6"),
            "1.0 M HClO₄ (25 °C)":   ("1.60e-5", "1.07e-2", "1.00e-6"),
            "0.5 M H₂SO₄ (25 °C)":   ("1.40e-5", "1.21e-2", "1.10e-6"),
            "1.0 M H₂SO₄ (25 °C)":   ("1.10e-5", "1.45e-2", "1.00e-6"),
        }
        prm = ttk.LabelFrame(win, text="Electrolyte parameters")
        prm.pack(fill=tk.X, padx=8, pady=(4, 2))
        _pr = ttk.Frame(prm)
        _pr.pack(fill=tk.X, padx=6, pady=3)
        d_var  = tk.StringVar(value="1.90e-5")
        nu_var = tk.StringVar(value="1.00e-2")
        c_var  = tk.StringVar(value="1.20e-6")
        _elyte_var = tk.StringVar(value="0.1 M KOH  (25 °C)")
        ttk.Label(_pr, text="Electrolyte:").pack(side=tk.LEFT)
        _elyte_cb = ttk.Combobox(_pr, textvariable=_elyte_var,
                                  values=list(_ELYTE_PRESETS.keys()),
                                  state="readonly", width=22)
        _elyte_cb.pack(side=tk.LEFT, padx=(4, 12))
        def _on_elyte_change(*_):
            preset = _ELYTE_PRESETS.get(_elyte_var.get())
            if preset:
                d_var.set(preset[0]); nu_var.set(preset[1]); c_var.set(preset[2])
        _elyte_cb.bind("<<ComboboxSelected>>", _on_elyte_change)
        for lbl_txt, var in (
            ("Dₒ₂ (cm²/s):", d_var),
            ("ν (cm²/s):",    nu_var),
            ("Cₒ₂ (mol/cm³):", c_var),
        ):
            ttk.Label(_pr, text=lbl_txt).pack(side=tk.LEFT, padx=(0, 2))
            ttk.Entry(_pr, textvariable=var, width=9).pack(side=tk.LEFT, padx=(0, 8))
        for v in (d_var, nu_var, c_var):
            v.trace_add("write", _schedule)

        # ── E-value controls ─────────────────────────────────────────
        ectrl = ttk.Frame(win)
        ectrl.pack(fill=tk.X, padx=8, pady=(2, 0))
        ttk.Label(ectrl, text="E values  (V vs RHE, comma-sep):").pack(side=tk.LEFT)
        e_vals_var = tk.StringVar(value="0.70, 0.75, 0.80, 0.85")
        ttk.Entry(ectrl, textvariable=e_vals_var, width=28).pack(side=tk.LEFT, padx=(4, 0))
        e_vals_var.trace_add("write", _schedule)

        # ── Figure — bottom items packed first ────────────────────────
        fig = Figure(figsize=(7.5, 4.0), dpi=100, constrained_layout=True)
        ax  = fig.add_subplot(111)
        cv  = FigureCanvasTkAgg(fig, master=win)
        _tb_row = ttk.Frame(win)
        tb  = NavigationToolbar2Tk(cv, _tb_row, pack_toolbar=False)
        tb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tb_row, text="Copy",
                  command=lambda f=fig: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        _passist = attach_plot_assistant(win, fig, ax, cv)
        _passist.pack(side=tk.TOP, fill=tk.X, padx=8, pady=(2, 0))

        res = tk.Text(win, height=6, state=tk.DISABLED, font=("Courier", 9),
                      wrap=tk.WORD)
        _kres_handle = tk.Frame(win, height=6, bg="#888888", cursor="sb_v_double_arrow")
        _tb_row.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        res.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        _kres_handle.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(1, 1))
        cv.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8)

        _kres_y0 = [0]; _kres_h0 = [6]
        def _kres_start(e): _kres_y0[0] = e.y_root; _kres_h0[0] = int(res.cget("height"))
        def _kres_drag(e):
            res.configure(height=max(2, _kres_h0[0] - int((e.y_root - _kres_y0[0]) / 16)))
        _kres_handle.bind("<ButtonPress-1>", _kres_start)
        _kres_handle.bind("<B1-Motion>", _kres_drag)

        # ── Compute ───────────────────────────────────────────────────
        def _compute():
            try:
                D  = float(d_var.get())
                nu = float(nu_var.get())
                C  = float(c_var.get())
            except ValueError:
                return
            try:
                e_vals = [float(x.strip()) for x in e_vals_var.get().split(",")
                          if x.strip()]
            except ValueError:
                return
            if not e_vals:
                return
            F = 96485.0
            B_factor = 0.62 * F * (D ** (2.0/3.0)) * (nu ** (-1.0/6.0)) * C * 1000.0

            # Group selected curves by (sample, catalyst) — one KL fit per group
            sel_by_grp = {}; grp_order = []
            for idx, (E_a, J_a, rpm, lbl, col, sn) in enumerate(all_valid):
                if not _ksel_vars.get(idx, tk.BooleanVar(value=True)).get():
                    continue
                m = re.match(r'^\[([^\]]+)\]', lbl)
                cat = m.group(1) if m else ""
                key = (sn, cat)
                if key not in sel_by_grp:
                    sel_by_grp[key] = []; grp_order.append(key)
                sel_by_grp[key].append((E_a, J_a, rpm, lbl, col))

            ax.clear(); lines = []
            _LS = ['-', '--', '-.', ':']
            _MK = ['o', 's', '^', 'D', 'v', 'P']
            n_ev = len(e_vals)

            for gi, grp_key in enumerate(grp_order):
                sn_g, cat_g = grp_key
                grp_lbl = (f"[{cat_g}] {sn_g}" if cat_g else sn_g)
                grp_ls  = _LS[gi % len(_LS)]
                grp_mk  = _MK[gi % len(_MK)]
                grp_curves = sel_by_grp[grp_key]
                lines.append(f"── {grp_lbl} ──")

                # All e_val lines for this group share the group's base colour
                # (catalyst colour from first curve), distinguished by lightness.
                _grp_base   = grp_curves[0][4] if grp_curves else _PALETTE[gi % len(_PALETTE)]
                _grp_shades = self._gradient_shades(_grp_base, n_ev)

                for ei, e_val in enumerate(e_vals):
                    c_kl = _grp_shades[ei]
                    inv_J = []; inv_sqw = []; rpm_labels = []
                    for E_arr, J_arr, rpm, label, _ in grp_curves:
                        if e_val < E_arr[0] or e_val > E_arr[-1]:
                            continue
                        j_at_e = float(np.interp(e_val, E_arr, J_arr))
                        if j_at_e == 0 or not np.isfinite(j_at_e):
                            continue
                        omega = 2.0 * math.pi * rpm / 60.0
                        inv_J.append(1.0 / j_at_e)
                        inv_sqw.append(1.0 / math.sqrt(omega))
                        rpm_labels.append(re.sub(r'^\[[^\]]+\]\s*', '',label))
                    if len(inv_J) < 2:
                        lines.append(f"  E={e_val:.3f} V: < 2 pts — skipped")
                        continue
                    x = np.array(inv_sqw); y = np.array(inv_J)
                    coeffs = np.polyfit(x, y, 1)
                    slope, intercept = coeffs
                    n   = (1.0 / (abs(slope) * B_factor)) if slope != 0 else float("nan")
                    j_k = (1.0 / intercept) if intercept != 0 else float("nan")
                    ax.scatter(x, y, color=c_kl, zorder=5, s=40, marker=grp_mk)
                    xfit = np.linspace(x.min(), x.max(), 60)
                    ax.plot(xfit, np.polyval(coeffs, xfit), color=c_kl,
                            linewidth=1.2, linestyle=grp_ls,
                            label=f"{grp_lbl}  E={e_val:.3f} V  n={n:.2f}")
                    for xi_pt, yi_pt, rl in zip(x, y, rpm_labels):
                        ax.annotate(rl, (xi_pt, yi_pt), fontsize=7, color=c_kl,
                                    xytext=(3, 3), textcoords="offset points")
                    lines.append(
                        f"  E={e_val:.3f} V:  n = {n:.2f}  Jᵏ = {j_k:+.3f} mA")

            ax.set_xlabel(r"$\omega^{-1/2}$  (rad s$^{-1}$)$^{-1/2}$")
            ax.set_ylabel(r"$|J|^{-1}$  (mA cm$^{-2}$)$^{-1}$")
            ax.set_title("Koutecky-Levich  1/|J| vs 1/√ω")
            ax.legend(fontsize=7, frameon=True)
            cv.draw()
            res.configure(state=tk.NORMAL)
            res.delete("1.0", tk.END)
            res.insert(tk.END, "\n".join(lines))
            res.configure(state=tk.DISABLED)

        _compute()

    def _open_report_window(self):
        """Extract J@E, SA@E, JL for all visible plotted samples — copy-to-Excel TSV."""
        RPMS = [400, 900, 1600, 2500]
        ref = self.ref_electrode_var.get()

        win = tk.Toplevel(self)
        win.title("ORR Report")
        win.geometry("1200x460")
        try: win.state('zoomed')
        except Exception: pass

        ctrl = ttk.Frame(win)
        ctrl.pack(fill=tk.X, padx=8, pady=(6, 2))
        ttk.Label(ctrl, text=f"E value (V vs {ref}):").pack(side=tk.LEFT)
        e_var = tk.StringVar(value="0.90")
        _ev_entry = ttk.Entry(ctrl, textvariable=e_var, width=6)
        _ev_entry.pack(side=tk.LEFT, padx=(2, 10))

        _copy_data = [None]

        def _copy_tsv():
            if _copy_data[0]:
                win.clipboard_clear()
                win.clipboard_append(_copy_data[0])

        def _compute_and_fill():
            try:
                e_tgt = float(e_var.get())
            except ValueError:
                messagebox.showerror("Report", "Invalid E value.", parent=win)
                return

            rows = []
            for sn, sentry in self.samples.items():
                if sentry.get("hidden", False):
                    continue
                if "ax" not in sentry:
                    continue
                cat_corrections = sentry.get("catalyst_corrections", {})
                curves = self._get_curves_for_sample(sn)
                if not curves:
                    continue

                cat_order = []
                curves_by_cat_rpm = {}
                for E_arr, J_arr, rpm, label, color in curves:
                    m = re.match(r'^\[([^\]]+)\]', label)
                    cat = m.group(1) if m else ""
                    if cat not in cat_order:
                        cat_order.append(cat)
                    rpm_r = int(round(rpm))
                    curves_by_cat_rpm[(cat, rpm_r)] = (E_arr, J_arr)

                for cat in cat_order:
                    cc = cat_corrections.get(cat, {})
                    try: area = float(cc.get("area", "") or 0)
                    except ValueError: area = 0.0
                    try: ecsa = float(cc.get("ecsa", "") or 0)
                    except ValueError: ecsa = 0.0

                    row_j = []; row_sa = []; row_jl = []

                    for rpm_t in RPMS:
                        best = None
                        for (c, r), data in curves_by_cat_rpm.items():
                            if c != cat:
                                continue
                            dist = abs(r - rpm_t)
                            if dist <= 50 and (best is None or dist < best[0]):
                                best = (dist, data)

                        if best is None:
                            row_j.append(""); row_sa.append(""); row_jl.append("")
                            continue

                        E_arr, J_arr = best[1]
                        j_lim = float(np.min(J_arr))

                        if e_tgt < E_arr[0] or e_tgt > E_arr[-1]:
                            j_at_e = None
                        else:
                            j_at_e = float(np.interp(e_tgt, E_arr, J_arr))

                        # J at E (mA)
                        if j_at_e is not None:
                            j_ma = j_at_e * area if area > 0 else j_at_e
                            row_j.append(f"{j_ma:.4f}")
                        else:
                            row_j.append("N/A")

                        # JL (mA/cm² if area set, else mA)
                        row_jl.append(f"{j_lim:.4f}")

                        # SA (mA/cm²_ECSA)
                        if (j_at_e is not None and j_lim < 0 and j_at_e < 0
                                and abs(j_lim - j_at_e) > 1e-12 and ecsa > 0):
                            j_k = j_at_e * j_lim / (j_lim - j_at_e)
                            row_sa.append(f"{abs(j_k) / ecsa:.4f}")
                        else:
                            row_sa.append("" if ecsa <= 0 else "N/A")

                    rows.append((sn, cat, row_j, row_sa, row_jl))

            e = e_tgt
            col_hdrs = (
                ["Sample", "Catalyst"]
                + [f"I at {e:.2f}V ({r} rpm) (mA)"         for r in RPMS]
                + [f"SA at {e:.2f}V ({r} rpm) (mA/cm2)"    for r in RPMS]
                + [f"JL ({r} rpm) (mA/cm2)"                  for r in RPMS]
            )
            keys_order = (
                ["Sample", "Catalyst"]
                + [f"J_{r}" for r in RPMS]
                + [f"SA_{r}" for r in RPMS]
                + [f"JL_{r}" for r in RPMS]
            )

            # Build TSV (for Excel copy)
            tsv_lines = ["\t".join(col_hdrs)]
            for sn, cat, row_j, row_sa, row_jl in rows:
                tsv_lines.append("\t".join([sn, cat] + row_j + row_sa + row_jl))
            _copy_data[0] = "\n".join(tsv_lines)

            # Build display (aligned columns)
            col_w = [max(len(h), 8) for h in col_hdrs]
            disp_lines = ["  ".join(h.ljust(w) for h, w in zip(col_hdrs, col_w))]
            disp_lines.append("-" * sum(w + 2 for w in col_w))
            for sn, cat, row_j, row_sa, row_jl in rows:
                vals = [sn, cat] + row_j + row_sa + row_jl
                disp_lines.append("  ".join(v.ljust(w) for v, w in zip(vals, col_w)))
            if not rows:
                disp_lines.append("(No visible plotted samples with data)")

            txt.configure(state=tk.NORMAL)
            txt.delete("1.0", tk.END)
            txt.insert(tk.END, "\n".join(disp_lines))
            txt.configure(state=tk.DISABLED)

        ttk.Button(ctrl, text="Compute", command=_compute_and_fill).pack(
            side=tk.LEFT, padx=(0, 6))
        ttk.Button(ctrl, text="Copy TSV (→ Excel)", command=_copy_tsv).pack(
            side=tk.LEFT, padx=(0, 6))
        ttk.Label(ctrl, text="(visible plotted samples only; ECSA_Hupd required for SA)",
                  foreground="gray", font=("", 8)).pack(side=tk.LEFT, padx=(6, 0))
        _ev_entry.bind("<Return>", lambda e: _compute_and_fill())

        txt_fr = ttk.Frame(win)
        txt_fr.pack(fill=tk.BOTH, expand=True, padx=8, pady=(2, 8))
        txt = tk.Text(txt_fr, font=("Courier", 9), wrap=tk.NONE, state=tk.DISABLED)
        sb_y = ttk.Scrollbar(txt_fr, orient=tk.VERTICAL,   command=txt.yview)
        sb_x = ttk.Scrollbar(txt_fr, orient=tk.HORIZONTAL, command=txt.xview)
        txt.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x.pack(side=tk.BOTTOM, fill=tk.X)
        txt.pack(fill=tk.BOTH, expand=True)

        _compute_and_fill()

    def _open_sa_window(self):
        """ECSA-normalised specific activity — all samples, sample-grouped selector."""
        all_curves = []  # (E_arr, J_arr, rpm, label, color, sname)
        for sn in self.samples:
            for E, J, rpm, lbl, col in self._get_curves_for_sample(sn):
                all_curves.append((E, J, rpm, lbl, col, sn))
        if not all_curves:
            messagebox.showwarning("SA Analysis", "No processed curves in any sample.")
            return
        all_curves = self._gradient_recolor(all_curves)
        ref = self.ref_electrode_var.get()

        # Collect unique catalysts across all samples
        _all_cats = []
        for _, _, _, lbl, _, _ in all_curves:
            m = re.match(r'^\[([^\]]+)\]', lbl)
            cat = m.group(1) if m else ""
            if cat not in _all_cats:
                _all_cats.append(cat)
        _cat_colors = {}
        for ci, cat in enumerate(_all_cats):
            for _, _, _, lbl, col, _ in all_curves:
                m = re.match(r'^\[([^\]]+)\]', lbl)
                if (m.group(1) if m else "") == cat:
                    _cat_colors[cat] = col; break
            if cat not in _cat_colors:
                _cat_colors[cat] = _PALETTE[ci % len(_PALETTE)]

        win = tk.Toplevel(self)
        win.title("Specific Activity (SA)")
        win.geometry("880x760")
        try: win.state('zoomed')
        except Exception: pass

        # Debounce
        _recompute_id = [None]
        def _schedule(*_):
            if _recompute_id[0]:
                try: win.after_cancel(_recompute_id[0])
                except Exception: pass
            _recompute_id[0] = win.after(350, _compute)

        # ── Theory ──────────────────────────────────────────────────────
        _th = ttk.LabelFrame(win, text="Specific Activity Theory")
        _th.pack(fill=tk.X, padx=8, pady=(6, 0))
        _th_g = ttk.Frame(_th); _th_g.pack(fill=tk.X, padx=6, pady=3)
        for _r, (_hd, _bd) in enumerate([
            ("Kinetic current:",  "Jᵏ = J · J_lim / (J_lim − J)   (Koutecky correction)"),
            ("Specific activity:","SA = |Jᵏ| / ECSA   [mA cm⁻²_ECSA]"),
        ]):
            ttk.Label(_th_g, text=_hd, font=("TkDefaultFont", 9, "bold"),
                      anchor=tk.W).grid(row=_r, column=0, sticky=tk.W, pady=2)
            ttk.Label(_th_g, text=f"  {_bd}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=1, sticky=tk.W, padx=(4,0), pady=2)

        # ── Notation — 2-column grid ──────────────────────────────────
        _sym = ttk.LabelFrame(win, text="Notation")
        _sym.pack(fill=tk.X, padx=8, pady=(0, 2))
        _sym_g = ttk.Frame(_sym); _sym_g.pack(fill=tk.X, padx=4, pady=3)
        for _i, (_s, _d) in enumerate([
            ("SA",    "specific activity = |Jᵏ| / ECSA  (mA cm⁻²_ECSA)"),
            ("ECSA",  "electrochemically active surface area (cm²)"),
            ("Jᵏ",   "kinetic J = J · J_lim / (J_lim − J)"),
            ("J_lim", "diffusion-limited plateau (most negative J)"),
            ("J",     "current density (mA cm⁻²)"),
        ]):
            _r, _c = _i // 2, (_i % 2) * 2
            ttk.Label(_sym_g, text=_s, font=("TkDefaultFont", 9, "bold"),
                      width=7, anchor=tk.E).grid(row=_r, column=_c, sticky=tk.E, padx=(8,2), pady=2)
            ttk.Label(_sym_g, text=f"= {_d}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=_c+1, sticky=tk.W, padx=(0,16), pady=2)

        # ── ECSA inputs per unique catalyst ───────────────────────────
        _stored_ecsa = {}
        for _sn_ec, _se_ec in self.samples.items():
            for _cid, _cc_ec in _se_ec.get("catalyst_corrections", {}).items():
                _ev = _cc_ec.get("ecsa", "")
                if _ev and _cid not in _stored_ecsa:
                    _stored_ecsa[_cid] = _ev
        _ecsa_fr = ttk.LabelFrame(win, text="ECSA (cm²) per catalyst")
        _ecsa_fr.pack(fill=tk.X, padx=8, pady=(2, 0))
        _ecsa_row = ttk.Frame(_ecsa_fr); _ecsa_row.pack(fill=tk.X, padx=6, pady=3)
        _ecsa_vars = {}
        for cat in _all_cats:
            ttk.Label(_ecsa_row, text=f"[{cat}]:" if cat else "(no cat):").pack(
                side=tk.LEFT, padx=(0, 2))
            _ecv = tk.StringVar(value=_stored_ecsa.get(cat, ""))
            _ecsa_vars[cat] = _ecv
            ttk.Entry(_ecsa_row, textvariable=_ecv, width=8).pack(
                side=tk.LEFT, padx=(0, 12))

        # ── Curve selector — Group label → Sample [cat] checkbox → RPM ──
        _ksel_fr = ttk.LabelFrame(
            win, text="Select curves  (☑ [Cat] = one sample = one catalyst at multiple RPMs)")
        _ksel_fr.pack(fill=tk.X, padx=8, pady=(4, 0))
        _sel_cv = tk.Canvas(_ksel_fr, height=80, bd=0, highlightthickness=0)
        _sel_sb = ttk.Scrollbar(_ksel_fr, orient=tk.VERTICAL, command=_sel_cv.yview)
        _sel_inner = tk.Frame(_sel_cv)
        _sel_inner.bind("<Configure>",
                        lambda e: _sel_cv.configure(scrollregion=_sel_cv.bbox("all")))
        _sel_cv.create_window((0, 0), window=_sel_inner, anchor="nw")
        _sel_cv.configure(yscrollcommand=_sel_sb.set)
        _sel_sb.pack(side=tk.RIGHT, fill=tk.Y)
        _sel_cv.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        _sel_cv.bind("<MouseWheel>",
                     lambda e: _sel_cv.yview_scroll(-1*(e.delta//120), "units"))

        _by_samp = {}; _samp_order = []
        for idx, (E_a, J_a, rpm, lbl, col, sn) in enumerate(all_curves):
            if sn not in _by_samp:
                _by_samp[sn] = []; _samp_order.append(sn)
            m = re.match(r'^\[([^\]]+)\]', lbl)
            cat = m.group(1) if m else ""
            _by_samp[sn].append((idx, E_a, J_a, rpm, lbl, col, cat))

        _ssel_vars = {}  # idx → BooleanVar

        for sn in _samp_order:
            tk.Label(_sel_inner, text=f"  ▸ Group: {sn}",
                     font=("TkDefaultFont", 8, "italic"), fg="#555555",
                     anchor=tk.W).pack(fill=tk.X, anchor=tk.W, pady=(4, 0))
            _by_cat = {}; _cat_ord = []
            for idx, E_a, J_a, rpm, lbl, col, cat in _by_samp[sn]:
                if cat not in _by_cat:
                    _by_cat[cat] = []; _cat_ord.append(cat)
                _by_cat[cat].append((idx, rpm, lbl, col))
            for cat in _cat_ord:
                cat_idxs = [it[0] for it in _by_cat[cat]]
                cat_bv = tk.BooleanVar(value=True)
                cat_row = tk.Frame(_sel_inner)
                cat_row.pack(fill=tk.X, anchor=tk.W, padx=(18, 0), pady=1)
                tk.Checkbutton(
                    cat_row,
                    text=f"[{cat}]" if cat else "(no cat)",
                    variable=cat_bv,
                    command=lambda idxs=cat_idxs, bv=cat_bv:
                        [_ssel_vars[i].set(bv.get()) for i in idxs],
                    font=("TkDefaultFont", 8, "bold")
                ).pack(side=tk.LEFT, padx=(0, 8))
                for idx, rpm, lbl, col in _by_cat[cat]:
                    bv = tk.BooleanVar(value=True)
                    _ssel_vars[idx] = bv
                    display = re.sub(r'^\[[^\]]+\]\s*', '',lbl)
                    tk.Checkbutton(cat_row, text=display, variable=bv).pack(
                        side=tk.LEFT, padx=3)

        # ── E-value controls ─────────────────────────────────────────
        ectrl = ttk.Frame(win); ectrl.pack(fill=tk.X, padx=8, pady=(4, 0))
        ttk.Label(ectrl, text="E values  (V vs RHE, comma-sep):").pack(side=tk.LEFT)
        e_vals_var = tk.StringVar(value="0.80, 0.85, 0.90")
        ttk.Entry(ectrl, textvariable=e_vals_var, width=26).pack(side=tk.LEFT, padx=(4, 10))
        ttk.Button(ectrl, text="Compute SA", command=lambda: _compute()).pack(side=tk.LEFT)

        # ── Figure — pack bottom items first ──────────────────────────
        fig = Figure(figsize=(7.5, 3.8), dpi=100, constrained_layout=True)
        ax  = fig.add_subplot(111)
        cv  = FigureCanvasTkAgg(fig, master=win)
        _tb_row = ttk.Frame(win)
        tb  = NavigationToolbar2Tk(cv, _tb_row, pack_toolbar=False)
        tb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tb_row, text="Copy",
                  command=lambda f=fig: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        _passist = attach_plot_assistant(win, fig, ax, cv)
        _passist.pack(side=tk.TOP, fill=tk.X, padx=8, pady=(2, 0))
        res = tk.Text(win, height=7, state=tk.DISABLED, font=("Courier", 9), wrap=tk.WORD)
        _sres_handle = tk.Frame(win, height=6, bg="#888888", cursor="sb_v_double_arrow")
        _tb_row.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        res.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        _sres_handle.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(1, 1))
        cv.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8)
        _sres_y0 = [0]; _sres_h0 = [7]
        def _sres_start(e): _sres_y0[0] = e.y_root; _sres_h0[0] = int(res.cget("height"))
        def _sres_drag(e):
            res.configure(height=max(2, _sres_h0[0] - int((e.y_root - _sres_y0[0]) / 16)))
        _sres_handle.bind("<ButtonPress-1>", _sres_start)
        _sres_handle.bind("<B1-Motion>", _sres_drag)

        def _compute():
            try:
                e_vals = [float(x.strip()) for x in e_vals_var.get().split(",")
                          if x.strip()]
            except ValueError:
                messagebox.showerror("SA", "Invalid E values.", parent=win)
                return
            if not e_vals:
                return
            ecsa_map = {}
            for cat, var in _ecsa_vars.items():
                raw = var.get().strip()
                try:
                    ecsa_map[cat] = float(raw) if raw else None
                except ValueError:
                    messagebox.showerror("SA", f"Invalid ECSA for [{cat}].", parent=win)
                    return

            sel_curves = []
            for idx, (E_arr, J_arr, rpm, lbl, col, sn) in enumerate(all_curves):
                if not _ssel_vars.get(idx, tk.BooleanVar(value=True)).get():
                    continue
                m = re.match(r'^\[([^\]]+)\]', lbl)
                cat = m.group(1) if m else ""
                sel_curves.append((idx, E_arr, J_arr, rpm, lbl, col, cat))

            ax.clear()
            lines = [f"{'Curve':<32}  {'E (V)':<7}  {'J':>10}  {'Jᵏ':>10}  SA"]
            lines.append("-" * 80)
            cat_e_sa = {}  # (cat, e_val) → [sa values]
            e_vals_sorted = sorted(e_vals)

            for _, E_arr, J_arr, rpm, lbl, col, cat in sel_curves:
                j_lim = float(np.min(J_arr))
                for e_val in e_vals:
                    if e_val < E_arr[0] or e_val > E_arr[-1]:
                        lines.append(f"  {lbl:<32}  E={e_val:.3f}: out of range"); continue
                    j_at_e = float(np.interp(e_val, E_arr, J_arr))
                    if j_lim >= 0 or j_at_e >= 0 or not np.isfinite(j_at_e):
                        lines.append(f"  {lbl:<32}  E={e_val:.3f}: no cathodic J"); continue
                    if abs(j_lim - j_at_e) < 1e-12:
                        continue
                    j_k = j_at_e * j_lim / (j_lim - j_at_e)
                    ecsa = ecsa_map.get(cat)
                    if ecsa and ecsa > 0:
                        sa = abs(j_k) / ecsa
                        sa_str = f"{sa:.4f}"
                        cat_e_sa.setdefault((cat, e_val), []).append(sa)
                    else:
                        sa_str = "N/A"
                    lines.append(
                        f"  {lbl:<32}  {e_val:.3f}  {j_at_e:>+10.4f}  {j_k:>+10.4f}  {sa_str}")

            if cat_e_sa:
                cats_with_sa = [c for c in _all_cats
                                if any((c, e) in cat_e_sa for e in e_vals)]
                x_pos = np.arange(len(e_vals_sorted))
                bar_w = 0.8 / max(len(cats_with_sa), 1)
                for ci, cat in enumerate(cats_with_sa):
                    sa_means = []; sa_errs = []
                    for e_val in e_vals_sorted:
                        vals = cat_e_sa.get((cat, e_val), [])
                        sa_means.append(float(np.mean(vals)) if vals else 0.0)
                        sa_errs.append(float(np.std(vals)) if len(vals) > 1 else 0.0)
                    offset = (ci - (len(cats_with_sa)-1)/2.0) * bar_w
                    ax.bar(x_pos + offset, sa_means, bar_w * 0.9,
                           yerr=sa_errs if any(s > 0 for s in sa_errs) else None,
                           label=f"[{cat}]" if cat else "(no cat)",
                           color=_cat_colors.get(cat, _PALETTE[ci % len(_PALETTE)]),
                           alpha=0.8, capsize=3)
                ax.set_xticks(x_pos)
                ax.set_xticklabels([f"{e:.3f} V" for e in e_vals_sorted])
                ax.set_ylabel("SA  (mA cm⁻²_ECSA)")
                ax.set_xlabel(f"E  (V vs {ref})")
                ax.set_title("Specific Activity")
                ax.legend(fontsize=8, frameon=True)
            else:
                ax.text(0.5, 0.5, "No SA data (check ECSA inputs and E range)",
                        ha="center", va="center", transform=ax.transAxes)
            cv.draw()
            res.configure(state=tk.NORMAL)
            res.delete("1.0", tk.END)
            res.insert(tk.END, "\n".join(lines))
            res.configure(state=tk.DISABLED)

        _compute()

    def _open_levich_window(self):
        """Standalone Levich plot: |J| vs √RPM — all samples, sample-grouped selector."""
        all_curves = []  # (E_arr, J_arr, rpm, label, color, sname)
        for sn in self.samples:
            for E, J, rpm, lbl, col in self._get_curves_for_sample(sn):
                if rpm > 0:
                    all_curves.append((E, J, rpm, lbl, col, sn))
        if not all_curves:
            messagebox.showwarning("Levich Plot",
                                   "No curves with valid RPM values in any sample.")
            return
        # NOTE: Levich plots one line per (group, e_val), not raw curves —
        # gradient is applied at plot time, not via _gradient_recolor.
        ref = self.ref_electrode_var.get()

        win = tk.Toplevel(self)
        win.title("Levich Plot  |J| vs √RPM")
        win.geometry("820x680")
        try: win.state('zoomed')
        except Exception: pass

        _recompute_id = [None]
        def _schedule(*_):
            if _recompute_id[0]:
                try: win.after_cancel(_recompute_id[0])
                except Exception: pass
            _recompute_id[0] = win.after(350, _compute)

        # ── Theory ───────────────────────────────────────────────────
        _th = ttk.LabelFrame(win, text="Theory")
        _th.pack(fill=tk.X, padx=8, pady=(6, 0))
        _th_g = ttk.Frame(_th); _th_g.pack(fill=tk.X, padx=6, pady=3)
        for _r, (_hd, _bd) in enumerate([
            ("Levich equation:", "|J| = B·√ω   (linear in √RPM at each potential)"),
            ("B factor:",        "B = 0.62·n·F·D^(2/3)·ν^(-1/6)·C   (ω = 2π·RPM/60)"),
        ]):
            ttk.Label(_th_g, text=_hd, font=("TkDefaultFont", 9, "bold"),
                      anchor=tk.W).grid(row=_r, column=0, sticky=tk.W, pady=2)
            ttk.Label(_th_g, text=f"  {_bd}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=1, sticky=tk.W, padx=(4, 0), pady=2)

        # ── Notation ─────────────────────────────────────────────────
        _sym = ttk.LabelFrame(win, text="Notation")
        _sym.pack(fill=tk.X, padx=8, pady=(4, 0))
        _sym_g = ttk.Frame(_sym); _sym_g.pack(fill=tk.X, padx=4, pady=3)
        for _i, (_s, _d) in enumerate([
            ("|J|",  "absolute cathodic current density (mA cm⁻²)"),
            ("√RPM", "square-root of rotation rate (rpm⁰·⁵)"),
            ("ω",    "angular velocity = 2π·RPM/60  (rad s⁻¹)"),
            ("B",    "Levich B factor (slope of |J| vs √RPM)"),
            ("n",    "electrons transferred per O₂"),
            ("F",    "Faraday constant = 96485 C mol⁻¹"),
            ("D",    "O₂ diffusion coefficient (cm² s⁻¹)"),
            ("ν",    "kinematic viscosity (cm² s⁻¹)"),
            ("C",    "O₂ bulk concentration (mol cm⁻³)"),
        ]):
            _r, _c = _i // 2, (_i % 2) * 2
            ttk.Label(_sym_g, text=_s, font=("TkDefaultFont", 9, "bold"),
                      width=7, anchor=tk.E).grid(row=_r, column=_c, sticky=tk.E, padx=(8, 2), pady=2)
            ttk.Label(_sym_g, text=f"= {_d}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=_c+1, sticky=tk.W, padx=(0, 16), pady=2)

        # ── Curve selector ────────────────────────────────────────────
        _sel_fr = ttk.LabelFrame(
            win, text="Select curves  (catalyst [cat] = sample, checkbox toggles all RPMs)")
        _sel_fr.pack(fill=tk.X, padx=8, pady=(4, 0))
        _sel_cv = tk.Canvas(_sel_fr, height=80, bd=0, highlightthickness=0)
        _sel_sb = ttk.Scrollbar(_sel_fr, orient=tk.VERTICAL, command=_sel_cv.yview)
        _sel_inner = tk.Frame(_sel_cv)
        _sel_inner.bind("<Configure>",
                        lambda e: _sel_cv.configure(scrollregion=_sel_cv.bbox("all")))
        _sel_cv.create_window((0, 0), window=_sel_inner, anchor="nw")
        _sel_cv.configure(yscrollcommand=_sel_sb.set)
        _sel_sb.pack(side=tk.RIGHT, fill=tk.Y)
        _sel_cv.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        _sel_cv.bind("<MouseWheel>",
                     lambda e: _sel_cv.yview_scroll(-1*(e.delta//120), "units"))

        _by_samp = {}; _samp_order = []
        for idx, (E_a, J_a, rpm, lbl, col, sn) in enumerate(all_curves):
            if sn not in _by_samp:
                _by_samp[sn] = []; _samp_order.append(sn)
            m = re.match(r'^\[([^\]]+)\]', lbl)
            cat = m.group(1) if m else ""
            _by_samp[sn].append((idx, E_a, J_a, rpm, lbl, col, cat))

        _lsel_vars = {}

        for sn in _samp_order:
            # Group label — non-interactive (user's "group")
            tk.Label(_sel_inner, text=f"  ▸ Group: {sn}",
                     font=("TkDefaultFont", 8, "italic"), fg="#555555",
                     anchor=tk.W).pack(fill=tk.X, anchor=tk.W, pady=(4, 0))
            _by_cat = {}; _cat_ord = []
            for idx, E_a, J_a, rpm, lbl, col, cat in _by_samp[sn]:
                if cat not in _by_cat:
                    _by_cat[cat] = []; _cat_ord.append(cat)
                _by_cat[cat].append((idx, rpm, lbl, col))
            for cat in _cat_ord:
                cat_idxs = [it[0] for it in _by_cat[cat]]
                cat_bv = tk.BooleanVar(value=True)
                cat_row = tk.Frame(_sel_inner)
                cat_row.pack(fill=tk.X, anchor=tk.W, padx=(18, 0), pady=1)
                # Catalyst [cat] = one sample (user's definition) — bold checkbox
                tk.Checkbutton(
                    cat_row,
                    text=f"[{cat}]" if cat else "(no cat)",
                    variable=cat_bv,
                    command=lambda idxs=cat_idxs, bv=cat_bv:
                        [_lsel_vars[i].set(bv.get()) for i in idxs],
                    font=("TkDefaultFont", 8, "bold")
                ).pack(side=tk.LEFT, padx=(0, 8))
                for idx, rpm, lbl, col in _by_cat[cat]:
                    bv = tk.BooleanVar(value=True)
                    _lsel_vars[idx] = bv
                    display = re.sub(r'^\[[^\]]+\]\s*', '',lbl)
                    tk.Checkbutton(cat_row, text=display, variable=bv).pack(
                        side=tk.LEFT, padx=3)
                    bv.trace_add("write", _schedule)

        # ── E-value controls ──────────────────────────────────────────
        ectrl = ttk.Frame(win); ectrl.pack(fill=tk.X, padx=8, pady=(4, 0))
        ttk.Label(ectrl, text="E values (V vs RHE, comma-sep):").pack(side=tk.LEFT)
        e_vals_var = tk.StringVar(value="0.40, 0.60, 0.70, 0.80, 0.90")
        ttk.Entry(ectrl, textvariable=e_vals_var, width=30).pack(
            side=tk.LEFT, padx=(4, 10))
        e_vals_var.trace_add("write", _schedule)

        # ── Figure — bottom first ─────────────────────────────────────
        _LS = ['-', '--', '-.', ':']
        _MK = ['o', 's', '^', 'D', 'v', 'P']
        fig = Figure(figsize=(7.5, 4.2), dpi=100, constrained_layout=True)
        ax  = fig.add_subplot(111)
        cv  = FigureCanvasTkAgg(fig, master=win)
        _tb_row = ttk.Frame(win)
        tb  = NavigationToolbar2Tk(cv, _tb_row, pack_toolbar=False)
        tb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tb_row, text="Copy",
                  command=lambda f=fig: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        _passist = attach_plot_assistant(win, fig, ax, cv)
        _passist.pack(side=tk.TOP, fill=tk.X, padx=8, pady=(2, 0))
        res = tk.Text(win, height=5, state=tk.DISABLED, font=("Courier", 9), wrap=tk.WORD)
        _handle = tk.Frame(win, height=5, bg="#aaaaaa", cursor="sb_v_double_arrow")
        _tb_row.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        res.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        _handle.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(1, 1))
        cv.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8)
        _y0 = [0]; _h0 = [5]
        def _start(e): _y0[0] = e.y_root; _h0[0] = int(res.cget("height"))
        def _drag(e):
            res.configure(height=max(2, _h0[0] - int((e.y_root - _y0[0]) / 16)))
        _handle.bind("<ButtonPress-1>", _start)
        _handle.bind("<B1-Motion>", _drag)

        def _compute():
            try:
                e_vals = [float(x.strip()) for x in e_vals_var.get().split(",")
                          if x.strip()]
            except ValueError:
                return
            if not e_vals: return

            # Group selected curves by (sample, catalyst)
            sel_by_grp = {}; grp_order = []
            for idx, (E_a, J_a, rpm, lbl, col, sn) in enumerate(all_curves):
                if not _lsel_vars.get(idx, tk.BooleanVar(value=True)).get():
                    continue
                m = re.match(r'^\[([^\]]+)\]', lbl)
                cat = m.group(1) if m else ""
                key = (sn, cat)
                if key not in sel_by_grp:
                    sel_by_grp[key] = []; grp_order.append(key)
                sel_by_grp[key].append((E_a, J_a, rpm, lbl, col))

            ax.clear(); lines = []
            n_ev = len(e_vals)

            for gi, grp_key in enumerate(grp_order):
                sn_g, cat_g = grp_key
                grp_lbl = (f"[{cat_g}] {sn_g}" if cat_g else sn_g)
                ls = _LS[gi % len(_LS)]; mk = _MK[gi % len(_MK)]
                grp_curves = sel_by_grp[grp_key]
                lines.append(f"── {grp_lbl} ──")

                # Group base colour (catalyst color from first curve), gradient by e_val
                _grp_base   = grp_curves[0][4] if grp_curves else _PALETTE[gi % len(_PALETTE)]
                _grp_shades = self._gradient_shades(_grp_base, n_ev)

                for ei, e_val in enumerate(e_vals):
                    c_lev = _grp_shades[ei]
                    pts = []
                    for E_arr, J_arr, rpm, _, _ in grp_curves:
                        if e_val < E_arr[0] or e_val > E_arr[-1]: continue
                        j = float(np.interp(e_val, E_arr, J_arr))
                        if np.isfinite(j): pts.append((rpm, abs(j)))
                    pts.sort(key=lambda x: x[0])
                    if not pts:
                        lines.append(f"  E={e_val:.3f} V: no pts"); continue
                    x = [math.sqrt(r) for r, _ in pts]
                    y = [j for _, j in pts]
                    ax.plot(x, y, mk + ls, color=c_lev, linewidth=1.4, markersize=5,
                            label=f"{grp_lbl}  E={e_val:.3f} V")
                    if len(x) >= 2:
                        coeffs = np.polyfit(x, y, 1)
                        lines.append(
                            f"  E={e_val:.3f} V:  slope={coeffs[0]:.4f}  "
                            f"intercept={coeffs[1]:.4f}")
                    else:
                        lines.append(f"  E={e_val:.3f} V:  {len(x)} pt(s)")

            ax.set_xlabel(r"$\sqrt{RPM}$  (rpm$^{0.5}$)")
            ax.set_ylabel(r"$|J|$  (mA cm$^{-2}$)")
            ax.set_title("Levich  |J| vs √RPM")
            ax.legend(fontsize=7, frameon=True)
            cv.draw()
            res.configure(state=tk.NORMAL)
            res.delete("1.0", tk.END)
            res.insert(tk.END, "\n".join(lines))
            res.configure(state=tk.DISABLED)

        _compute()

    # ════════════════════════════════════════════════════════════════
    # Limiting Current Comparison
    # ════════════════════════════════════════════════════════════════
    def _open_lc_comparison_window(self):
        """Plot |J_lim| vs √RPM for each sample/catalyst with Levich theory baseline."""
        all_valid = []   # (sname, label, color, rpm, sqrt_rpm, j_lim)
        for sn in self.samples:
            for E, J, rpm, lbl, col in self._get_curves_for_sample(sn):
                if rpm > 0 and len(J) > 0:
                    j_lim    = float(np.min(J))
                    sqrt_rpm = math.sqrt(rpm)
                    all_valid.append((sn, lbl, col, rpm, sqrt_rpm, j_lim))

        if len(all_valid) < 2:
            messagebox.showwarning(
                "Lim. Current Compare",
                "Need at least 2 RPM curves with numeric RPM values.")
            return
        # Tuple shape here is (sn, lbl, col, rpm, sqrt_rpm, j_lim) — adjust indices
        all_valid = self._gradient_recolor(
            all_valid, label_idx=1, color_idx=2, sort_idx=3, sn_idx=0)

        win = tk.Toplevel(self)
        win.title("Limiting Current Comparison  |J_lim| vs √RPM")
        win.geometry("860x720")
        try: win.state('zoomed')
        except Exception: pass

        _recompute_id = [None]
        def _schedule(*_):
            if _recompute_id[0]:
                try: win.after_cancel(_recompute_id[0])
                except Exception: pass
            _recompute_id[0] = win.after(300, _compute)

        # ── Theory section ──────────────────────────────────────────
        _th = ttk.LabelFrame(win, text="Levich Theory Baseline")
        _th.pack(fill=tk.X, padx=8, pady=(6, 0))
        _th_g = ttk.Frame(_th); _th_g.pack(fill=tk.X, padx=6, pady=3)
        for _r, (_hd, _bd) in enumerate([
            ("Levich equation:",
             "|J_lim| = B·√ω = B·√(2π/60)·√RPM   where B = 0.62·n·F·D^(2/3)·ν^(-1/6)·C"),
        ]):
            ttk.Label(_th_g, text=_hd, font=("TkDefaultFont", 9, "bold"),
                      anchor=tk.W).grid(row=_r, column=0, sticky=tk.W, pady=2)
            ttk.Label(_th_g, text=f"  {_bd}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=1, sticky=tk.W, padx=(4, 0), pady=2)

        # ── Electrolyte selector + n + area ─────────────────────────
        _ec_fr = ttk.Frame(win); _ec_fr.pack(fill=tk.X, padx=8, pady=(4, 0))
        ttk.Label(_ec_fr, text="Electrolyte:").pack(side=tk.LEFT)
        _elec_var = tk.StringVar(value=list(_ELECTROLYTES.keys())[0])
        ttk.Combobox(_ec_fr, textvariable=_elec_var,
                     values=list(_ELECTROLYTES.keys()),
                     state="readonly", width=16).pack(side=tk.LEFT, padx=(4, 12))
        _elec_var.trace_add("write", _schedule)
        ttk.Label(_ec_fr, text="n (electrons):").pack(side=tk.LEFT)
        _n_var = tk.StringVar(value="4")
        ttk.Entry(_ec_fr, textvariable=_n_var, width=4).pack(side=tk.LEFT, padx=(4, 12))
        _n_var.trace_add("write", _schedule)
        _show_pred_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(_ec_fr, text="Show Levich predicted",
                        variable=_show_pred_var,
                        command=_schedule).pack(side=tk.LEFT, padx=(0, 8))
        _show_theory_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(_ec_fr, text="Show theory line",
                        variable=_show_theory_var,
                        command=_schedule).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(_ec_fr, text="Copy Plot",
                   command=lambda: _copy_plot()).pack(side=tk.LEFT, padx=(12, 0))

        # ── X-axis range ──────────────────────────────────────────────
        _ax_fr = ttk.Frame(win); _ax_fr.pack(fill=tk.X, padx=8, pady=(2, 0))
        ttk.Label(_ax_fr, text="X (√RPM):").pack(side=tk.LEFT)
        _xmin_var = tk.StringVar(value="")
        _xmax_var = tk.StringVar(value="")
        ttk.Label(_ax_fr, text="min").pack(side=tk.LEFT, padx=(4, 2))
        ttk.Entry(_ax_fr, textvariable=_xmin_var, width=6).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Label(_ax_fr, text="max").pack(side=tk.LEFT, padx=(0, 2))
        ttk.Entry(_ax_fr, textvariable=_xmax_var, width=6).pack(side=tk.LEFT)
        _xmin_var.trace_add("write", _schedule)
        _xmax_var.trace_add("write", _schedule)

        # ── Curve selector ───────────────────────────────────────────
        _sel_lf = ttk.LabelFrame(win, text="Samples / Catalysts")
        _sel_lf.pack(fill=tk.X, padx=8, pady=(4, 0))
        _sel_cv = tk.Canvas(_sel_lf, height=80, bd=0, highlightthickness=0)
        _sel_sb = ttk.Scrollbar(_sel_lf, orient=tk.VERTICAL, command=_sel_cv.yview)
        _sel_inner = tk.Frame(_sel_cv)
        _sel_inner.bind("<Configure>",
                        lambda e: _sel_cv.configure(scrollregion=_sel_cv.bbox("all")))
        _sel_cv.create_window((0, 0), window=_sel_inner, anchor="nw")
        _sel_cv.configure(yscrollcommand=_sel_sb.set)
        _sel_sb.pack(side=tk.RIGHT, fill=tk.Y)
        _sel_cv.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        _sel_cv.bind("<MouseWheel>",
                     lambda e: _sel_cv.yview_scroll(-1*(e.delta//120), "units"))

        # Group by sample → catalysts within sample
        _by_samp = {}; _samp_order = []
        for idx, (sn, lbl, col, rpm, sq, jl) in enumerate(all_valid):
            if sn not in _by_samp: _by_samp[sn] = []; _samp_order.append(sn)
            _by_samp[sn].append(idx)

        _lsel_vars  = {}   # idx → BooleanVar
        _samp_bvars = {}   # sname → BooleanVar (group toggle)

        for sn in _samp_order:
            sn_idxs = _by_samp[sn]
            sn_bv   = tk.BooleanVar(value=True)
            _samp_bvars[sn] = sn_bv
            sn_row = tk.Frame(_sel_inner); sn_row.pack(anchor=tk.W, pady=(2, 0))
            tk.Label(sn_row, text=f"  ▸ {sn}",
                     font=("TkDefaultFont", 8, "italic"), fg="#555555").pack(
                side=tk.LEFT)
            def _toggle_sn(ids=sn_idxs, bv=sn_bv):
                for _i in ids:
                    _lsel_vars[_i].set(bv.get())
                _schedule()
            tk.Checkbutton(sn_row, text="(all)", variable=sn_bv,
                           command=_toggle_sn,
                           font=("TkDefaultFont", 8, "bold")).pack(side=tk.LEFT)
            # Unique catalysts within sample. When label has no "[cat]" prefix
            # (e.g. when filename parsing didn't yield a catalyst), group all
            # such curves into a single "(no cat)" bucket instead of treating
            # each unique label (which includes the RPM) as its own catalyst.
            def _label_cat(lbl):
                m = re.match(r'^\[([^\]]+)\]', lbl)
                return m.group(1) if m else "(no cat)"
            _seen_cat = []
            for idx in sn_idxs:
                _, lbl, col, _, _, _ = all_valid[idx]
                cat = _label_cat(lbl)
                if cat not in _seen_cat:
                    _seen_cat.append(cat)
                    cat_idxs = [i for i in sn_idxs
                                if _label_cat(all_valid[i][1]) == cat]
                    cat_bv = tk.BooleanVar(value=True)
                    cat_row = tk.Frame(_sel_inner); cat_row.pack(anchor=tk.W)
                    for i in cat_idxs:
                        _lsel_vars[i] = tk.BooleanVar(value=True)
                    def _toggle_cat(idxs=cat_idxs, bv=cat_bv):
                        for _i in idxs:
                            _lsel_vars[_i].set(bv.get())
                        _schedule()
                    tk.Checkbutton(cat_row, text=f"    [{cat}]", variable=cat_bv,
                                   font=("TkDefaultFont", 8),
                                   command=_toggle_cat).pack(side=tk.LEFT)

        # Fallback: any idx not yet in _lsel_vars
        for idx in range(len(all_valid)):
            _lsel_vars.setdefault(idx, tk.BooleanVar(value=True))

        # ── Figure — bottom-packed first ─────────────────────────────
        fig = Figure(figsize=(7.5, 3.8), dpi=100, constrained_layout=True)
        ax  = fig.add_subplot(111)
        cv  = FigureCanvasTkAgg(fig, master=win)
        _tb_row = ttk.Frame(win)
        tb  = NavigationToolbar2Tk(cv, _tb_row, pack_toolbar=False)
        tb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tb_row, text="Copy",
                  command=lambda f=fig: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        _passist = attach_plot_assistant(win, fig, ax, cv)
        _passist.pack(side=tk.TOP, fill=tk.X, padx=8, pady=(2, 0))
        res = tk.Text(win, height=5, state=tk.DISABLED, font=("Courier", 9), wrap=tk.WORD)
        _hdl = tk.Frame(win, height=6, bg="#888888", cursor="sb_v_double_arrow")
        _tb_row.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        res.pack(side=tk.BOTTOM, fill=tk.X, padx=8)
        _hdl.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(1, 1))
        cv.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8)
        _lcy0 = [0]; _lch0 = [5]
        def _lc_start(e): _lcy0[0] = e.y_root; _lch0[0] = int(res.cget("height"))
        def _lc_drag(e):
            res.configure(height=max(2, _lch0[0] - int((e.y_root - _lcy0[0]) / 16)))
        _hdl.bind("<ButtonPress-1>", _lc_start)
        _hdl.bind("<B1-Motion>",     _lc_drag)

        # ── Compute ───────────────────────────────────────────────────
        _LS = ['-', '--', '-.', ':']
        _MK = ['o', 's', '^', 'D', 'v', 'P']
        F_const = 96485.0

        def _copy_plot():
            try:
                import io
                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
                buf.seek(0)
                try:
                    import win32clipboard
                    from PIL import Image
                    img = Image.open(buf)
                    out2 = io.BytesIO()
                    img.convert('RGB').save(out2, 'BMP')
                    data = out2.getvalue()[14:]
                    win32clipboard.OpenClipboard()
                    win32clipboard.EmptyClipboard()
                    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
                    win32clipboard.CloseClipboard()
                except ImportError:
                    from tkinter.filedialog import asksaveasfilename
                    path = asksaveasfilename(
                        parent=win, defaultextension='.png',
                        filetypes=[('PNG image', '*.png')])
                    if path:
                        buf.seek(0)
                        with open(path, 'wb') as f:
                            f.write(buf.read())
            except Exception as exc:
                messagebox.showerror("Copy Plot", str(exc), parent=win)

        def _compute():
            ax.clear()
            lines_out = []
            # Data from _get_curves_for_sample is already in mA/cm²
            # (_process_pair divides I_net by area when area > 0)
            y_unit = "mA/cm²"

            # Group selected points by sample→catalyst. Curves whose labels
            # have no "[cat]" prefix are grouped together under "(no cat)" so
            # they form a single line rather than one isolated point per RPM.
            _grp = {}; _grp_order = []
            for idx, (sn, lbl, col, rpm, sq, jl) in enumerate(all_valid):
                if not _lsel_vars.get(idx, tk.BooleanVar(value=True)).get():
                    continue
                m = re.match(r'^\[([^\]]+)\]', lbl)
                cat = m.group(1) if m else "(no cat)"
                key = (sn, cat)
                if key not in _grp: _grp[key] = []; _grp_order.append(key)
                _grp[key].append((rpm, sq, abs(jl), col))

            _palette_idx = 0
            for (sn, cat), pts in [(k, _grp[k]) for k in _grp_order]:
                pts.sort(key=lambda t: t[0])
                sq_arr  = np.array([t[1] for t in pts])
                jl_arr  = np.array([t[2] for t in pts])
                col     = pts[0][3]
                mk      = _MK[_palette_idx % len(_MK)]
                _palette_idx += 1
                lbl_txt = f"[{cat}] {sn}" if len(_grp) > 1 else f"[{cat}]"
                # Data: solid line
                ax.plot(sq_arr, jl_arr, marker=mk, ls='-', color=col,
                        lw=1.6, ms=5, label=lbl_txt)
                # Levich predicted from lowest RPM — dashed
                _slope_pred = None
                if _show_pred_var.get() and pts[0][1] > 0:
                    _slope_pred = pts[0][2] / pts[0][1]   # J_0 / √RPM_0
                    x_pred = np.linspace(0, sq_arr.max() * 1.1, 60)
                    ax.plot(x_pred, _slope_pred * x_pred, ls="--", color=col,
                            lw=0.9, alpha=0.65, label="_nolegend_")
                # Multi-line block per group (header + indented rpm rows + fit summary)
                lines_out.append(f"── [{cat}] {sn} ──")
                for t in pts:
                    lines_out.append(
                        f"  {int(t[0]):>5d} rpm   |J_lim| = {t[2]:.4f}  {y_unit}")
                if len(pts) >= 2:
                    _coeffs = np.polyfit(sq_arr, jl_arr, 1)
                    lines_out.append(
                        f"  Linear fit:   slope = {_coeffs[0]:.5f}  "
                        f"intercept = {_coeffs[1]:+.5f}  {y_unit}")
                if _slope_pred is not None:
                    lines_out.append(
                        f"  Pred. slope (from lowest RPM) = {_slope_pred:.5f}  "
                        f"{y_unit}/rpm^0.5")
                lines_out.append("")   # blank separator between groups

            # Theoretical baseline — always in mA/cm² (B × √ω)
            if _show_theory_var.get():
                try:
                    elec = _ELECTROLYTES.get(_elec_var.get(), list(_ELECTROLYTES.values())[0])
                    n, D, nu, C = elec
                    n  = float(_n_var.get()) if _n_var.get().strip() else n
                    # B in mA/(cm²·(rad/s)^0.5) — pure current density, no area factor
                    # B in mA/(cm²·(rad/s)^0.5); convert to mA/(cm²·rpm^0.5)
                    # ω = 2π/60 · RPM  →  √ω = √(2π/60) · √RPM
                    B_rad  = 0.62 * n * F_const * (D ** (2/3)) * (nu ** (-1/6)) * C * 1000.0
                    _conv  = math.sqrt(2 * math.pi / 60.0)   # (rad/s)^0.5 per rpm^0.5
                    B_rpm  = B_rad * _conv                    # mA/(cm²·rpm^0.5)
                    sq_max = max((pts[-1][1] for pts in _grp.values()), default=50.0)
                    x_th   = np.linspace(0, sq_max * 1.1, 80)
                    # Theory always in mA/cm² regardless of data normalisation
                    ax.plot(x_th, B_rpm * x_th, "k--", lw=1.5, alpha=0.7,
                            label=f"Theory ({_elec_var.get()}, n={n:.0f}) [mA/cm²]")
                    lines_out.append(
                        f"Theory B={B_rpm:.5f} mA/cm²/rpm^0.5  "
                        f"[{_elec_var.get()}, n={n:.0f}]")
                except (ValueError, KeyError):
                    pass

            ax.set_xlabel(r"$\sqrt{RPM}$  (rpm$^{0.5}$)", fontsize=9)
            ax.set_ylabel(rf"$|J_{{lim}}|$  ({y_unit})", fontsize=9)
            ax.set_title(r"Limiting Current  $|J_{lim}|$ vs $\sqrt{RPM}$", fontsize=10)
            ax.tick_params(labelsize=8)
            ax.legend(fontsize=7, frameon=True)
            ax.grid(True, alpha=0.3)
            try:
                xlo = float(_xmin_var.get())
                xhi = float(_xmax_var.get())
                if xlo < xhi:
                    ax.set_xlim(xlo, xhi)
            except ValueError:
                pass
            cv.draw_idle()
            res.configure(state=tk.NORMAL)
            res.delete("1.0", tk.END)
            res.insert(tk.END, "\n".join(lines_out) if lines_out
                       else "Select at least one curve above.")
            res.configure(state=tk.DISABLED)

        # ── Click-to-annotate ─────────────────────────────────────────
        _ann_obj = [None]    # current annotation artist
        _plot_pts = []       # [(sq, jl, label), ...] populated by _compute

        def _on_lc_click(event):
            if event.inaxes is not ax or event.xdata is None:
                return
            if not _plot_pts:
                return
            # Find nearest point in display coordinates
            x_click, y_click = event.xdata, event.ydata
            xlim = ax.get_xlim(); ylim = ax.get_ylim()
            x_span = xlim[1] - xlim[0] or 1.0
            y_span = ylim[1] - ylim[0] or 1.0
            best_dist, best_pt = float("inf"), None
            for (sq, jl, lbl) in _plot_pts:
                dx = (sq - x_click) / x_span
                dy = (jl - y_click) / y_span
                d = dx*dx + dy*dy
                if d < best_dist:
                    best_dist = d
                    best_pt = (sq, jl, lbl)
            if best_pt is None:
                return
            sq, jl, lbl = best_pt
            rpm_val = sq * sq
            # Remove old annotation
            if _ann_obj[0] is not None:
                try: _ann_obj[0].remove()
                except Exception: pass
                _ann_obj[0] = None
            txt = f"{lbl}\nRPM = {rpm_val:.0f}\n√RPM = {sq:.2f}\n|J| = {jl:.4f} mA/cm²"
            _ann_obj[0] = ax.annotate(
                txt, xy=(sq, jl), xytext=(12, 12), textcoords="offset points",
                bbox=dict(boxstyle="round,pad=0.4", fc="lightyellow", ec="#aaaaaa", alpha=0.9),
                arrowprops=dict(arrowstyle="->", color="#555555", lw=0.8),
                fontsize=8, zorder=10)
            _ann_obj[0].set_in_layout(False)
            cv.draw_idle()

        def _on_lc_rclick(event):
            if event.button == 3 and _ann_obj[0] is not None:
                try: _ann_obj[0].remove()
                except Exception: pass
                _ann_obj[0] = None
                cv.draw_idle()

        cv.mpl_connect("button_press_event", _on_lc_click)
        cv.mpl_connect("button_press_event", _on_lc_rclick)

        # Patch _compute to also fill _plot_pts
        _orig_compute = _compute
        def _compute():
            _plot_pts.clear()
            if _ann_obj[0] is not None:
                try: _ann_obj[0].remove()
                except Exception: pass
                _ann_obj[0] = None
            _orig_compute()
            # Collect marker points from lines on the axes
            for line in ax.get_lines():
                lbl = line.get_label()
                if lbl.startswith("_"):
                    continue
                xd = line.get_xdata()
                yd = line.get_ydata()
                mk = line.get_marker()
                if mk in (None, "None", "") or len(xd) == 0:
                    continue
                for xi, yi in zip(xd, yd):
                    _plot_pts.append((float(xi), float(yi), lbl))

        _compute()

    def _open_comparison_window(self):
        """Sample Comparison: Ratio / Levich / KL / Kinetic for all catalysts."""
        curves = self._get_active_curves()
        if not curves:
            messagebox.showwarning("Sample Comparison",
                                   "No processed curves for active sample.")
            return

        _by_cat = {}; _cat_order = []
        for idx, (E_arr, J_arr, rpm, lbl, col) in enumerate(curves):
            m = re.match(r'^\[([^\]]+)\]', lbl)
            cat = m.group(1) if m else ""
            if cat not in _by_cat:
                _by_cat[cat] = []; _cat_order.append(cat)
            _by_cat[cat].append((idx, E_arr, J_arr, rpm, lbl, col))

        sname = self.active_sample
        ref   = self.ref_electrode_var.get()
        _LS = ['-', '--', '-.', ':']
        _MK = ['o', 's', '^', 'D', 'v', 'P']
        # Use catalyst's registered color (from catalyst_styles, exposed via
        # _get_curves_for_sample). Fall back to _PALETTE if missing.
        cat_colors = {}
        for ci, cat in enumerate(_cat_order):
            items = _by_cat.get(cat, [])
            # Each item tuple is (idx, E_arr, J_arr, rpm, lbl, col) — col at idx 5
            cat_colors[cat] = (items[0][5] if items and items[0][5]
                               else _PALETTE[ci % len(_PALETTE)])

        win = tk.Toplevel(self)
        win.title(f"Sample Comparison — {sname}")
        win.geometry("940x720")
        try: win.state('zoomed')
        except Exception: pass

        # ── Global controls (E values only) ──────────────────────────
        _ctrl = ttk.Frame(win); _ctrl.pack(fill=tk.X, padx=8, pady=(6, 0))
        ttk.Label(_ctrl, text="E values (V vs RHE, comma-sep):").pack(side=tk.LEFT)
        e_vals_var = tk.StringVar(value="0.40, 0.60, 0.70, 0.80, 0.85, 0.90")
        ttk.Entry(_ctrl, textvariable=e_vals_var, width=34).pack(side=tk.LEFT, padx=(4, 10))
        ttk.Button(_ctrl, text="Plot All", command=lambda: _plot_all()).pack(side=tk.LEFT)

        # ── Notation ─────────────────────────────────────────────────
        _sym = ttk.LabelFrame(win, text="Notation")
        _sym.pack(fill=tk.X, padx=8, pady=(2, 0))
        _sym_g = ttk.Frame(_sym); _sym_g.pack(fill=tk.X, padx=4, pady=3)
        for _i, (_s, _d) in enumerate([
            ("|J|",   "absolute cathodic current density (mA cm⁻²)"),
            ("Jᵏ",    "kinetic J = J·J_lim / (J_lim − J)"),
            ("J_lim", "diffusion plateau current (mA cm⁻²)"),
            ("SA",    "|Jᵏ| / ECSA  (mA cm⁻²_ECSA)"),
            ("ECSA",  "electrochemical active surface area (cm²)"),
            ("ω",     "angular velocity = 2π·RPM/60  (rad s⁻¹)"),
        ]):
            _r, _c = _i // 2, (_i % 2) * 2
            ttk.Label(_sym_g, text=_s, font=("TkDefaultFont", 9, "bold"),
                      width=7, anchor=tk.E).grid(row=_r, column=_c, sticky=tk.E, padx=(8, 2), pady=2)
            ttk.Label(_sym_g, text=f"= {_d}", font=("TkDefaultFont", 9),
                      anchor=tk.W).grid(row=_r, column=_c+1, sticky=tk.W, padx=(0, 16), pady=2)

        # ── Notebook ─────────────────────────────────────────────────
        nb = ttk.Notebook(win); nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        # ── Helpers ──────────────────────────────────────────────────
        def _rpm_j(cat, e_val):
            out = []
            for _, E_a, J_a, rpm, _, _ in _by_cat.get(cat, []):
                if rpm <= 0 or e_val < E_a[0] or e_val > E_a[-1]: continue
                j = float(np.interp(e_val, E_a, J_a))
                if np.isfinite(j): out.append((rpm, j))
            return sorted(out, key=lambda x: x[0])

        def _best_curve(cat):
            items = _by_cat.get(cat, [])
            if not items: return None, None
            best = max(items, key=lambda x: x[3])
            return best[1], best[2]

        def _get_e_vals():
            try: return [float(x.strip()) for x in e_vals_var.get().split(",") if x.strip()]
            except ValueError: return []

        # ── Tab 1: Ratio — own numerator/denominator controls ─────────
        tab_r = ttk.Frame(nb); nb.add(tab_r, text="Ratio  |Jn|/|Jd|")
        _rctrl = ttk.Frame(tab_r); _rctrl.pack(fill=tk.X, padx=6, pady=(4, 0))
        ttk.Label(_rctrl, text="Numerator:").pack(side=tk.LEFT)
        num_var = tk.StringVar(value=_cat_order[0] if _cat_order else "")
        ttk.Combobox(_rctrl, textvariable=num_var, values=_cat_order,
                     state="readonly", width=16).pack(side=tk.LEFT, padx=(4, 10))
        ttk.Label(_rctrl, text="÷  Denominator:").pack(side=tk.LEFT)
        den_var = tk.StringVar(
            value=_cat_order[1] if len(_cat_order) > 1 else _cat_order[0] if _cat_order else "")
        ttk.Combobox(_rctrl, textvariable=den_var, values=_cat_order,
                     state="readonly", width=16).pack(side=tk.LEFT, padx=(4, 10))
        fig_r = Figure(figsize=(7.5, 4.2), dpi=100, constrained_layout=True)
        ax_r  = fig_r.add_subplot(111)
        cv_r  = FigureCanvasTkAgg(fig_r, master=tab_r)
        _tbr_r = ttk.Frame(tab_r); _tbr_r.pack(side=tk.BOTTOM, fill=tk.X)
        NavigationToolbar2Tk(cv_r, _tbr_r, pack_toolbar=False).pack(
            side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tbr_r, text="Copy",
                  command=lambda f=fig_r: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        attach_plot_assistant(tab_r, fig_r, ax_r, cv_r).pack(
            side=tk.TOP, fill=tk.X, padx=4, pady=(2, 0))
        cv_r.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        num_var.trace_add("write", lambda *_: _plot_ratio())
        den_var.trace_add("write", lambda *_: _plot_ratio())

        # ── Tab 2: Levich ─────────────────────────────────────────────
        tab_l = ttk.Frame(nb); nb.add(tab_l, text="Levich  |J| vs √RPM")
        fig_l = Figure(figsize=(7.5, 4.2), dpi=100, constrained_layout=True)
        ax_l  = fig_l.add_subplot(111)
        cv_l  = FigureCanvasTkAgg(fig_l, master=tab_l)
        _tbr_l = ttk.Frame(tab_l); _tbr_l.pack(side=tk.BOTTOM, fill=tk.X)
        NavigationToolbar2Tk(cv_l, _tbr_l, pack_toolbar=False).pack(
            side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tbr_l, text="Copy",
                  command=lambda f=fig_l: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        attach_plot_assistant(tab_l, fig_l, ax_l, cv_l).pack(
            side=tk.TOP, fill=tk.X, padx=4, pady=(2, 0))
        cv_l.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # ── Tab 3: KL ─────────────────────────────────────────────────
        tab_kl = ttk.Frame(nb); nb.add(tab_kl, text="KL  1/|J| vs 1/√ω")
        fig_kl = Figure(figsize=(7.5, 4.2), dpi=100, constrained_layout=True)
        ax_kl  = fig_kl.add_subplot(111)
        cv_kl  = FigureCanvasTkAgg(fig_kl, master=tab_kl)
        _tbr_kl = ttk.Frame(tab_kl); _tbr_kl.pack(side=tk.BOTTOM, fill=tk.X)
        NavigationToolbar2Tk(cv_kl, _tbr_kl, pack_toolbar=False).pack(
            side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tbr_kl, text="Copy",
                  command=lambda f=fig_kl: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        attach_plot_assistant(tab_kl, fig_kl, ax_kl, cv_kl).pack(
            side=tk.TOP, fill=tk.X, padx=4, pady=(2, 0))
        cv_kl.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # ── Tab 4: Kinetic — ECSA inputs in tab ──────────────────────
        tab_kin = ttk.Frame(nb); nb.add(tab_kin, text="Kinetic  Jᵏ & SA")
        _kctrl = ttk.Frame(tab_kin); _kctrl.pack(fill=tk.X, padx=6, pady=(4, 0))
        ttk.Label(_kctrl, text="ECSA (cm²) per catalyst:").pack(side=tk.LEFT)
        _ecsa_vars = {}
        _stored_ecsa_cmp = {}
        for _sn_ec2, _se_ec2 in self.samples.items():
            for _cid2, _cc_ec2 in _se_ec2.get("catalyst_corrections", {}).items():
                _ev2 = _cc_ec2.get("ecsa", "")
                if _ev2 and _cid2 not in _stored_ecsa_cmp:
                    _stored_ecsa_cmp[_cid2] = _ev2
        for cat in _cat_order:
            ttk.Label(_kctrl, text=f" [{cat}]:").pack(side=tk.LEFT)
            ev = tk.StringVar(value=_stored_ecsa_cmp.get(cat, ""))
            _ecsa_vars[cat] = ev
            ttk.Entry(_kctrl, textvariable=ev, width=7).pack(side=tk.LEFT, padx=(2, 4))
        fig_kin = Figure(figsize=(7.5, 4.2), dpi=100, constrained_layout=True)
        ax_ta   = fig_kin.add_subplot(121)
        ax_sa   = fig_kin.add_subplot(122)
        cv_kin  = FigureCanvasTkAgg(fig_kin, master=tab_kin)
        _tbr_kin = ttk.Frame(tab_kin); _tbr_kin.pack(side=tk.BOTTOM, fill=tk.X)
        NavigationToolbar2Tk(cv_kin, _tbr_kin, pack_toolbar=False).pack(
            side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_tbr_kin, text="Copy",
                  command=lambda f=fig_kin: copy_figure_to_clipboard(f),
                  relief=tk.RAISED, borderwidth=1, padx=6).pack(
                      side=tk.LEFT, padx=(4, 2), pady=1)
        attach_plot_assistant(tab_kin, fig_kin, [ax_ta, ax_sa], cv_kin).pack(
            side=tk.TOP, fill=tk.X, padx=4, pady=(2, 0))
        cv_kin.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # ── Plot Ratio ────────────────────────────────────────────────
        def _plot_ratio():
            e_vals = _get_e_vals()
            if not e_vals: return
            num_cat = num_var.get(); den_cat = den_var.get()
            # Gradient of numerator catalyst's color across E values
            _r_base = cat_colors.get(num_cat, _PALETTE[0])
            e_colors = self._gradient_shades(_r_base, len(e_vals))
            ax_r.clear()
            for e_val, c_e in zip(e_vals, e_colors):
                d_pts = _rpm_j(den_cat, e_val); n_pts = _rpm_j(num_cat, e_val)
                d_d = {r: j for r, j in d_pts}; n_d = {r: j for r, j in n_pts}
                common = sorted(set(d_d) & set(n_d))
                if not common: continue
                ratio = [abs(n_d[r]) / abs(d_d[r]) for r in common
                         if d_d[r] != 0 and np.isfinite(n_d[r]) and np.isfinite(d_d[r])]
                if not ratio: continue
                ax_r.plot(common[:len(ratio)], ratio, 'o-', color=c_e,
                          linewidth=1.5, markersize=5, label=f"E = {e_val:.2f} V")
            ax_r.axhline(1.0, color="gray", lw=0.8, ls="--")
            ax_r.set_xlabel("RPM")
            ax_r.set_ylabel(f"|J[{num_cat}]| / |J[{den_cat}]|")
            ax_r.set_title(f"Current Ratio  [{num_cat}] / [{den_cat}]")
            ax_r.legend(fontsize=7, frameon=True)
            cv_r.draw()

        # ── Plot All (Levich + KL + Kinetic) ─────────────────────────
        def _plot_all():
            e_vals = _get_e_vals()
            if not e_vals:
                messagebox.showerror("Comparison", "Invalid E values.", parent=win)
                return
            lev_ev = e_vals[:5]; kl_ev = e_vals[:4]
            # Per-catalyst gradient shades (E values within a catalyst share base color)
            lev_shades = {cat: self._gradient_shades(
                              cat_colors.get(cat, _PALETTE[ci % len(_PALETTE)]),
                              len(lev_ev))
                          for ci, cat in enumerate(_cat_order)}
            kl_shades  = {cat: self._gradient_shades(
                              cat_colors.get(cat, _PALETTE[ci % len(_PALETTE)]),
                              len(kl_ev))
                          for ci, cat in enumerate(_cat_order)}

            # Levich
            ax_l.clear()
            for ci, cat in enumerate(_cat_order):
                ls = _LS[ci % len(_LS)]; mk = _MK[ci % len(_MK)]
                shades = lev_shades[cat]
                for ei, e_val in enumerate(lev_ev):
                    c_e = shades[ei]
                    pts = _rpm_j(cat, e_val)
                    if not pts: continue
                    x = [math.sqrt(r) for r, _ in pts]
                    y = [abs(j) for _, j in pts]
                    ax_l.plot(x, y, mk + ls, color=c_e, linewidth=1.4, markersize=5,
                              label=f"[{cat}] {e_val:.2f} V")
            ax_l.set_xlabel(r"$\sqrt{RPM}$  (rpm$^{0.5}$)")
            ax_l.set_ylabel(r"$|J|$  (mA cm$^{-2}$)")
            ax_l.set_title("Levich  |J| vs √RPM")
            ax_l.legend(fontsize=6, frameon=True, ncol=max(1, len(_cat_order)))
            cv_l.draw()

            # KL
            ax_kl.clear()
            for ci, cat in enumerate(_cat_order):
                ls = _LS[ci % len(_LS)]; mk = _MK[ci % len(_MK)]
                shades = kl_shades[cat]
                for ei, e_val in enumerate(kl_ev):
                    c_e = shades[ei]
                    pts = _rpm_j(cat, e_val)
                    x = []; y = []
                    for r, j in pts:
                        if j == 0: continue
                        x.append(1.0 / math.sqrt(2.0 * math.pi * r / 60.0))
                        y.append(1.0 / abs(j))
                    if len(x) < 2: continue
                    ax_kl.plot(x, y, mk + ls, color=c_e, linewidth=1.4, markersize=4,
                               label=f"[{cat}] {e_val:.2f} V")
            ax_kl.set_xlabel(r"$\omega^{-1/2}$  (rad s$^{-1}$)$^{-1/2}$")
            ax_kl.set_ylabel(r"$|J|^{-1}$  (mA cm$^{-2}$)$^{-1}$")
            ax_kl.set_title("Koutecky-Levich  1/|J| vs 1/√ω")
            ax_kl.legend(fontsize=6, frameon=True, ncol=max(1, len(_cat_order)))
            cv_kl.draw()

            # Kinetic
            ax_ta.clear(); ax_sa.clear()
            for ci, cat in enumerate(_cat_order):
                E_a, J_a = _best_curve(cat)
                if E_a is None: continue
                j_lim = float(np.min(J_a))
                if j_lim >= 0: continue
                color = cat_colors[cat]; ls = _LS[ci % len(_LS)]
                mask = (E_a >= 0.85) & (E_a <= 0.95)
                if mask.sum() >= 3:
                    E_k = E_a[mask]; J_k = J_a[mask]
                    with np.errstate(divide="ignore", invalid="ignore"):
                        jk = J_k * j_lim / (j_lim - J_k)
                    good = np.isfinite(jk) & (jk < 0)
                    if good.sum() >= 3:
                        with np.errstate(divide="ignore", invalid="ignore"):
                            log_jk = np.log10(np.abs(jk[good]))
                        fin = np.isfinite(log_jk)
                        ax_ta.plot(log_jk[fin], E_k[good][fin], ls,
                                   color=color, linewidth=1.5, label=f"[{cat}]")
            ax_ta.set_xlabel(r"$\log_{10}|J^k|$")
            ax_ta.set_ylabel(f"E  (V vs {ref})")
            ax_ta.set_title("Tafel (Koutecky-corrected, 0.85–0.95 V)")
            ax_ta.legend(fontsize=7)

            sa_e_list = []
            for target in [0.90, 0.85]:
                closest = min(e_vals, key=lambda v: abs(v - target))
                if abs(closest - target) < 0.06 and closest not in sa_e_list:
                    sa_e_list.append(closest)
            if not sa_e_list:
                sa_e_list = e_vals[-2:] if len(e_vals) >= 2 else e_vals
            x_sa = np.arange(len(sa_e_list))
            bw = 0.8 / max(len(_cat_order), 1)
            any_ecsa = False
            for ci, cat in enumerate(_cat_order):
                raw = _ecsa_vars[cat].get().strip()
                try: ecsa = float(raw) if raw else None
                except ValueError: ecsa = None
                if not ecsa or ecsa <= 0: continue
                any_ecsa = True
                E_a, J_a = _best_curve(cat)
                if E_a is None: continue
                j_lim = float(np.min(J_a))
                sa_vals = []
                for e_val in sa_e_list:
                    if e_val < E_a[0] or e_val > E_a[-1] or j_lim >= 0:
                        sa_vals.append(0.0); continue
                    j = float(np.interp(e_val, E_a, J_a))
                    if j >= 0 or abs(j_lim - j) < 1e-12:
                        sa_vals.append(0.0); continue
                    jk = j * j_lim / (j_lim - j)
                    sa_vals.append(abs(jk) / ecsa)
                offset = (ci - (len(_cat_order)-1)/2.0) * bw
                ax_sa.bar(x_sa + offset, sa_vals, bw * 0.85,
                          color=cat_colors[cat], alpha=0.8, label=f"[{cat}]")
            ax_sa.set_xticks(x_sa)
            ax_sa.set_xticklabels([f"{e:.2f} V" for e in sa_e_list])
            ax_sa.set_ylabel(r"SA  (mA cm$^{-2}_{ECSA}$)")
            ax_sa.set_title("Specific Activity")
            ax_sa.legend(fontsize=7)
            if not any_ecsa:
                ax_sa.text(0.5, 0.5, "Enter ECSA above for SA",
                           ha="center", va="center", transform=ax_sa.transAxes,
                           color="gray", fontsize=8)
            cv_kin.draw()
            _plot_ratio()

        _plot_all()

    # ════════════════════════════════════════════════════════════════
    # Excel export
    # ════════════════════════════════════════════════════════════════
    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """Excel sheet names: ≤31 chars, none of []:*?/\\, non-empty."""
        cleaned = re.sub(r'[\[\]:*?/\\]', "_", str(name)).strip()
        return (cleaned[:31] or "Sample")

    def _export_sample_excel(self):
        """Export the ACTIVE sample to a single Excel sheet.

        Only pairs currently enabled/active in the plot are written, using the
        exact same processed curves as the plot (`_build_curve_records`, which
        applies per-catalyst iR, N2-subtraction, RHE conversion and current
        density). RPMs are laid out left→right as (E, J) column pairs.
        """
        if not self.active_sample or self.active_sample not in self.samples:
            messagebox.showwarning("ORR Export", "Select a sample first.")
            return

        records = self._build_curve_records(self.active_sample)
        if not records:
            messagebox.showwarning(
                "ORR Export",
                "No enabled pairs in the active sample could be processed.\n"
                "(Only pairs currently active in the plot are exported.)")
            return

        ref   = self.ref_electrode_var.get() or "RHE"
        e_hdr = f"E (V vs {ref})"

        path = filedialog.asksaveasfilename(
            title="Export ORR Sample to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{self.active_sample}_ORR.xlsx",
            parent=self)
        if not path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = self._sanitize_sheet_name(self.active_sample)

            # ── Metadata / header block (column A = labels) ──────────────
            ws.cell(1, 1, "Sample")
            ws.cell(1, 2, self.active_sample)
            ws.cell(2, 1, "R_sol_N2 (Ω)")
            ws.cell(3, 1, "R_sol_O2 (Ω)")
            ws.cell(4, 1, f"E_ref → {ref} (V)")
            ws.cell(5, 1, "Area (cm²)")
            ws.cell(6, 1, "ECSA_Hupd (cm²)")
            # row 7 blank
            ws.cell(8, 1, "RPM")
            # row 9 = E/J sub-headers ; data from row 10

            DATA_ROW0 = 10
            for idx, rec in enumerate(records):
                c_e = 2 + 2 * idx        # E column for this RPM
                c_j = c_e + 1            # J column for this RPM
                area = rec["area"]
                j_hdr = "J (mA cm⁻²)" if area > 0 else "I_net (mA)"

                ws.cell(2, c_e, rec["r_n2"])
                ws.cell(3, c_e, rec["r_o2"])
                ws.cell(4, c_e, rec["e_ref"])
                ws.cell(5, c_e, area if area > 0 else "")
                _ecsa = rec.get("ecsa", "")
                ws.cell(6, c_e, _ecsa if _ecsa not in (None, "") else "")
                ws.cell(8, c_e, rec["label"])
                ws.cell(9, c_e, e_hdr)
                ws.cell(9, c_j, j_hdr)

                E_arr, J_arr = rec["E"], rec["J"]
                for k in range(len(E_arr)):
                    ws.cell(DATA_ROW0 + k, c_e, float(E_arr[k]))
                    ws.cell(DATA_ROW0 + k, c_j, float(J_arr[k]))

            wb.save(path)
            self._log(f"Exported '{self.active_sample}' "
                      f"({len(records)} RPM) → {os.path.basename(path)}")
            messagebox.showinfo("ORR Export", f"Saved:\n{path}", parent=self)
        except Exception as exc:
            messagebox.showerror("ORR Export", f"Export failed:\n{exc}", parent=self)

    # ════════════════════════════════════════════════════════════════
    # Logging
    # ════════════════════════════════════════════════════════════════
    def _log(self, msg: str):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    # ════════════════════════════════════════════════════════════════
    # Session save / restore
    # ════════════════════════════════════════════════════════════════
    def get_session_state(self, data_store: dict) -> dict:
        self._save_active_sample_state()

        samples_list = []
        for sname, sentry in self.samples.items():
            rec = {"name": sname}
            pairs_rec = []
            for pair in sentry.get("pairs", []):
                pr = {}
                for k, v in pair.items():
                    if k in _PAIR_RUNTIME:
                        continue
                    if k in ("df_n2", "df_o2"):
                        continue
                    try:
                        import json; json.dumps(v)
                        pr[k] = v
                    except (TypeError, ValueError):
                        pass
                # Store DataFrames by hash
                for gas in ("n2", "o2"):
                    df = pair.get(f"df_{gas}")
                    if df is not None:
                        h = _sm.df_hash(df)
                        data_store[h] = df
                        pr[f"df_{gas}_hash"] = h
                pairs_rec.append(pr)
            rec["pairs"] = pairs_rec
            for k, v in sentry.items():
                if k in _SAMPLE_RUNTIME or k == "pairs":
                    continue
                try:
                    import json; json.dumps(v)
                    rec[k] = v
                except (TypeError, ValueError):
                    pass
            rec["reflines"] = [list(r) for r in sentry.get("reflines", [])]
            samples_list.append(rec)

        return {
            "active_sample": self.active_sample,
            "plot_w_var":    self.plot_w_var.get(),
            "plot_h_var":    self.plot_h_var.get(),
            "grid_cols_var": self._grid_cols_var.get(),
            "samples": samples_list,
        }

    def restore_session_state(self, state: dict, data_store: dict) -> None:
        old = self._suppress_replot
        self._suppress_replot = True

        # Destroy existing sample figures
        for sentry in self.samples.values():
            pf = sentry.get("plot_frame")
            if pf is not None:
                try:
                    pf.destroy()
                except Exception:
                    pass
        self.samples.clear()
        self.active_sample = None
        self.sample_lb.clear()
        self._zoom_sample = None

        # Clear loaded files
        self.loaded_files.clear()
        self._loaded_keys.clear()
        children = self.loaded_tv.get_children()
        if children:
            self.loaded_tv.delete(*children)

        # Panel-level vars
        try:
            self.plot_w_var.set(state.get("plot_w_var", "10.5"))
            self.plot_h_var.set(state.get("plot_h_var", "5.5"))
            self._grid_cols_var.set(state.get("grid_cols_var", "2"))
        except Exception:
            pass

        # Restore samples
        for srec in state.get("samples", []):
            sname = srec.get("name", "")
            if not sname:
                continue
            sentry: dict = {"pairs": [], "reflines": []}
            for k, v in srec.items():
                if k in ("name", "pairs"):
                    continue
                sentry[k] = v
            sentry["reflines"] = [tuple(r) for r in sentry.get("reflines", [])]

            # Restore pairs + re-attach DataFrames from data_store
            for prec in srec.get("pairs", []):
                pair = dict(prec)
                for gas in ("n2", "o2"):
                    h = pair.pop(f"df_{gas}_hash", None)
                    pair[f"df_{gas}"] = data_store.get(h) if h else None
                sentry["pairs"].append(pair)

            self.samples[sname] = sentry
            # Migrate old flat correction values to per-catalyst format
            if "catalyst_corrections" not in sentry:
                sentry["catalyst_corrections"] = {}
            old_r_n2 = sentry.get("r_sol_n2", 0.0)
            old_r_o2 = sentry.get("r_sol_o2", 0.0)
            old_eref = sentry.get("e_ref", 0.0)
            old_area = sentry.get("area", "")
            for p in sentry.get("pairs", []):
                c = p.get("catalyst_id", "")
                if c:
                    sentry["catalyst_corrections"].setdefault(
                        c, {"r_sol_n2": old_r_n2, "r_sol_o2": old_r_o2,
                            "e_ref": old_eref, "area": old_area})
            self.sample_lb.insert(tk.END, sname,
                                  checked=not srec.get("hidden", False))
            self._create_sample_figure(sname)

        self._suppress_replot = old
        active_sample = state.get("active_sample")
        if active_sample and active_sample in self.samples:
            keys = list(self.samples.keys())
            self.sample_lb.selection_set(keys.index(active_sample))
            self._switch_active_sample(active_sample)
        elif self.samples:
            first = next(iter(self.samples))
            self.sample_lb.selection_set(0)
            self._switch_active_sample(first)

        self._apply_plot_size()
        self._relayout_figures()
